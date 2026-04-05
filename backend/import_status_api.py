import io
import json
import os
import re
import base64
import binascii
import sqlite3
import time
import xml.etree.ElementTree as ET
from collections import defaultdict, deque
from decimal import Decimal, InvalidOperation
import hashlib
import hmac
import secrets
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Set, Tuple

import mysql.connector
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, Query, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", "3306")),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", "showlang"),
    "database": os.getenv("DB_NAME", "trans_fields_mapping"),
    "connection_timeout": int(os.getenv("DB_CONNECTION_TIMEOUT", "60")),
    "read_timeout": int(os.getenv("DB_READ_TIMEOUT", "3600")),
    "write_timeout": int(os.getenv("DB_WRITE_TIMEOUT", "3600")),
}
DB_DRIVER = str(os.getenv("DB_DRIVER", "mysql")).strip().lower() or "mysql"
SQLITE_DB_PATH = str(os.getenv("SQLITE_DB_PATH", "backend/data/trans_fields_mapping.db")).strip() or "backend/data/trans_fields_mapping.db"
IS_SQLITE = DB_DRIVER == "sqlite"

DB_SSL_CA = str(os.getenv("DB_SSL_CA", "")).strip()
DB_SSL_DISABLED = str(os.getenv("DB_SSL_DISABLED", "false")).strip().lower() == "true"
DB_SSL_VERIFY_CERT = str(os.getenv("DB_SSL_VERIFY_CERT", "false")).strip().lower() == "true"
DB_SSL_VERIFY_IDENTITY = str(os.getenv("DB_SSL_VERIFY_IDENTITY", "false")).strip().lower() == "true"
DB_SSL_CA_PATH = Path(DB_SSL_CA) if DB_SSL_CA else None

if DB_SSL_DISABLED:
    DB_CONFIG["ssl_disabled"] = True
elif DB_SSL_CA_PATH and DB_SSL_CA_PATH.exists():
    DB_CONFIG["ssl_ca"] = DB_SSL_CA
    DB_CONFIG["ssl_verify_cert"] = DB_SSL_VERIFY_CERT
    DB_CONFIG["ssl_verify_identity"] = DB_SSL_VERIFY_IDENTITY

CORS_ALLOW_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "CORS_ALLOW_ORIGINS",
        "http://localhost:8088,http://127.0.0.1:8088,http://localhost:5500,http://127.0.0.1:5500,http://localhost:5173,http://127.0.0.1:5173,http://localhost:3000,http://127.0.0.1:3000,null",
    ).split(",")
    if origin.strip()
]
CORS_ALLOW_ORIGIN_REGEX = str(os.getenv("CORS_ALLOW_ORIGIN_REGEX", "")).strip() or None

AUTH_COOKIE_NAME = os.getenv("AUTH_COOKIE_NAME", "trans_fields_mapping_session")
AUTH_COOKIE_SECURE = os.getenv("AUTH_COOKIE_SECURE", "false").strip().lower() == "true"
AUTH_COOKIE_SAMESITE = str(os.getenv("AUTH_COOKIE_SAMESITE", "lax")).strip().lower()
AUTH_COOKIE_DOMAIN = str(os.getenv("AUTH_COOKIE_DOMAIN", "")).strip() or None
AUTH_SESSION_HOURS = int(os.getenv("AUTH_SESSION_HOURS", "12"))
AUTH_PASSWORD_MIN_LEN = int(os.getenv("AUTH_PASSWORD_MIN_LEN", "8"))
AUTH_LOGIN_MAX_FAILS = int(os.getenv("AUTH_LOGIN_MAX_FAILS", "5"))
AUTH_TEMP_LOCK_MINUTES = int(os.getenv("AUTH_TEMP_LOCK_MINUTES", "15"))
DEFAULT_ADMIN_USERNAME = os.getenv("DEFAULT_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = str(os.getenv("DEFAULT_ADMIN_PASSWORD", "")).strip()

if AUTH_COOKIE_SAMESITE not in {"lax", "strict", "none"}:
    AUTH_COOKIE_SAMESITE = "lax"

PROG_CODE_TABLE = "prog_code"
ALLOWED_TABLES = {"rstran", "rstrant", "rstranrule", "rstranfield", "rstranstepcnst", "rstransteprout", "rsoadso", "rsoadsot", "rsds", "rsdst", "rsdssegfd", "rsdssegfdt", "rsksnew", "rsksnewt", "rsksfieldnew", "rsksfieldnewt", "dd03l", "dd02t", "dd03t", "dd04t", "rsdiobj", "rsdiobjt", "bw_object_name", PROG_CODE_TABLE}
IMPORT_SCHEMA_TABLES = {"rstran", "rstrant", "rstranrule", "rstranfield", "rstranstepcnst", "rstransteprout", "rsoadso", "rsoadsot", "rsds", "rsdst", "rsdssegfd", "rsdssegfdt", "rsksnew", "rsksnewt", "rsksfieldnew", "rsksfieldnewt", "dd03l", "dd02t", "dd03t", "dd04t", "rsdiobj", "rsdiobjt", "bw_object_name", PROG_CODE_TABLE}
RSTRAN_MAPPING_RULE_TABLE = "rstran_mapping_rule"
RSTRAN_MAPPING_RULE_FULL_TABLE = "rstran_mapping_rule_full"
BW_OBJECT_FIELD_INVENTORY_TABLE = "bw_object_field_inventory"
DERIVED_STATUS_TABLES = frozenset({
    RSTRAN_MAPPING_RULE_TABLE,
    RSTRAN_MAPPING_RULE_FULL_TABLE,
    BW_OBJECT_FIELD_INVENTORY_TABLE,
})
IMPORT_STATUS_TRACKED_TABLES = frozenset(set(ALLOWED_TABLES) | set(DERIVED_STATUS_TABLES))
DD03L_METADATA_PATH = Path(os.getenv("DD03L_METADATA_PATH", "/app/Excel Data/P4C-DD03L-Table Fields-RS_DD-0317.xlsx"))
DATA_QUERY_EXCLUDED_TABLES = frozenset({"import_status", "users", "user_sessions"})
DATA_QUERY_ALLOWED_JOIN_TYPES = {
    "inner": "INNER JOIN",
    "left": "LEFT JOIN",
    "right": "RIGHT JOIN",
}
DATA_QUERY_ALLOWED_FILTER_OPERATORS = frozenset({"in", "range"})
DATA_QUERY_PREVIEW_LIMIT = 10
DATA_QUERY_DEFAULT_LIMIT = 200
DATA_QUERY_MAX_LIMIT = 1000
DDIC_METADATA_TABLES = {
    "dd03l": {"metadata_tabname": "DD03L", "comment": "DD03L table field metadata"},
    "dd02t": {"metadata_tabname": "DD02T", "comment": "DD02T table text metadata"},
    "dd03t": {"metadata_tabname": "DD03T", "comment": "DD03T field text metadata"},
    "dd04t": {"metadata_tabname": "DD04T", "comment": "DD04T data element text metadata"},
    "rsdiobj": {"metadata_tabname": "RSDIOBJ", "comment": "InfoObject properties metadata"},
    "rsdiobjt": {"metadata_tabname": "RSDIOBJT", "comment": "InfoObject text metadata"},
    "rsoadso": {"metadata_tabname": "RSOADSO", "comment": "ADSO technical metadata"},
    "rsdssegfd": {"metadata_tabname": "RSDSSEGFD", "comment": "RSDS segment field metadata"},
    "rsdssegfdt": {"metadata_tabname": "RSDSSEGFDT", "comment": "RSDS segment field text metadata"},
    "rsksnew": {"metadata_tabname": "RSKSNEW", "comment": "InfoSource master metadata"},
    "rsksnewt": {"metadata_tabname": "RSKSNEWT", "comment": "InfoSource text metadata"},
    "rsksfieldnew": {"metadata_tabname": "RSKSFIELDNEW", "comment": "InfoSource field metadata"},
    "rsksfieldnewt": {"metadata_tabname": "RSKSFIELDNEWT", "comment": "InfoSource field text metadata"},
    "rstrant": {"metadata_tabname": "RSTRANT", "comment": "Transformation text metadata"},
    "rstranstepcnst": {"metadata_tabname": "RSTRANSTEPCNST", "comment": "Transformation constant rule metadata"},
    "rstransteprout": {"metadata_tabname": "RSTRANSTEPROUT", "comment": "Transformation formula and routine rule metadata"},
}
_DDIC_METADATA_CACHE: Tuple[pd.DataFrame, Dict[str, int]] | None = None
_TABLE_COLUMN_NAMES_CACHE: Dict[str, Set[str]] = {}
_TABLE_EXISTS_CACHE: Dict[str, bool] = {}
_RSIOBJ_TYPE_LOOKUP_CACHE: Dict[str, str] | None = None
_ADSO_FIELD_TEXT_LOOKUP_CACHE: Dict[Tuple[str, str], Dict[Tuple[str, str], str]] = {}
_ADSO_XML_FIELD_LOOKUP_CACHE: Dict[str, Dict[str, Tuple[str, str, str]]] = {}
_TRCS_FIELD_LOOKUP_CACHE: Dict[str, Dict[str, Tuple[str, str, str]]] = {}
RSDIOBJ_IOBJTP_TO_FIELD_TYPE = {
    "CHA": "I",
    "KYF": "I",
    "UNI": "I",
}
RSDIOBJ_BACKED_FIELD_TYPES = frozenset({"I", "K", "U"})

app = FastAPI(title="Dataflow Import Status API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOW_ORIGINS,
    allow_origin_regex=CORS_ALLOW_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class ImportStatusUpdate(BaseModel):
    table_name: str


class ProgCodeUpsertRequest(BaseModel):
    prog_id: str
    prog_type: str = ""
    prog_code: str = ""


class AbapToSqlRequest(BaseModel):
    output_table: str = ""
    abap_code: str = ""


class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    username: str
    role: str
    session_token: str = ""


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class AdminCreateUserRequest(BaseModel):
    username: str
    password: str
    role: str


class AdminResetPasswordRequest(BaseModel):
    new_password: str


class PathSelectionRequest(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    SOURCE: str = Field(default="", alias="source_name")
    SOURCESYS: str = Field(default="", alias="source_system")
    TARGETNAME: str = Field(default="", alias="target_name")
    TRANID: str = Field(default="", alias="tran_id")
    WAYPOINTS: List[str] = Field(default_factory=list, alias="waypoints")


class PathSegmentRequest(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    SOURCE: str = Field(default="", alias="source")
    TARGETNAME: str = Field(default="", alias="target")
    SOURCETYPE: str = Field(default="", alias="source_type")
    TARGETTYPE: str = Field(default="", alias="target_type")
    SOURCESUBTYPE: str = Field(default="", alias="source_subtype")
    TARGETSUBTYPE: str = Field(default="", alias="target_subtype")
    SOURCESYS: str = Field(default="", alias="source_system")
    TARGETSYSTEM: str = Field(default="", alias="target_system")
    TRANIDS: List[str] = Field(default_factory=list, alias="tran_ids")


class PathMappingRequest(BaseModel):
    segments: List[PathSegmentRequest] = []
    include_logic: bool = False
    include_text: bool = False


class DataQueryFilterRequest(BaseModel):
    field_ref: str = ""
    operator: str = "in"
    values: List[str] = []
    range_start: str = ""
    range_end: str = ""


class DataQueryJoinConditionRequest(BaseModel):
    main_field: str = ""
    join_field: str = ""


class DataQueryRequest(BaseModel):
    main_table: str = ""
    join_table: str = ""
    join_type: str = "left"
    select_fields: List[str] = []
    main_join_field: str = ""
    join_join_field: str = ""
    join_conditions: List[DataQueryJoinConditionRequest] = []
    filters: List[DataQueryFilterRequest] = []
    limit: int = DATA_QUERY_PREVIEW_LIMIT
    offset: int = 0


STEP_LEVEL_ROUTINE_KINDS = {"START", "END", "EXPERT"}
RULE_LEVEL_ROUTINE_KINDS = {"FORMULA", "NORMAL", "ROUTINE"}
ROUTINE_KIND_SORT_ORDER = {
    "GLOBAL": 0,
    "START": 1,
    "END": 2,
    "EXPERT": 3,
    "FORMULA": 10,
    "NORMAL": 11,
    "ROUTINE": 12,
    "CONSTANT": 20,
}


def utcnow() -> datetime:
    return datetime.utcnow()


def normalize_username(value: str) -> str:
    return str(value or "").strip().lower()


def normalize_bw_object_lookup(value: str | None) -> str:
    return str(value or "").strip().upper()


def validate_password_strength(password: str) -> None:
    error = get_password_strength_error(password)
    if error:
        raise HTTPException(status_code=400, detail=error)


def get_password_strength_error(password: str) -> Optional[str]:
    raw = str(password or "")
    if len(raw) < AUTH_PASSWORD_MIN_LEN:
        return f"密码至少 {AUTH_PASSWORD_MIN_LEN} 位"
    if not any(c.islower() for c in raw):
        return "密码必须包含小写字母"
    if not any(c.isupper() for c in raw):
        return "密码必须包含大写字母"
    if not any(c.isdigit() for c in raw):
        return "密码必须包含数字"
    if not any(not c.isalnum() for c in raw):
        return "密码必须包含特殊字符"
    return None


def hash_password(password: str, iterations: int = 210000) -> str:
    salt = secrets.token_bytes(16)
    hashed = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    salt_b64 = base64.b64encode(salt).decode("ascii")
    hash_b64 = base64.b64encode(hashed).decode("ascii")
    return f"pbkdf2_sha256${iterations}${salt_b64}${hash_b64}"


def verify_password(password: str, encoded: str) -> bool:
    try:
        algo, iterations_txt, salt_b64, hash_b64 = str(encoded or "").split("$", 3)
        if algo != "pbkdf2_sha256":
            return False
        iterations = int(iterations_txt)
        salt = base64.b64decode(salt_b64.encode("ascii"))
        expected = base64.b64decode(hash_b64.encode("ascii"))
    except Exception:
        return False

    actual = hashlib.pbkdf2_hmac("sha256", str(password or "").encode("utf-8"), salt, iterations)
    return hmac.compare_digest(actual, expected)


def hash_session_token(raw_token: str) -> str:
    return hashlib.sha256(str(raw_token or "").encode("utf-8")).hexdigest()


def parse_datetime_like(value: object) -> object:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return value

    # Fast reject non datetime-like text before trying strptime.
    if len(text) < 19:
        return value
    if text[4:5] != "-" or text[7:8] != "-":
        return value
    sep = text[10:11]
    if sep not in {" ", "T"}:
        return value
    if text[13:14] != ":" or text[16:17] != ":":
        return value

    if len(text) == 19:
        fmt = f"%Y-%m-%d{sep}%H:%M:%S"
    elif len(text) > 20 and text[19:20] == "." and text[20:].isdigit():
        fmt = f"%Y-%m-%d{sep}%H:%M:%S.%f"
    else:
        return value

    try:
        return datetime.strptime(text, fmt)
    except ValueError:
        return value
    return value


def sqlite_value(value: object) -> object:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return 1 if value else 0
    return value


def sqlite_normalize_sql(sql: str) -> str:
    normalized = str(sql or "")
    normalized = normalized.replace("%s", "?")
    normalized = re.sub(r"\bNOW\(\)", "CURRENT_TIMESTAMP", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"DATE_FORMAT\(([^,]+),\s*'%Y-%m-%d %H:%i'\)", r"strftime('%Y-%m-%d %H:%M', \1)", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\bINSERT\s+IGNORE\s+INTO\b", "INSERT OR IGNORE INTO", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\bTRUNCATE\s+TABLE\s+(`?[A-Za-z0-9_]+`?)", r"DELETE FROM \1", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\bCAST\((.*?)\s+AS\s+UNSIGNED\)", r"CAST(\1 AS INTEGER)", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\bCAST\((.*?)\s+AS\s+CHAR\)", r"CAST(\1 AS TEXT)", normalized, flags=re.IGNORECASE)
    normalized = normalized.replace("<=>", "IS")
    return normalized


class SQLiteCompatCursor:
    def __init__(self, cursor: sqlite3.Cursor, dictionary: bool = False):
        self._cursor = cursor
        self._dictionary = bool(dictionary)

    def execute(self, sql: str, params: object = None):
        normalized = sqlite_normalize_sql(sql)
        if params is None:
            self._cursor.execute(normalized)
        elif isinstance(params, (list, tuple)):
            self._cursor.execute(normalized, tuple(sqlite_value(item) for item in params))
        else:
            self._cursor.execute(normalized, params)
        return self

    def executemany(self, sql: str, seq_of_params: object):
        normalized = sqlite_normalize_sql(sql)
        parsed_rows = []
        for row in seq_of_params or []:
            if isinstance(row, (list, tuple)):
                parsed_rows.append(tuple(sqlite_value(item) for item in row))
            else:
                parsed_rows.append(row)
        self._cursor.executemany(normalized, parsed_rows)
        return self

    def _convert_row(self, row: sqlite3.Row | tuple | None):
        if row is None:
            return None
        if self._dictionary:
            keys = [item[0] for item in (self._cursor.description or [])]
            return {key: parse_datetime_like(row[key]) for key in keys}
        if isinstance(row, sqlite3.Row):
            return tuple(parse_datetime_like(value) for value in row)
        return tuple(parse_datetime_like(value) for value in row)

    def fetchone(self):
        return self._convert_row(self._cursor.fetchone())

    def fetchall(self):
        return [self._convert_row(row) for row in self._cursor.fetchall()]

    def fetchmany(self, size: int | None = None):
        rows = self._cursor.fetchmany(size) if size is not None else self._cursor.fetchmany()
        return [self._convert_row(row) for row in rows]

    def close(self):
        self._cursor.close()

    def __getattr__(self, name: str):
        return getattr(self._cursor, name)


class SQLiteCompatConnection:
    def __init__(self, connection: sqlite3.Connection):
        self._connection = connection

    def cursor(self, dictionary: bool = False):
        return SQLiteCompatCursor(self._connection.cursor(), dictionary=dictionary)

    def commit(self):
        self._connection.commit()

    def rollback(self):
        self._connection.rollback()

    def close(self):
        self._connection.close()

    def __getattr__(self, name: str):
        return getattr(self._connection, name)


def sqlite_resolve_table_name(conn: SQLiteCompatConnection | sqlite3.Connection, table_name: str) -> str:
    raw_conn = conn._connection if isinstance(conn, SQLiteCompatConnection) else conn
    cur = raw_conn.cursor()
    try:
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND LOWER(name) = LOWER(?) LIMIT 1",
            (str(table_name or "").strip(),),
        )
        row = cur.fetchone()
        return str(row[0]) if row and row[0] else str(table_name or "").strip()
    finally:
        cur.close()


def get_conn():
    if IS_SQLITE:
        sqlite_path = Path(SQLITE_DB_PATH)
        sqlite_path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(str(sqlite_path), check_same_thread=False)
        connection.row_factory = sqlite3.Row
        return SQLiteCompatConnection(connection)
    return mysql.connector.connect(**DB_CONFIG)


def ensure_status_table() -> None:
    if IS_SQLITE:
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS `import_status` (
                  `table_name` TEXT NOT NULL,
                  `last_import_at` TEXT NOT NULL,
                  `last_import_count` INTEGER NOT NULL DEFAULT 0,
                  `updated_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`table_name`)
                )
                """
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS `import_status` (
          `table_name` VARCHAR(64) NOT NULL,
          `last_import_at` DATETIME NOT NULL,
          `last_import_count` INT NOT NULL DEFAULT 0,
          `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`table_name`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
    )

    # Backward compatible migration for existing table.
    cur.execute(
        """
        SELECT COUNT(*)
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'import_status' AND COLUMN_NAME = 'last_import_count'
        """,
        (DB_CONFIG["database"],),
    )
    has_count_col = cur.fetchone()[0] > 0
    if not has_count_col:
        cur.execute("ALTER TABLE import_status ADD COLUMN last_import_count INT NOT NULL DEFAULT 0 AFTER last_import_at")

    conn.commit()
    cur.close()
    conn.close()


def migrate_legacy_infoobject_tables() -> None:
    if IS_SQLITE:
        return
    conn = get_conn()
    cur = conn.cursor()
    try:
        for old_name, new_name in (("rsiobj", "rsdiobj"), ("rsiobjt", "rsdiobjt")):
            cur.execute(
                """
                SELECT COUNT(*)
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """,
                (DB_CONFIG["database"], old_name),
            )
            old_exists = int(cur.fetchone()[0] or 0) > 0
            cur.execute(
                """
                SELECT COUNT(*)
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """,
                (DB_CONFIG["database"], new_name),
            )
            new_exists = int(cur.fetchone()[0] or 0) > 0

            if old_exists and not new_exists:
                cur.execute(f"RENAME TABLE `{old_name}` TO `{new_name}`")

        cur.execute(
            """
            UPDATE import_status AS legacy
            LEFT JOIN import_status AS modern ON modern.table_name = 'rsdiobj'
            SET legacy.table_name = 'rsdiobj'
            WHERE legacy.table_name = 'rsiobj'
              AND modern.table_name IS NULL
            """
        )
        cur.execute(
            """
            UPDATE import_status AS legacy
            LEFT JOIN import_status AS modern ON modern.table_name = 'rsdiobjt'
            SET legacy.table_name = 'rsdiobjt'
            WHERE legacy.table_name = 'rsiobjt'
              AND modern.table_name IS NULL
            """
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_bw_object_name_schema() -> None:
    """Keep bw_object_name schema aligned with business rules.

    - SOURCESYS can be NULL.
    - Use NAME_EN / NAME_DE fields.
    - Remove legacy OBJECT_NAME field.
    """
    conn = get_conn()
    cur = conn.cursor()

    def ensure_bw_object_name_indexes() -> None:
        cur.execute("SHOW INDEX FROM `bw_object_name` WHERE Key_name = 'idx_bw_object_lookup'")
        if not cur.fetchall():
            cur.execute(
                "CREATE INDEX `idx_bw_object_lookup` ON `bw_object_name` (`BW_OBJECT`, `BW_OBJECT_TYPE`, `SOURCESYS`)"
            )

        cur.execute("SHOW INDEX FROM `bw_object_name` WHERE Key_name = 'idx_bw_object_sourcesys'")
        if not cur.fetchall():
            cur.execute(
                "CREATE INDEX `idx_bw_object_sourcesys` ON `bw_object_name` (`BW_OBJECT`, `SOURCESYS`)"
            )

        cur.execute("SHOW INDEX FROM `bw_object_name` WHERE Key_name = 'idx_bw_object_norm_lookup'")
        if not cur.fetchall():
            cur.execute(
                "CREATE INDEX `idx_bw_object_norm_lookup` ON `bw_object_name` (`BW_OBJECT_NORM`, `BW_OBJECT_TYPE`, `SOURCESYS`)"
            )

        cur.execute("SHOW INDEX FROM `bw_object_name` WHERE Key_name = 'idx_bw_object_norm_sourcesys'")
        if not cur.fetchall():
            cur.execute(
                "CREATE INDEX `idx_bw_object_norm_sourcesys` ON `bw_object_name` (`BW_OBJECT_NORM`, `SOURCESYS`)"
            )

    try:
        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name'
            """,
            (DB_CONFIG["database"],),
        )
        if int(cur.fetchone()[0]) == 0:
            cur.execute(
                """
                CREATE TABLE `bw_object_name` (
                  `BW_OBJECT` varchar(40) NOT NULL COMMENT 'BW object',
                                    `BW_OBJECT_NORM` varchar(40) NOT NULL COMMENT 'BW object normalized to uppercase',
                  `SOURCESYS` varchar(25) NULL COMMENT 'Source System',
                  `BW_OBJECT_TYPE` varchar(20) NULL COMMENT 'BW object type',
                  `NAME_EN` varchar(255) NULL COMMENT 'Object Name (EN)',
                  `NAME_DE` varchar(255) NULL COMMENT 'Object Name (DE)',
                                    `NAME_EN_NORM` varchar(255) NULL COMMENT 'Object Name (EN) normalized to uppercase',
                                    `NAME_DE_NORM` varchar(255) NULL COMMENT 'Object Name (DE) normalized to uppercase',
                  KEY `idx_bw_object_sourcesys` (`BW_OBJECT`, `SOURCESYS`),
                                    KEY `idx_bw_object_lookup` (`BW_OBJECT`, `BW_OBJECT_TYPE`, `SOURCESYS`),
                                    KEY `idx_bw_object_norm_sourcesys` (`BW_OBJECT_NORM`, `SOURCESYS`),
                                    KEY `idx_bw_object_norm_lookup` (`BW_OBJECT_NORM`, `BW_OBJECT_TYPE`, `SOURCESYS`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
                """
            )
            conn.commit()
            return

        cur.execute(
            """
            SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_COMMENT
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name'
            """,
            (DB_CONFIG["database"],),
        )
        col_rows = cur.fetchall()
        col_map = {
            str(name): {
                "type": str(col_type),
                "nullable": str(is_nullable).upper() == "YES",
                "comment": str(comment or ""),
            }
            for name, col_type, is_nullable, comment in col_rows
        }

        # Migrate OBJECT_NAME -> NAME_EN when needed.
        if "NAME_EN" not in col_map:
            if "OBJECT_NAME" in col_map:
                src = col_map["OBJECT_NAME"]
                nullable_sql = "NULL" if src["nullable"] else "NOT NULL"
                escaped_comment = src["comment"].replace("'", "''")
                comment_sql = f" COMMENT '{escaped_comment}'" if src["comment"] else ""
                cur.execute(
                    f"ALTER TABLE `bw_object_name` CHANGE COLUMN `OBJECT_NAME` `NAME_EN` {src['type']} {nullable_sql}{comment_sql}"
                )
            else:
                cur.execute("ALTER TABLE `bw_object_name` ADD COLUMN `NAME_EN` VARCHAR(255) NULL")

        # Ensure NAME_DE exists.
        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name' AND COLUMN_NAME = 'NAME_DE'
            """,
            (DB_CONFIG["database"],),
        )
        if int(cur.fetchone()[0]) == 0:
            cur.execute("ALTER TABLE `bw_object_name` ADD COLUMN `NAME_DE` VARCHAR(255) NULL")

        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name' AND COLUMN_NAME = 'BW_OBJECT_NORM'
            """,
            (DB_CONFIG["database"],),
        )
        if int(cur.fetchone()[0]) == 0:
            cur.execute(
                "ALTER TABLE `bw_object_name` ADD COLUMN `BW_OBJECT_NORM` VARCHAR(40) NULL COMMENT 'BW object normalized to uppercase' AFTER `BW_OBJECT`"
            )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name' AND COLUMN_NAME = 'NAME_EN_NORM'
            """,
            (DB_CONFIG["database"],),
        )
        if int(cur.fetchone()[0]) == 0:
            cur.execute(
                "ALTER TABLE `bw_object_name` ADD COLUMN `NAME_EN_NORM` VARCHAR(255) NULL COMMENT 'Object Name (EN) normalized to uppercase' AFTER `NAME_EN`"
            )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name' AND COLUMN_NAME = 'NAME_DE_NORM'
            """,
            (DB_CONFIG["database"],),
        )
        if int(cur.fetchone()[0]) == 0:
            cur.execute(
                "ALTER TABLE `bw_object_name` ADD COLUMN `NAME_DE_NORM` VARCHAR(255) NULL COMMENT 'Object Name (DE) normalized to uppercase' AFTER `NAME_DE`"
            )

        # If OBJECT_NAME still exists after migration, merge then drop.
        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name' AND COLUMN_NAME = 'OBJECT_NAME'
            """,
            (DB_CONFIG["database"],),
        )
        has_object_name = int(cur.fetchone()[0]) > 0
        if has_object_name:
            cur.execute(
                """
                UPDATE `bw_object_name`
                SET `NAME_EN` = COALESCE(NULLIF(TRIM(`NAME_EN`), ''), NULLIF(TRIM(`OBJECT_NAME`), ''))
                """
            )
            cur.execute("ALTER TABLE `bw_object_name` DROP COLUMN `OBJECT_NAME`")

        cur.execute(
            """
            SELECT IS_NULLABLE, COLUMN_TYPE, COLUMN_COMMENT
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'bw_object_name' AND COLUMN_NAME = 'SOURCESYS'
            """,
            (DB_CONFIG["database"],),
        )
        row = cur.fetchone()
        if not row:
            conn.commit()
            return

        is_nullable, column_type, column_comment = row
        if str(is_nullable).upper() != "YES":
            cur.execute("SHOW KEYS FROM `bw_object_name` WHERE Key_name = 'PRIMARY'")
            pk_rows = cur.fetchall()
            pk_cols = [r[4] for r in pk_rows]
            can_relax_sourcesys = True
            if "SOURCESYS" in pk_cols:
                try:
                    cur.execute("ALTER TABLE `bw_object_name` DROP PRIMARY KEY")
                except mysql.connector.Error as exc:
                    # TiDB clustered index can reject DROP PRIMARY KEY (e.g. error 8200).
                    # In that case keep existing schema and avoid blocking API startup.
                    print(f"[startup] Skip bw_object_name PK migration on current DB engine: {exc}")
                    can_relax_sourcesys = False

            if can_relax_sourcesys:
                escaped_comment = str(column_comment or "").replace("'", "''")
                comment_sql = f" COMMENT '{escaped_comment}'" if column_comment is not None else ""
                cur.execute(f"ALTER TABLE `bw_object_name` MODIFY COLUMN `SOURCESYS` {column_type} NULL{comment_sql}")

                ensure_bw_object_name_indexes()
        else:
            ensure_bw_object_name_indexes()

        cur.execute(
            """
            UPDATE `bw_object_name`
            SET
              `BW_OBJECT_NORM` = UPPER(TRIM(COALESCE(`BW_OBJECT`, ''))),
              `NAME_EN_NORM` = NULLIF(UPPER(TRIM(COALESCE(`NAME_EN`, ''))), ''),
              `NAME_DE_NORM` = NULLIF(UPPER(TRIM(COALESCE(`NAME_DE`, ''))), '')
            WHERE COALESCE(`BW_OBJECT_NORM`, '') <> UPPER(TRIM(COALESCE(`BW_OBJECT`, '')))
               OR COALESCE(`NAME_EN_NORM`, '') <> COALESCE(NULLIF(UPPER(TRIM(COALESCE(`NAME_EN`, ''))), ''), '')
               OR COALESCE(`NAME_DE_NORM`, '') <> COALESCE(NULLIF(UPPER(TRIM(COALESCE(`NAME_DE`, ''))), ''), '')
            """
        )

        cur.execute(
            """
            UPDATE `bw_object_name`
            SET `BW_OBJECT_NORM` = UPPER(TRIM(COALESCE(`BW_OBJECT`, '')))
            WHERE `BW_OBJECT_NORM` IS NULL
            """
        )

        try:
            cur.execute("ALTER TABLE `bw_object_name` MODIFY COLUMN `BW_OBJECT_NORM` VARCHAR(40) NOT NULL COMMENT 'BW object normalized to uppercase'")
        except mysql.connector.Error as exc:
            print(f"[startup] Skip bw_object_name BW_OBJECT_NORM tighten on current DB engine: {exc}")

        conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_rstran_schema() -> None:
    """Keep rstran schema aligned with SOURCE naming rule."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'rstran'
            """,
            (DB_CONFIG["database"],),
        )
        if int(cur.fetchone()[0]) == 0:
            conn.commit()
            return

        cur.execute(
            """
            SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'rstran'
            """,
            (DB_CONFIG["database"],),
        )
        rows = cur.fetchall()
        col_map = {str(name): (str(col_type), str(is_nullable).upper() == "YES") for name, col_type, is_nullable in rows}

        if "SOURCE" not in col_map:
            if "DATASOURCE" in col_map:
                src_type, src_nullable = col_map["DATASOURCE"]
                nullable_sql = "NULL" if src_nullable else "NOT NULL"
                cur.execute(f"ALTER TABLE `rstran` ADD COLUMN `SOURCE` {src_type} {nullable_sql}")
                cur.execute("UPDATE `rstran` SET `SOURCE` = `DATASOURCE` WHERE `SOURCE` IS NULL")
            else:
                cur.execute("ALTER TABLE `rstran` ADD COLUMN `SOURCE` VARCHAR(255) NULL")

        # Backfill SOURCE from SOURCENAME first token when still empty.
        cur.execute(
            """
            UPDATE `rstran`
            SET `SOURCE` = NULLIF(TRIM(SUBSTRING_INDEX(TRIM(`SOURCENAME`), ' ', 1)), '')
            WHERE (`SOURCE` IS NULL OR TRIM(`SOURCE`) = '')
              AND `SOURCENAME` IS NOT NULL
              AND TRIM(`SOURCENAME`) <> ''
            """
        )

        conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_rstran_mapping_rule_table() -> None:
    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{RSTRAN_MAPPING_RULE_TABLE}` (
                  `tran_id` TEXT NOT NULL,
                  `rule_id` INTEGER NOT NULL,
                  `step_id` INTEGER NOT NULL,
                  `seg_id` TEXT NOT NULL DEFAULT '',
                  `pair_index` INTEGER NOT NULL,
                  `ruleposit` TEXT NOT NULL DEFAULT '',
                  `source_field` TEXT NULL,
                  `target_field` TEXT NULL,
                  `target_keyflag` TEXT NULL,
                  `rule_type` TEXT NULL,
                                    `group_type` TEXT NULL,
                  `aggr` TEXT NULL,
                  `updated_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`tran_id`, `rule_id`, `step_id`, `seg_id`, `pair_index`)
                )
                """
            )
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_tran` ON `{RSTRAN_MAPPING_RULE_TABLE}` (`tran_id`)")
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_source` ON `{RSTRAN_MAPPING_RULE_TABLE}` (`tran_id`, `source_field`)")
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_target` ON `{RSTRAN_MAPPING_RULE_TABLE}` (`tran_id`, `target_field`)")
        else:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{RSTRAN_MAPPING_RULE_TABLE}` (
                  `tran_id` VARCHAR(64) NOT NULL,
                  `rule_id` INT NOT NULL,
                  `step_id` INT NOT NULL,
                  `seg_id` VARCHAR(64) NOT NULL DEFAULT '',
                  `pair_index` INT NOT NULL,
                  `ruleposit` VARCHAR(64) NOT NULL DEFAULT '',
                  `source_field` VARCHAR(255) NULL,
                  `target_field` VARCHAR(255) NULL,
                  `target_keyflag` VARCHAR(8) NULL,
                  `rule_type` VARCHAR(32) NULL,
                  `group_type` VARCHAR(32) NULL,
                  `aggr` VARCHAR(64) NULL,
                  `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  PRIMARY KEY (`tran_id`, `rule_id`, `step_id`, `seg_id`, `pair_index`),
                  KEY `idx_mapping_rule_tran` (`tran_id`),
                  KEY `idx_mapping_rule_source` (`tran_id`, `source_field`),
                  KEY `idx_mapping_rule_target` (`tran_id`, `target_field`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
            )

        existing_columns = set(get_table_columns(RSTRAN_MAPPING_RULE_TABLE))
        if "group_type" not in existing_columns:
            if IS_SQLITE:
                cur.execute(f"ALTER TABLE `{RSTRAN_MAPPING_RULE_TABLE}` ADD COLUMN `group_type` TEXT NULL")
            else:
                cur.execute(f"ALTER TABLE `{RSTRAN_MAPPING_RULE_TABLE}` ADD COLUMN `group_type` VARCHAR(32) NULL AFTER `rule_type`")
        conn.commit()
    finally:
        cur.close()
        conn.close()


def mysql_type_for_dd03l_metadata(row: pd.Series) -> str:
    datatype = str(row.get("DATATYPE") or "").strip().upper()
    length = int(row.get("LENG") or 0)
    decimals = int(row.get("DECIMALS") or 0)

    if datatype in {"CHAR", "NUMC", "CLNT", "LANG", "CUKY", "UNIT", "DATS", "TIMS"}:
        return f"CHAR({max(length, 1)})"
    if datatype == "INT1":
        return "TINYINT UNSIGNED"
    if datatype == "INT2":
        return "SMALLINT"
    if datatype in {"INT4", "INT"}:
        return "INT"
    if datatype == "INT8":
        return "BIGINT"
    if datatype in {"DEC", "CURR", "QUAN", "DECIMAL", "NUMERIC"}:
        precision = max(length, decimals + 1, 1)
        return f"DECIMAL({precision},{decimals})"
    if datatype in {"FLTP", "FLOAT"}:
        return "DOUBLE"
    if datatype in {"STRG", "STRING", "SSTRING", "RAWSTRING"}:
        return "TEXT"
    if datatype in {"RAW", "LRAW"}:
        return f"VARBINARY({max(length, 1)})"
    return f"VARCHAR({max(length, 1)})"


def get_dd03l_observed_char_lengths(frame: pd.DataFrame) -> Dict[str, int]:
    observed: Dict[str, int] = {}
    for column in frame.columns:
        column_name = str(column or "").strip()
        if not column_name:
            continue
        try:
            observed[column_name] = int(frame[column_name].astype(str).str.len().max() or 0)
        except Exception:
            observed[column_name] = 0
    return observed


def mysql_type_for_dd03l_storage(row: pd.Series, observed_lengths: Dict[str, int]) -> str:
    field_name = str(row.get("FIELDNAME") or "").strip()
    base_type = mysql_type_for_dd03l_metadata(row)
    match = re.fullmatch(r"CHAR\((\d+)\)", base_type)
    if not match:
        return base_type

    metadata_length = int(match.group(1))
    observed_length = int(observed_lengths.get(field_name, 0) or 0)
    if field_name.upper() in {"LANGU", "SPRAS"}:
        return f"CHAR({max(metadata_length, observed_length, 2)})"
    return f"CHAR({max(metadata_length, observed_length, 1)})"


def get_ddic_metadata_cache() -> Tuple[pd.DataFrame, Dict[str, int]] | Tuple[None, None]:
    global _DDIC_METADATA_CACHE
    if _DDIC_METADATA_CACHE is not None:
        return _DDIC_METADATA_CACHE
    if not DD03L_METADATA_PATH.exists():
        return None, None

    frame = pd.read_excel(DD03L_METADATA_PATH, sheet_name=0, header=2).fillna("")
    frame.columns = [str(column).strip() for column in frame.columns]
    observed_lengths = get_dd03l_observed_char_lengths(frame)
    _DDIC_METADATA_CACHE = (frame, observed_lengths)
    return _DDIC_METADATA_CACHE


def ensure_ddic_metadata_table_schema(table_name: str) -> None:
    config = DDIC_METADATA_TABLES.get(str(table_name or "").strip().lower())
    if not config:
        return
    if not DD03L_METADATA_PATH.exists():
        print(f"[startup] Skip dd03l schema ensure: metadata file not found at {DD03L_METADATA_PATH}")
        return

    frame, observed_lengths = get_ddic_metadata_cache()
    if frame is None or observed_lengths is None:
        print(f"[startup] Skip {table_name} schema ensure: metadata cache unavailable")
        return

    metadata_tabname = str(config.get("metadata_tabname") or "").strip().upper()
    fields = frame[frame["TABNAME"].astype(str).str.strip().str.upper() == metadata_tabname].copy()
    if fields.empty:
        print(f"[startup] Skip {table_name} schema ensure: no {metadata_tabname} field definitions found in metadata file")
        return

    ordered_fields = fields.sort_values(by=["POSITION", "FIELDNAME"], kind="stable")
    column_sql = []
    primary_keys = []
    processed_fields: Set[str] = set()

    for _, row in ordered_fields.iterrows():
        field_name = str(row.get("FIELDNAME") or "").strip()
        if not field_name or field_name.startswith(".") or field_name in processed_fields:
            continue
        processed_fields.add(field_name)
        data_type_sql = mysql_type_for_dd03l_storage(row, observed_lengths)
        not_null = str(row.get("NOTNULL") or "").strip().upper() == "X" or str(row.get("KEYFLAG") or "").strip().upper() == "X"
        null_sql = "NOT NULL" if not_null else "NULL"
        column_sql.append(f"`{field_name}` {data_type_sql} {null_sql} COMMENT '{field_name}'")
        if str(row.get("KEYFLAG") or "").strip().upper() == "X":
            primary_keys.append(f"`{field_name}`")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            """,
            (DB_CONFIG["database"], table_name),
        )
        if int(cur.fetchone()[0]) == 0:
            body = [f"  {item}" for item in column_sql]
            if primary_keys:
                body.append(f"  PRIMARY KEY ({', '.join(primary_keys)})")
            table_comment = str(config.get("comment") or table_name).replace("'", "''")
            cur.execute(
                f"CREATE TABLE `{table_name}` (\n"
                + ",\n".join(body)
                + f"\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='{table_comment}';"
            )
            conn.commit()
            return

        cur.execute(
            """
            SELECT COLUMN_NAME, COLUMN_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            """,
            (DB_CONFIG["database"], table_name),
        )
        existing_rows = cur.fetchall()
        existing_columns = {str(item[0] or "").strip() for item in existing_rows}
        existing_column_types = {str(item[0] or "").strip(): str(item[1] or "").strip() for item in existing_rows}

        pseudo_columns = sorted(column_name for column_name in existing_columns if column_name.startswith("."))
        for column_name in pseudo_columns:
            cur.execute(f"ALTER TABLE `{table_name}` DROP COLUMN `{column_name}`")
        if pseudo_columns:
            existing_columns -= set(pseudo_columns)

        modifications = []
        for _, row in ordered_fields.iterrows():
            field_name = str(row.get("FIELDNAME") or "").strip()
            if not field_name or field_name.startswith(".") or field_name not in existing_columns:
                continue
            data_type_sql = mysql_type_for_dd03l_storage(row, observed_lengths)
            data_type_sql = choose_preserved_column_type(existing_column_types.get(field_name, ""), data_type_sql)
            not_null = str(row.get("NOTNULL") or "").strip().upper() == "X" or str(row.get("KEYFLAG") or "").strip().upper() == "X"
            null_sql = "NOT NULL" if not_null else "NULL"
            modifications.append(f"MODIFY COLUMN `{field_name}` {data_type_sql} {null_sql} COMMENT '{field_name}'")

        if modifications:
            cur.execute(f"ALTER TABLE `{table_name}`\n  " + ",\n  ".join(modifications))

        for _, row in ordered_fields.iterrows():
            field_name = str(row.get("FIELDNAME") or "").strip()
            if not field_name or field_name.startswith(".") or field_name in existing_columns:
                continue
            data_type_sql = mysql_type_for_dd03l_storage(row, observed_lengths)
            not_null = str(row.get("NOTNULL") or "").strip().upper() == "X" or str(row.get("KEYFLAG") or "").strip().upper() == "X"
            null_sql = "NOT NULL" if not_null else "NULL"
            cur.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{field_name}` {data_type_sql} {null_sql} COMMENT '{field_name}'")

        conn.commit()
    finally:
        cur.close()
        conn.close()


def sqlite_type_for_ddic_metadata(datatype: object, length: object, decimals: object) -> str:
    normalized_type = str(datatype or "").strip().upper()
    normalized_length = str(length or "").strip()
    normalized_decimals = str(decimals or "").strip()
    if normalized_type in {"INT1", "INT2", "INT4", "INT8"}:
        return "INTEGER"
    if normalized_type in {"DEC", "CURR", "QUAN", "FLTP"}:
        return "REAL"
    if normalized_type in {"DATS", "TIMS", "CHAR", "NUMC", "LANG", "CLNT", "UNIT", "CUKY", "ACCP", "STRG", "RAW", "LCHR", "SSTR"}:
        return "TEXT"
    if normalized_length and normalized_decimals:
        return "TEXT"
    return "TEXT"


def ensure_sqlite_ddic_metadata_table_schema(table_name: str) -> None:
    normalized_table_name = str(table_name or "").strip().lower()
    config = DDIC_METADATA_TABLES.get(normalized_table_name)
    if not config or not IS_SQLITE:
        return
    if table_exists(normalized_table_name):
        return
    if not table_exists("dd03l"):
        return

    metadata_tabname = str(config.get("metadata_tabname") or "").strip().upper()
    if not metadata_tabname:
        return

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT FIELDNAME, POSITION, KEYFLAG, "NOTNULL" AS NOTNULL_FLAG, DATATYPE, LENG, DECIMALS
            FROM dd03l
            WHERE UPPER(TRIM(TABNAME)) = UPPER(%s)
            ORDER BY POSITION, FIELDNAME
            """,
            (metadata_tabname,),
        )
        rows = cur.fetchall()
        if not rows:
            return

        column_defs: List[str] = []
        primary_keys: List[str] = []
        seen_fields: Set[str] = set()
        for row in rows:
            field_name = str(row.get("FIELDNAME") or "").strip()
            if not field_name or field_name in seen_fields or field_name.startswith("."):
                continue
            seen_fields.add(field_name)
            data_type_sql = sqlite_type_for_ddic_metadata(row.get("DATATYPE"), row.get("LENG"), row.get("DECIMALS"))
            not_null = str(row.get("NOTNULL_FLAG") or "").strip().upper() == "X" or str(row.get("KEYFLAG") or "").strip().upper() == "X"
            null_sql = "NOT NULL" if not_null else "NULL"
            column_defs.append(f'"{field_name}" {data_type_sql} {null_sql}')
            if str(row.get("KEYFLAG") or "").strip().upper() == "X":
                primary_keys.append(f'"{field_name}"')

        if not column_defs:
            return

        body = list(column_defs)
        if primary_keys:
            body.append(f"PRIMARY KEY ({', '.join(primary_keys)})")
        raw_conn = conn._connection if isinstance(conn, SQLiteCompatConnection) else conn
        raw_cur = raw_conn.cursor()
        try:
            raw_cur.execute(f'CREATE TABLE IF NOT EXISTS "{normalized_table_name}" (\n  ' + ',\n  '.join(body) + '\n)')
            raw_conn.commit()
        finally:
            raw_cur.close()

        _TABLE_EXISTS_CACHE.pop(normalized_table_name, None)
        _TABLE_COLUMN_NAMES_CACHE.pop(normalized_table_name, None)
    finally:
        cur.close()
        conn.close()


def ensure_import_target_table_schema(table_name: str) -> None:
    normalized_table_name = str(table_name or "").strip().lower()
    if not normalized_table_name:
        return
    if table_exists(normalized_table_name):
        return
    if normalized_table_name == PROG_CODE_TABLE:
        ensure_prog_code_schema()
        return
    if IS_SQLITE and normalized_table_name in DDIC_METADATA_TABLES:
        ensure_sqlite_ddic_metadata_table_schema(normalized_table_name)


def ensure_prog_code_schema() -> None:
    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{PROG_CODE_TABLE}` (
                  `prog_id` TEXT NOT NULL,
                  `prog_type` TEXT NOT NULL DEFAULT '',
                  `prog_code` TEXT NOT NULL DEFAULT '',
                  `updated_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`prog_id`)
                )
                """
            )
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_{PROG_CODE_TABLE}_type` ON `{PROG_CODE_TABLE}` (`prog_type`)")
        else:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{PROG_CODE_TABLE}` (
                  `prog_id` VARCHAR(40) NOT NULL,
                  `prog_type` VARCHAR(20) NOT NULL DEFAULT '',
                  `prog_code` LONGTEXT NOT NULL,
                  `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  PRIMARY KEY (`prog_id`),
                  KEY `idx_{PROG_CODE_TABLE}_type` (`prog_type`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='Program code import table';
                """
            )
        conn.commit()
        _TABLE_EXISTS_CACHE.pop(PROG_CODE_TABLE, None)
        _TABLE_COLUMN_NAMES_CACHE.pop(PROG_CODE_TABLE, None)
    finally:
        cur.close()
        conn.close()


def normalize_prog_code_text(value: object) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n")


def normalize_abap_object_name(value: object) -> str:
    return str(value or "").strip().replace("->", "=>").upper()


def normalize_abap_statement_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def strip_abap_inline_comment(line: str) -> str:
    if '"' not in line:
        return line
    return line.split('"', 1)[0]


def extract_abap_object_references(code: str) -> List[Dict[str, str]]:
    normalized_code = normalize_prog_code_text(code)
    found: List[Dict[str, str]] = []
    seen: Set[Tuple[str, str, str]] = set()

    patterns: List[Tuple[str, re.Pattern[str]]] = [
        ("function", re.compile(r"CALL\s+FUNCTION\s+'([^']+)'", re.IGNORECASE)),
        ("include", re.compile(r"\bINCLUDE\s+([A-Za-z0-9_/]+)\b", re.IGNORECASE)),
        ("program", re.compile(r"\bSUBMIT\s+([A-Za-z0-9_/]+)\b", re.IGNORECASE)),
        ("method", re.compile(r"CALL\s+METHOD\s+([A-Za-z0-9_/]+(?:=>|->)[A-Za-z0-9_/]+|[A-Za-z0-9_/]+)\b", re.IGNORECASE)),
        ("method", re.compile(r"\b([A-Za-z0-9_/]+=>[A-Za-z0-9_/]+)\b")),
    ]

    for ref_type, pattern in patterns:
        for match in pattern.finditer(normalized_code):
            name = normalize_abap_object_name(match.group(1))
            key = (ref_type, name, "")
            if not name or key in seen:
                continue
            seen.add(key)
            found.append({"type": ref_type, "name": name})

    perform_program_pattern = re.compile(r"\bPERFORM\s+([A-Za-z0-9_/]+)\s+IN\s+PROGRAM\s+([A-Za-z0-9_/]+)", re.IGNORECASE)
    for match in perform_program_pattern.finditer(normalized_code):
        form_name = normalize_abap_object_name(match.group(1))
        program_name = normalize_abap_object_name(match.group(2))
        key = ("form", form_name, program_name)
        if form_name and key not in seen:
            seen.add(key)
            found.append({"type": "form", "name": form_name, "program": program_name})

    perform_pattern = re.compile(r"\bPERFORM\s+([A-Za-z0-9_/]+)\b", re.IGNORECASE)
    for match in perform_pattern.finditer(normalized_code):
        form_name = normalize_abap_object_name(match.group(1))
        key = ("form", form_name, "")
        if form_name and key not in seen:
            seen.add(key)
            found.append({"type": "form", "name": form_name})

    return found


def fetch_prog_code_exact(prog_id: str) -> Dict[str, Any] | None:
    normalized = normalize_abap_object_name(prog_id)
    if not normalized:
        return None
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT `prog_id`, `prog_type`, `prog_code`, `updated_at`
            FROM `{PROG_CODE_TABLE}`
            WHERE UPPER(`prog_id`) = UPPER(%s)
            LIMIT 1
            """,
            (normalized,),
        )
        row = cur.fetchone()
        return row if row else None
    finally:
        cur.close()
        conn.close()


def search_prog_code_entries(keyword: str, limit: int = 50) -> List[Dict[str, Any]]:
    normalized_keyword = str(keyword or "").strip()
    if not normalized_keyword:
        return []

    resolved_limit = max(1, min(int(limit or 50), 200))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        base_sql = f"""
            SELECT `prog_id`, `prog_type`, `updated_at`
            FROM `{PROG_CODE_TABLE}`
        """
        if "*" in normalized_keyword:
            like_pattern = (
                normalized_keyword
                .replace("\\", "\\\\")
                .replace("%", "\\%")
                .replace("_", "\\_")
                .replace("*", "%")
            )
            cur.execute(
                base_sql + """
                WHERE UPPER(`prog_id`) LIKE UPPER(%s) ESCAPE '\\'
                ORDER BY LENGTH(`prog_id`) ASC, UPPER(`prog_id`) ASC
                LIMIT %s
                """,
                (like_pattern, resolved_limit),
            )
        else:
            cur.execute(
                base_sql + """
                WHERE UPPER(`prog_id`) = UPPER(%s)
                ORDER BY UPPER(`prog_id`) ASC
                LIMIT %s
                """,
                (normalized_keyword, resolved_limit),
            )
        rows = cur.fetchall() or []
    finally:
        cur.close()
        conn.close()

    return [
        {
            "prog_id": str(row.get("prog_id") or "").strip(),
            "prog_type": str(row.get("prog_type") or "").strip(),
            "updated_at": str(row.get("updated_at") or "").strip(),
        }
        for row in rows
    ]


def fetch_prog_code_by_body(pattern: str, preferred_prog_id: str = "") -> Dict[str, Any] | None:
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT `prog_id`, `prog_type`, `prog_code`, `updated_at`
            FROM `{PROG_CODE_TABLE}`
            WHERE UPPER(`prog_code`) LIKE UPPER(%s)
            ORDER BY CASE WHEN UPPER(`prog_id`) = UPPER(%s) THEN 0 ELSE 1 END, LENGTH(`prog_id`) ASC
            LIMIT 1
            """,
            (pattern, preferred_prog_id),
        )
        row = cur.fetchone()
        return row if row else None
    finally:
        cur.close()
        conn.close()


def resolve_prog_code_reference(reference: Dict[str, str]) -> Tuple[Dict[str, Any] | None, str]:
    ref_name = normalize_abap_object_name(reference.get("name"))
    ref_type = str(reference.get("type") or "").strip().lower()
    ref_program = normalize_abap_object_name(reference.get("program"))

    if ref_type == "form" and ref_program:
        program_entry = fetch_prog_code_exact(ref_program)
        if program_entry and re.search(rf"\bFORM\s+{re.escape(ref_name)}\b", str(program_entry.get("prog_code") or ""), re.IGNORECASE):
            return program_entry, f"program:{ref_program}"

    exact_entry = fetch_prog_code_exact(ref_name)
    if exact_entry:
        return exact_entry, "prog_id_exact"

    search_pattern = ""
    if ref_type == "form":
        search_pattern = f"%FORM {ref_name}.%"
    elif ref_type == "function":
        search_pattern = f"%FUNCTION {ref_name}%"
    elif ref_type == "method":
        search_pattern = f"%METHOD {ref_name.split('=>')[-1]}.%"
    elif ref_type in {"include", "program"}:
        search_pattern = f"%PROGRAM {ref_name}.%"

    if search_pattern:
        fallback_entry = fetch_prog_code_by_body(search_pattern, preferred_prog_id=ref_name)
        if fallback_entry:
            return fallback_entry, "body_search"

    return None, "not_found"


def resolve_abap_call_chain(code: str, max_depth: int = 40) -> Dict[str, Any]:
    detected_refs = extract_abap_object_references(code)
    queue: deque[Dict[str, str]] = deque(detected_refs)
    processed_keys: Set[Tuple[str, str, str]] = set()
    matched_entries: List[Dict[str, Any]] = []
    matched_prog_ids: Set[str] = set()
    missing_refs: List[Dict[str, str]] = []

    while queue and len(processed_keys) < max_depth:
        reference = queue.popleft()
        key = (
            str(reference.get("type") or "").strip().lower(),
            normalize_abap_object_name(reference.get("name")),
            normalize_abap_object_name(reference.get("program")),
        )
        if not key[1] or key in processed_keys:
            continue
        processed_keys.add(key)

        entry, matched_by = resolve_prog_code_reference(reference)
        if not entry:
            missing_refs.append({
                "type": key[0],
                "name": key[1],
                "program": key[2],
            })
            continue

        prog_id = normalize_abap_object_name(entry.get("prog_id"))
        if prog_id in matched_prog_ids:
            continue
        matched_prog_ids.add(prog_id)
        matched_entries.append({
            "prog_id": prog_id,
            "prog_type": str(entry.get("prog_type") or ""),
            "prog_code": normalize_prog_code_text(entry.get("prog_code")),
            "matched_by": matched_by,
            "source_ref": key[1],
        })

        for nested_ref in extract_abap_object_references(str(entry.get("prog_code") or "")):
            queue.append(nested_ref)

    return {
        "detected_refs": detected_refs,
        "matched_entries": matched_entries,
        "missing_refs": missing_refs,
    }


def normalize_abap_identifier_to_sql(name: str, fallback: str = "FIELD") -> str:
    normalized = str(name or "").strip()
    normalized = normalized.replace("<", "").replace(">", "")
    normalized = normalized.replace("->", "_").replace("=>", "_").replace("~", "_")
    normalized = normalized.replace("-", "_").replace("/", "_")
    normalized = re.sub(r"[^A-Za-z0-9_]", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    if not normalized:
        normalized = fallback
    if normalized[0].isdigit():
        normalized = f"_{normalized}"
    return normalized.upper()


def render_sql_identifier(name: str, fallback: str = "FIELD") -> str:
    raw_name = normalize_abap_statement_text(name).rstrip(".")
    if not raw_name:
        raw_name = fallback
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", raw_name):
        return raw_name
    escaped_name = raw_name.replace('"', '""')
    return f'"{escaped_name}"'


def render_abap_assignment_target(target: str, fallback: str = "FIELD") -> str:
    raw_target = normalize_abap_statement_text(target).rstrip(".")
    result_field_match = re.match(r"^<[^>]+>-(.+)$", raw_target)
    if result_field_match:
        return render_sql_identifier(result_field_match.group(1), fallback=fallback)
    upper_target = normalize_abap_object_name(raw_target)
    if upper_target.startswith("LS_RANGE-"):
        return render_sql_identifier(raw_target.split("-", 1)[1], fallback=fallback)
    return render_sql_identifier(raw_target, fallback=fallback)


def abap_expression_to_sql(expression: str) -> str:
    normalized = normalize_abap_statement_text(expression).rstrip(".")
    if not normalized:
        return "NULL"
    upper_expr = normalized.upper()
    if upper_expr in {"SPACE", "INITIAL"}:
        return "NULL"
    if upper_expr == "ABAP_TRUE":
        return "1"
    if upper_expr == "ABAP_FALSE":
        return "0"
    if re.fullmatch(r"'[^']*'", normalized):
        return normalized
    converted = normalized
    converted = converted.replace("->", ".").replace("=>", ".").replace("~", ".")
    converted = re.sub(r"<([A-Za-z0-9_/-]+)>", lambda match: match.group(1).replace("-", ".").replace("/", "_"), converted)
    converted = re.sub(r"\b([A-Za-z0-9_]+)-(/[^\s\.,]+)\b", lambda match: f'{match.group(1)}."{match.group(2)}"', converted)
    converted = re.sub(r"\b([A-Za-z0-9_]+)-([A-Za-z0-9_]+)\b", r"\1.\2", converted)
    converted = re.sub(r"\bL[TSVRO]?_[A-Za-z0-9_]+\.((?:\"/[^\"]+\")|(?:[A-Za-z0-9_]+))", r"\1", converted)
    converted = converted.replace(" EQ ", " = ").replace(" NE ", " <> ")
    return converted


def is_read_table_key_continuation(previous_line: str, current_line: str) -> bool:
    prev = normalize_abap_statement_text(previous_line).upper()
    current = normalize_abap_statement_text(current_line).upper()
    return bool(prev and "READ TABLE" in prev and "WITH KEY" in prev and current.startswith("FIELDNAME ="))


def is_non_output_abap_target(target: str) -> bool:
    normalized = normalize_abap_object_name(target)
    if not normalized:
        return True
    if normalized in {"P_SUBRC", "SY-TABIX", "SY_SUBRC", "SY-TABIX"}:
        return True
    if re.match(r"^<[^>]+>-.+$", str(target or "").strip()):
        return False
    if normalized.startswith(("L_", "LT_", "LS_", "LV_", "LO_", "LR_", "<")) and not normalized.startswith("LS_RANGE-"):
        return True
    return False


def normalize_abap_output_alias(target: str) -> str:
    normalized = str(target or "").strip()
    result_field_match = re.match(r"^<[^>]+>-(.+)$", normalized)
    if result_field_match:
        return normalize_abap_identifier_to_sql(result_field_match.group(1), fallback="FIELD")
    upper_target = normalize_abap_object_name(normalized)
    if upper_target.startswith("LS_RANGE-"):
        return normalize_abap_identifier_to_sql(upper_target.split("-", 1)[1], fallback="FIELD")
    return normalize_abap_identifier_to_sql(normalized, fallback="FIELD")


def is_supported_abap_direct_target(target: str) -> bool:
    raw_target = str(target or "").strip()
    upper_target = normalize_abap_object_name(raw_target)
    if re.match(r"^<(?:result_fields|source_fields)>-.+$", raw_target, re.IGNORECASE):
        return True
    return upper_target.startswith("LS_RANGE-")


def should_keep_program_output_target(target: str) -> bool:
    raw_target = normalize_abap_statement_text(target).rstrip(".")
    if not raw_target:
        return False
    if is_supported_abap_direct_target(raw_target):
        return True
    normalized_target = normalize_abap_object_name(raw_target)
    return bool(normalized_target) and normalized_target.startswith(("LV_", "GV_", "CV_", "RV_", "EV_"))


def parse_abap_call_method_blocks(code: str) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    lines = normalize_prog_code_text(code).split("\n")
    index = 0

    while index < len(lines):
        line = normalize_abap_statement_text(strip_abap_inline_comment(lines[index]))
        call_match = re.match(r"^CALL\s+METHOD\s+([^\s\.]+)\.?$", line, re.IGNORECASE)
        if not call_match:
            index += 1
            continue

        method_name = normalize_abap_object_name(call_match.group(1))
        block = {
            "method_name": method_name,
            "statement": line,
            "exporting": {},
            "importing": {},
            "changing": {},
            "receiving": {},
        }

        index += 1
        section = ""
        while index < len(lines):
            block_line = normalize_abap_statement_text(strip_abap_inline_comment(lines[index]))
            if not block_line or block_line.startswith("*"):
                index += 1
                continue

            upper_line = block_line.upper().rstrip(".")
            if upper_line in {"EXPORTING", "IMPORTING", "CHANGING", "RECEIVING"}:
                section = upper_line.lower()
                index += 1
                continue

            param_match = re.match(r"^([A-Za-z0-9_]+)\s*=\s*(.+?)(\.)?$", block_line, re.IGNORECASE)
            if section and param_match:
                formal_name = normalize_abap_object_name(param_match.group(1))
                actual_expr = normalize_abap_statement_text(param_match.group(2))
                block[section][formal_name] = actual_expr
                reached_end = bool(param_match.group(3))
                index += 1
                if reached_end:
                    break
                continue

            break

        blocks.append(block)

    return blocks


def parse_abap_call_function_blocks(code: str) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    lines = normalize_prog_code_text(code).split("\n")
    index = 0

    while index < len(lines):
        line = normalize_abap_statement_text(strip_abap_inline_comment(lines[index]))
        call_match = re.match(r"^CALL\s+FUNCTION\s+'([^']+)'\.?$", line, re.IGNORECASE)
        if not call_match:
            index += 1
            continue

        function_name = normalize_abap_object_name(call_match.group(1))
        block = {
            "function_name": function_name,
            "statement": line,
            "exporting": {},
            "importing": {},
            "changing": {},
            "tables": {},
        }

        index += 1
        section = ""
        while index < len(lines):
            block_line = normalize_abap_statement_text(strip_abap_inline_comment(lines[index]))
            if not block_line or block_line.startswith("*"):
                index += 1
                continue

            upper_line = block_line.upper().rstrip(".")
            if upper_line in {"EXPORTING", "IMPORTING", "CHANGING", "TABLES"}:
                section = upper_line.lower()
                index += 1
                continue

            if upper_line == "EXCEPTIONS":
                break

            param_match = re.match(r"^([A-Za-z0-9_]+)\s*=\s*(.+?)(\.)?$", block_line, re.IGNORECASE)
            if section and param_match:
                formal_name = normalize_abap_object_name(param_match.group(1))
                actual_expr = normalize_abap_statement_text(param_match.group(2))
                block[section][formal_name] = actual_expr
                reached_end = bool(param_match.group(3))
                index += 1
                if reached_end:
                    break
                continue

            break

        blocks.append(block)

    return blocks


def rewrite_method_expr_to_actual_sql(expression: str, actual_mapping: Dict[str, str]) -> str:
    rewritten = normalize_abap_statement_text(expression).rstrip(".")
    for formal_name, actual_expr in sorted(actual_mapping.items(), key=lambda item: len(item[0]), reverse=True):
        rewritten = re.sub(
            rf"\b{re.escape(formal_name)}\b",
            f"({normalize_abap_statement_text(actual_expr)})",
            rewritten,
            flags=re.IGNORECASE,
        )
    return abap_expression_to_sql(rewritten)


def extract_method_lookup_semantics(method_code: str) -> Dict[str, Any]:
    normalized_code = normalize_prog_code_text(method_code)
    select_match = re.search(
        r"SELECT\s+\*\s+FROM\s+([^\s]+)\s+INTO\s+TABLE\s+([A-Za-z0-9_<>\-/]+)\.",
        normalized_code,
        re.IGNORECASE,
    )
    read_match = re.search(
        r"READ\s+TABLE\s+([A-Za-z0-9_<>\-/]+)[\s\S]*?WITH\s+KEY\s+([A-Za-z0-9_<>\-/]+)\s*=\s*([A-Za-z0-9_<>\-/]+)\.",
        normalized_code,
        re.IGNORECASE,
    )
    if_block_match = re.search(
        r"IF\s+SY-SUBRC\s*=\s*0\.(?P<success>[\s\S]*?)ELSE\.(?P<fallback>[\s\S]*?)ENDIF\.",
        normalized_code,
        re.IGNORECASE,
    )
    if not (select_match and read_match and if_block_match):
        return {}

    success_block = if_block_match.group("success")
    fallback_block = if_block_match.group("fallback")
    success_assignments = re.findall(r"([A-Za-z0-9_<>\-/]+)\s*=\s*(.+?)\.", success_block, re.IGNORECASE)
    fallback_assignments = re.findall(r"([A-Za-z0-9_<>\-/]+)\s*=\s*(.+?)\.", fallback_block, re.IGNORECASE)
    fallback_map = {normalize_abap_object_name(target): normalize_abap_statement_text(expr) for target, expr in fallback_assignments}

    outputs: Dict[str, Dict[str, str]] = {}
    for target, expr in success_assignments:
        normalized_target = normalize_abap_object_name(target)
        if not normalized_target:
            continue
        outputs[normalized_target] = {
            "success_expr": normalize_abap_statement_text(expr),
            "fallback_expr": fallback_map.get(normalized_target, ""),
        }

    if not outputs:
        return {}

    return {
        "source_table": select_match.group(1).strip(),
        "lookup_table": normalize_abap_object_name(read_match.group(1)),
        "key_field": normalize_abap_statement_text(read_match.group(2)),
        "formal_input": normalize_abap_object_name(read_match.group(3)),
        "outputs": outputs,
    }


def build_method_output_expr_sql(semantics: Dict[str, Any], formal_output: str, actual_mapping: Dict[str, str]) -> str:
    source_table = semantics.get("source_table") or "source_table"
    key_field = semantics.get("key_field") or "lookup_key"
    input_expr_sql = rewrite_method_expr_to_actual_sql(semantics.get("formal_input") or "NULL", actual_mapping)
    output_semantics = (semantics.get("outputs") or {}).get(normalize_abap_object_name(formal_output)) or {}
    success_expr = normalize_abap_statement_text(output_semantics.get("success_expr") or "NULL")
    fallback_expr = normalize_abap_statement_text(output_semantics.get("fallback_expr") or "")

    field_match = re.match(r"^<[^>]+>-([A-Za-z0-9_/]+)$", success_expr, re.IGNORECASE)
    if field_match:
        lookup_field = normalize_abap_statement_text(field_match.group(1))
        fallback_expr_sql = rewrite_method_expr_to_actual_sql(fallback_expr or semantics.get("formal_input") or "NULL", actual_mapping)
        return (
            f"COALESCE(( SELECT DISTINCT {lookup_field} FROM {source_table} "
            f"WHERE {key_field} = {input_expr_sql} ), {fallback_expr_sql})"
        )

    success_expr_sql = rewrite_method_expr_to_actual_sql(success_expr, actual_mapping)
    fallback_expr_sql = rewrite_method_expr_to_actual_sql(fallback_expr or "NULL", actual_mapping)
    return (
        f"CASE WHEN EXISTS ( SELECT 1 FROM {source_table} WHERE {key_field} = {input_expr_sql} ) "
        f"THEN {success_expr_sql} ELSE {fallback_expr_sql} END"
    )


def derive_abap_function_call_rules(code: str, resolution: Dict[str, Any]) -> Tuple[List[Dict[str, str]], Set[str]]:
    function_rules: List[Dict[str, str]] = []
    ignored_aliases: Set[str] = set()
    matched_entries = resolution.get("matched_entries") or []
    matched_by_prog_id = {
        normalize_abap_object_name(entry.get("prog_id")): entry
        for entry in matched_entries
        if normalize_abap_object_name(entry.get("prog_id"))
    }

    for block in parse_abap_call_function_blocks(code):
        for formal_name in list((block.get("importing") or {}).keys()) + list((block.get("changing") or {}).keys()):
            ignored_aliases.add(normalize_abap_output_alias(formal_name))

        function_entry = matched_by_prog_id.get(normalize_abap_object_name(block.get("function_name")))
        if not function_entry:
            continue

        semantics = extract_method_lookup_semantics(str(function_entry.get("prog_code") or ""))
        if not semantics:
            continue

        actual_mapping = {}
        actual_mapping.update(block.get("exporting") or {})
        actual_mapping.update(block.get("changing") or {})
        actual_mapping.update(block.get("tables") or {})

        statement = normalize_abap_statement_text(block.get("statement") or "CALL FUNCTION")
        source_table = semantics.get("source_table") or "source_table"
        key_field = semantics.get("key_field") or "lookup_key"
        for formal_output in (semantics.get("outputs") or {}).keys():
            actual_target = str(
                (block.get("importing") or {}).get(formal_output)
                or (block.get("changing") or {}).get(formal_output)
                or ""
            ).strip()
            if not actual_target or not should_keep_program_output_target(actual_target):
                continue

            function_rules.append({
                "target": actual_target,
                "alias": normalize_abap_output_alias(actual_target),
                "expr_sql": build_method_output_expr_sql(semantics, formal_output, actual_mapping),
                "statement": f"{statement} -> 从 {source_table} 按 {key_field} 查值后回写 {actual_target}",
            })

    return function_rules, ignored_aliases


def derive_abap_method_call_rules(code: str, resolution: Dict[str, Any]) -> Tuple[List[Dict[str, str]], Set[str]]:
    method_rules: List[Dict[str, str]] = []
    ignored_aliases: Set[str] = set()
    matched_entries = resolution.get("matched_entries") or []
    matched_by_prog_id = {
        normalize_abap_object_name(entry.get("prog_id")): entry
        for entry in matched_entries
        if normalize_abap_object_name(entry.get("prog_id"))
    }

    for block in parse_abap_call_method_blocks(code):
        for formal_name in list((block.get("importing") or {}).keys()) + list((block.get("receiving") or {}).keys()):
            ignored_aliases.add(normalize_abap_output_alias(formal_name))

        method_entry = matched_by_prog_id.get(normalize_abap_object_name(block.get("method_name")))
        if not method_entry:
            continue

        semantics = extract_method_lookup_semantics(str(method_entry.get("prog_code") or ""))
        if not semantics:
            continue

        actual_mapping = {}
        actual_mapping.update(block.get("exporting") or {})
        actual_mapping.update(block.get("changing") or {})
        actual_mapping.update(block.get("receiving") or {})

        statement = normalize_abap_statement_text(block.get("statement") or "CALL METHOD")
        source_table = semantics.get("source_table") or "source_table"
        key_field = semantics.get("key_field") or "lookup_key"
        for formal_output in (semantics.get("outputs") or {}).keys():
            actual_target = str(
                (block.get("importing") or {}).get(formal_output)
                or (block.get("changing") or {}).get(formal_output)
                or (block.get("receiving") or {}).get(formal_output)
                or ""
            ).strip()
            if not actual_target or not should_keep_program_output_target(actual_target):
                continue

            method_rules.append({
                "target": actual_target,
                "alias": normalize_abap_output_alias(actual_target),
                "expr_sql": build_method_output_expr_sql(semantics, formal_output, actual_mapping),
                "statement": f"{statement} -> 从 {source_table} 按 {key_field} 查值后回写 {actual_target}",
            })

    return method_rules, ignored_aliases


def finalize_abap_output_rules(rule_order: List[str], rule_map: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    finalized: List[Dict[str, str]] = []
    seen_aliases: Set[str] = set()
    range_aliases = {"SIGN", "OPTION", "FIELDNAME", "IOBJNM", "LOW", "HIGH"}
    has_range_struct = any(alias in range_aliases for alias in rule_order)

    for alias in rule_order:
        rule = rule_map.get(alias)
        if not rule:
            continue
        if has_range_struct and alias not in range_aliases:
            continue
        if alias in seen_aliases:
            continue
        seen_aliases.add(alias)
        finalized.append(rule)

    return finalized


def collect_abap_field_rules(code: str) -> List[Dict[str, str]]:
    rule_order: List[str] = []
    rule_map: Dict[str, Dict[str, str]] = {}
    previous_line = ""

    def store_rule(target: str, expr_sql: str, statement: str) -> None:
        if is_non_output_abap_target(target) or not is_supported_abap_direct_target(target):
            return
        alias = normalize_abap_output_alias(target)
        payload = {
            "target": target,
            "alias": alias,
            "expr_sql": expr_sql or "NULL",
            "statement": normalize_abap_statement_text(statement),
        }
        if alias not in rule_map:
            rule_order.append(alias)
        rule_map[alias] = payload

    for raw_line in normalize_prog_code_text(code).split("\n"):
        line = normalize_abap_statement_text(strip_abap_inline_comment(raw_line))
        if not line or line.startswith("*"):
            continue

        if is_read_table_key_continuation(previous_line, line):
            previous_line = line
            continue

        assign_match = re.match(r"^([A-Za-z0-9_<>\-/]+)\s*=\s*(.+?)\.$", line, re.IGNORECASE)
        if assign_match:
            store_rule(assign_match.group(1), abap_expression_to_sql(assign_match.group(2)), line)
            previous_line = line
            continue

        previous_line = line

    return finalize_abap_output_rules(rule_order, rule_map)


def detect_abap_result_merge_semantics(code: str) -> Dict[str, Any]:
    normalized_code = normalize_prog_code_text(code)
    upper_code = normalized_code.upper()
    return {
        "has_range_append": "APPEND LINES OF L_T_RESULT TO L_T_RANGE" in upper_code,
        "has_range_insert": bool(re.search(r"\bINSERT\s+LS_RANGE\s+INTO\s+TABLE\s+L_T_RANGE\b", upper_code)),
        "has_logsys_lookup": bool(re.search(r"READ\s+TABLE\s+L_T_RANGE\s+WITH\s+KEY[\s\S]*FIELDNAME\s*=\s*'LOGSYS'", upper_code)),
        "has_modify_existing": "MODIFY L_T_RANGE INDEX L_IDX" in upper_code,
        "has_append_new": bool(re.search(r"\bELSE\.\s*APPEND\s+L_T_RANGE\.", upper_code, re.DOTALL)),
    }


def extract_select_source_summary(code: str) -> Dict[str, str]:
    normalized_code = normalize_prog_code_text(code)
    select_match = re.search(r"SELECT\s+\*\s+FROM\s+([^\s]+)\s+INTO\s+TABLE\s+([A-Za-z0-9_<>\-/]+)\s+WHERE\s+([^\.]+)\.", normalized_code, re.IGNORECASE)
    if not select_match:
        return {}
    return {
        "source_table": select_match.group(1).strip(),
        "into_table": select_match.group(2).strip(),
        "where_clause": normalize_abap_statement_text(select_match.group(3)),
    }


def unwrap_sql_literal(value: str) -> str:
    normalized = str(value or "").strip()
    if re.fullmatch(r"'[^']*'", normalized):
        return normalized[1:-1]
    return normalized


def get_rule_by_alias(rules: List[Dict[str, str]], alias: str) -> Dict[str, str] | None:
    for rule in rules:
        if str(rule.get("alias") or "").upper() == alias.upper():
            return rule
    return None


def map_sign_option_to_filter_semantics(sign_value: str, option_value: str) -> Dict[str, str]:
    # Fixed translation rules for ABAP selection conditions:
    # I + EQ -> IN
    # I + BT -> BETWEEN
    # I + CP -> LIKE
    # E + EQ -> NOT IN
    # E + BT -> NOT BETWEEN
    # E + CP -> NOT LIKE
    sign_token = unwrap_sql_literal(sign_value).upper()
    option_token = unwrap_sql_literal(option_value).upper()

    semantics = {
        ("I", "EQ"): {"label": "IN", "predicate": "IN"},
        ("E", "EQ"): {"label": "NOT IN", "predicate": "NOT IN"},
        ("I", "BT"): {"label": "BETWEEN", "predicate": "BETWEEN"},
        ("E", "BT"): {"label": "NOT BETWEEN", "predicate": "NOT BETWEEN"},
        ("I", "CP"): {"label": "LIKE", "predicate": "LIKE"},
        ("E", "CP"): {"label": "NOT LIKE", "predicate": "NOT LIKE"},
    }
    return semantics.get((sign_token, option_token), {
        "label": f"{sign_token}/{option_token}",
        "predicate": option_token or "UNKNOWN",
    })


def build_range_selection_summary(rules: List[Dict[str, str]], source_summary: Dict[str, str]) -> Dict[str, str]:
    sign_rule = get_rule_by_alias(rules, "SIGN")
    option_rule = get_rule_by_alias(rules, "OPTION")
    field_rule = get_rule_by_alias(rules, "FIELDNAME")
    low_rule = get_rule_by_alias(rules, "LOW")
    high_rule = get_rule_by_alias(rules, "HIGH")
    iobj_rule = get_rule_by_alias(rules, "IOBJNM")

    sign_expr = str(sign_rule.get("expr_sql") if sign_rule else "'I'")
    option_expr = str(option_rule.get("expr_sql") if option_rule else "'EQ'")
    field_expr = str(field_rule.get("expr_sql") if field_rule else "''")
    low_expr = str(low_rule.get("expr_sql") if low_rule else "NULL")
    high_expr = str(high_rule.get("expr_sql") if high_rule else "NULL")
    iobj_expr = str(iobj_rule.get("expr_sql") if iobj_rule else "NULL")
    semantics = map_sign_option_to_filter_semantics(sign_expr, option_expr)

    return {
        "field_name": unwrap_sql_literal(field_expr),
        "sign": unwrap_sql_literal(sign_expr),
        "option": unwrap_sql_literal(option_expr),
        "filter_label": semantics["label"],
        "predicate": semantics["predicate"],
        "low_expr": low_expr,
        "high_expr": high_expr,
        "iobj_expr": iobj_expr,
        "source_table": source_summary.get("source_table") or "source_table",
        "where_clause": source_summary.get("where_clause") or "1 = 1",
    }


def collect_abap_comment_hints(code: str, limit: int = 4) -> List[str]:
    hints: List[str] = []
    seen: Set[str] = set()
    for raw_line in normalize_prog_code_text(code).split("\n"):
        stripped = raw_line.strip()
        hint = ""
        if stripped.startswith("*"):
            hint = stripped.lstrip("*").strip()
        elif '"' in raw_line:
            hint = raw_line.split('"', 1)[1].strip()
        normalized_hint = normalize_abap_statement_text(hint)
        normalized_hint = re.sub(r'^["*\s]+', '', normalized_hint).strip()
        upper_hint = normalized_hint.upper()
        if re.fullmatch(r"[-=*\s]+", normalized_hint):
            continue
        if upper_hint in {"IMPORTING", "EXPORTING", "TABLES", "CHANGING"}:
            continue
        if upper_hint.startswith("VALUE(") or "LOKALE SCHNITTSTELLE" in upper_hint:
            continue
        if len(normalized_hint) < 4 or normalized_hint in seen:
            continue
        seen.add(normalized_hint)
        hints.append(normalized_hint)
        if len(hints) >= limit:
            break
    return hints


def collect_abap_processing_notes(code: str) -> List[str]:
    notes: List[str] = []
    seen: Set[str] = set()

    def add_note(text: str) -> None:
        normalized = normalize_abap_statement_text(text)
        if normalized and normalized not in seen:
            seen.add(normalized)
            notes.append(normalized)

    normalized_code = normalize_prog_code_text(code).upper()
    if "CONVERSION_EXIT_ALPHA_INPUT" in normalized_code:
        add_note("调用 CONVERSION_EXIT_ALPHA_INPUT，对字段做前导零补齐。")
    if "RSKC_CHAVL_CHECK" in normalized_code:
        add_note("调用 RSKC_CHAVL_CHECK，校验字符是否满足 BW 合法值规则。")
    if "ZFMSM_CHANGE_UNALLOWED_CHA" in normalized_code:
        add_note("调用 ZFMSM_CHANGE_UNALLOWED_CHA，对非法字符进行替换或清洗。")
    if re.search(r"\bCONDENSE\b", normalized_code):
        add_note("存在 CONDENSE 处理，说明文本会去首尾并压缩多余空格。")
    if re.search(r"\bTRANSLATE\b.+\bUPPER\s+CASE\b", normalized_code):
        add_note("存在大写转换逻辑。")
    if re.search(r"\bTRANSLATE\b.+\bLOWER\s+CASE\b", normalized_code):
        add_note("存在小写转换逻辑。")
    if re.search(r"\bSELECT\b", normalized_code):
        add_note("存在数据库读取逻辑，实际源表和过滤条件需要结合完整上下文确认。")
    return notes


def rewrite_package_pointer_sql(expression: str, target_table: str) -> str:
    rewritten = str(expression or "")
    table_ref = render_sql_identifier(target_table, fallback="TARGET_TABLE")
    rewritten = re.sub(r"\bresult_fields\.", f"{table_ref}.", rewritten, flags=re.IGNORECASE)
    rewritten = re.sub(r"\bsource_fields\.", f"{table_ref}.", rewritten, flags=re.IGNORECASE)
    rewritten = re.sub(r"\bresult_package\.", f"{table_ref}.", rewritten, flags=re.IGNORECASE)
    rewritten = re.sub(r"\bsource_package\.", f"{table_ref}.", rewritten, flags=re.IGNORECASE)
    return rewritten


def build_abap_location_text(output_table: str, resolution: Dict[str, Any]) -> str:
    detected_refs = resolution.get("detected_refs") or []
    matched_entries = resolution.get("matched_entries") or []
    missing_refs = resolution.get("missing_refs") or []
    lines = [
        f"生成表: {output_table}",
        f"识别对象数: {len(detected_refs)}",
        f"PROG_CODE 命中: {len(matched_entries)}",
        f"未命中对象: {len(missing_refs)}",
    ]

    if detected_refs:
        lines.append("")
        lines.append("识别到的对象:")
        for ref in detected_refs:
            program_suffix = f" | PROGRAM {ref.get('program')}" if ref.get("program") else ""
            lines.append(f"- {str(ref.get('type') or '').upper()}: {ref.get('name')}{program_suffix}")

    if matched_entries:
        lines.append("")
        lines.append("已递归定位源码:")
        for entry in matched_entries:
            match_reason = str(entry.get("matched_by") or "")
            lines.append(f"- {entry.get('prog_id')} ({match_reason})")

    if missing_refs:
        lines.append("")
        lines.append("未命中对象:")
        for ref in missing_refs:
            program_suffix = f" | PROGRAM {ref.get('program')}" if ref.get("program") else ""
            lines.append(f"- {str(ref.get('type') or '').upper()}: {ref.get('name')}{program_suffix}")

    return "\n".join(lines)


def build_abap_pseudo_sql(output_table: str, rules: List[Dict[str, str]], resolution: Dict[str, Any], comment_hints: List[str], processing_notes: List[str], full_code: str = "") -> str:
    target_table = normalize_abap_statement_text(output_table).strip() or "TARGET_TABLE"
    target_table_sql = render_sql_identifier(target_table, fallback="TARGET_TABLE")
    matched_entries = resolution.get("matched_entries") or []
    missing_refs = resolution.get("missing_refs") or []
    sql_lines: List[str] = [
        f"/* 生成表: {target_table} */",
        f"/* 识别对象 {len(resolution.get('detected_refs') or [])} 个，命中 PROG_CODE {len(matched_entries)} 个，未命中 {len(missing_refs)} 个。 */",
    ]

    if matched_entries:
        sql_lines.append(f"/* 已递归读取: {', '.join(str(entry.get('prog_id') or '') for entry in matched_entries)} */")
    if missing_refs:
        sql_lines.append(f"/* 未命中: {', '.join(str(ref.get('name') or '') for ref in missing_refs)} */")
    for hint in comment_hints:
        sql_lines.append(f"/* 源注释: {hint} */")

    merged_code = full_code or "\n\n".join([
        str(entry.get("prog_code") or "") for entry in matched_entries
    ])
    merge_semantics = detect_abap_result_merge_semantics(merged_code)
    source_summary = extract_select_source_summary(merged_code)

    effective_rules = rules or [{
        "alias": "PLACEHOLDER_FIELD",
        "expr_sql": "NULL",
        "statement": "未识别到明确字段赋值，请结合 ABAP 代码补充映射。",
    }]

    effective_aliases = {str(rule.get("alias") or "") for rule in effective_rules}

    if source_summary and effective_aliases.issuperset({"SIGN", "OPTION", "FIELDNAME", "LOW"}):
        range_summary = build_range_selection_summary(effective_rules, source_summary)
        source_table = range_summary.get("source_table") or "source_table"
        where_clause = range_summary.get("where_clause") or "1 = 1"
        field_name = range_summary.get("field_name") or "FIELD"
        filter_label = range_summary.get("filter_label") or "IN"
        low_expr = range_summary.get("low_expr") or "NULL"
        high_expr = range_summary.get("high_expr") or "NULL"

        if range_summary.get("option") in {"EQ", "BT", "CP"} and range_summary.get("sign") in {"I", "E"}:
            option_token = range_summary.get("option")
            sign_token = range_summary.get("sign")
            if option_token == "EQ":
                predicate = "IN" if sign_token == "I" else "NOT IN"
                action_label = "IN" if sign_token == "I" else "NOT IN"
            elif option_token == "BT":
                predicate = "BETWEEN" if sign_token == "I" else "NOT BETWEEN"
                action_label = predicate
            else:
                predicate = "LIKE" if sign_token == "I" else "NOT LIKE"
                action_label = predicate

            sql_lines.append("")
            sql_lines.append(f"/* 给 {field_name} 添加 {action_label} 过滤，值来自 {source_table}，前提是 {where_clause} */")
            if option_token == "EQ":
                sql_lines.append(f"{field_name} {predicate} ( SELECT DISTINCT {low_expr} FROM {source_table} WHERE {where_clause} );")
            elif option_token == "BT":
                sql_lines.append(f"{field_name} {predicate} ( SELECT DISTINCT {low_expr}, {high_expr} FROM {source_table} WHERE {where_clause} );")
            else:
                sql_lines.append(f"{field_name} {predicate} ( SELECT DISTINCT REPLACE({low_expr}, '*', '%') FROM {source_table} WHERE {where_clause} );")

            if merge_semantics.get("has_logsys_lookup") or merge_semantics.get("has_range_insert") or merge_semantics.get("has_range_append"):
                sql_lines.append("")
                sql_lines.append(f"/* 若结果对象中已存在 FIELDNAME='{field_name}' 的槽位，则按原索引回写；否则追加新记录。 */")

            if processing_notes:
                sql_lines.append("")
                for note in processing_notes:
                    sql_lines.append(f"/* {note} */")
            return "\n".join(sql_lines)

        sql_lines.append("")
        sql_lines.append(f"/* 步骤1：为字段 {field_name} 生成 {filter_label} 选择条件，值来源 {source_table}。 */")
        sql_lines.append("WITH RANGE_SELECTION AS (")
        sql_lines.append("  SELECT")
        for index, rule in enumerate(effective_rules):
            suffix = "," if index < len(effective_rules) - 1 else ""
            sql_lines.append(f"    /* {rule.get('statement') or '字段映射'} */")
            sql_lines.append(f"    {rule.get('expr_sql') or 'NULL'} AS {rule.get('alias') or 'PLACEHOLDER_FIELD'}{suffix}")
        sql_lines.append(f"  FROM {source_table}")
        sql_lines.append(f"  WHERE {where_clause}")
        sql_lines.append(")")
        sql_lines.append("")
        sql_lines.append(f"/* 步骤2：直接输出 {field_name} 的选择条件语义。 */")
        sql_lines.append("SELECT")
        sql_lines.append(f"  '{field_name}' AS FIELDNAME,")
        sql_lines.append(f"  '{filter_label}' AS FILTER_TYPE,")
        sql_lines.append("  LOW AS LOW,")
        if "HIGH" in effective_aliases:
            sql_lines.append("  HIGH AS HIGH,")
        else:
            sql_lines.append("  NULL AS HIGH,")
        if "IOBJNM" in effective_aliases:
            sql_lines.append("  IOBJNM AS IOBJNM")
        else:
            sql_lines.append("  NULL AS IOBJNM")
        sql_lines.append("FROM RANGE_SELECTION;")

        sql_lines.append("")
        sql_lines.append(f"/* 步骤3：{field_name} 的选择条件解释 */")
        if range_summary.get("option") == "EQ":
            sql_lines.append(f"/* {field_name} {filter_label} 单值集合；值来自 {source_table} 的字段表达式 {low_expr}。 */")
        elif range_summary.get("option") == "BT":
            sql_lines.append(f"/* {field_name} {filter_label} 区间；下限来自 {low_expr}，上限来自 {high_expr}。 */")
        elif range_summary.get("option") == "CP":
            sql_lines.append(f"/* {field_name} {filter_label} 模式匹配；模式值来自 {low_expr}。 */")
        else:
            sql_lines.append(f"/* {field_name} 使用 SIGN={range_summary.get('sign')}、OPTION={range_summary.get('option')} 生成选择条件。 */")

        if merge_semantics.get("has_logsys_lookup") or merge_semantics.get("has_range_insert") or merge_semantics.get("has_range_append"):
            sql_lines.append("")
            sql_lines.append(f"/* 步骤4：若结果对象中已存在 FIELDNAME='{field_name}' 的槽位，则按原索引回写；否则追加新记录。 */")
        return "\n".join(sql_lines + ([f"/* {note} */" for note in processing_notes] if processing_notes else []))

    sql_lines.append("")
    sql_lines.append(f"UPDATE {target_table_sql}")
    sql_lines.append("SET")

    for index, rule in enumerate(effective_rules):
        sql_lines.append(f"  /* {rule.get('statement') or '字段映射'} */")
        suffix = "," if index < len(effective_rules) - 1 else ""
        sql_lines.append(
            f"  {render_abap_assignment_target(str(rule.get('target') or rule.get('alias') or 'PLACEHOLDER_FIELD'))} = {rewrite_package_pointer_sql(rule.get('expr_sql') or 'NULL', target_table)}{suffix}"
        )

    sql_lines.append(";")

    if processing_notes:
        sql_lines.append("")
        for note in processing_notes:
            sql_lines.append(f"/* {note} */")

    return "\n".join(sql_lines)


def ensure_dd03l_schema() -> None:
    ensure_ddic_metadata_table_schema("dd03l")


def ensure_dd02t_schema() -> None:
    ensure_ddic_metadata_table_schema("dd02t")


def ensure_dd03t_schema() -> None:
    ensure_ddic_metadata_table_schema("dd03t")


def ensure_dd04t_schema() -> None:
    ensure_ddic_metadata_table_schema("dd04t")


def ensure_rsdiobj_schema() -> None:
    ensure_ddic_metadata_table_schema("rsdiobj")


def ensure_rsdiobjt_schema() -> None:
    ensure_ddic_metadata_table_schema("rsdiobjt")


def ensure_rsoadso_schema() -> None:
    ensure_ddic_metadata_table_schema("rsoadso")


def ensure_rsdssegfd_schema() -> None:
    ensure_ddic_metadata_table_schema("rsdssegfd")


def ensure_rsdssegfdt_schema() -> None:
    ensure_ddic_metadata_table_schema("rsdssegfdt")


def ensure_rsksnew_schema() -> None:
    ensure_ddic_metadata_table_schema("rsksnew")


def ensure_rsksnewt_schema() -> None:
    ensure_ddic_metadata_table_schema("rsksnewt")


def ensure_rsksfieldnew_schema() -> None:
    ensure_ddic_metadata_table_schema("rsksfieldnew")


def ensure_rsksfieldnewt_schema() -> None:
    ensure_ddic_metadata_table_schema("rsksfieldnewt")


def ensure_rstranstepcnst_schema() -> None:
    ensure_ddic_metadata_table_schema("rstranstepcnst")


def ensure_rstransteprout_schema() -> None:
    ensure_ddic_metadata_table_schema("rstransteprout")


def ensure_rstrant_schema() -> None:
    ensure_ddic_metadata_table_schema("rstrant")


def ensure_rstran_mapping_rule_full_table() -> None:
    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (
                  `tran_id` TEXT NOT NULL,
                  `rule_id` INTEGER NOT NULL,
                  `step_id` INTEGER NOT NULL,
                  `seg_id` TEXT NOT NULL DEFAULT '',
                  `pair_index` INTEGER NOT NULL DEFAULT 0,
                  `display_order` INTEGER NOT NULL DEFAULT 0,
                  `source_object` TEXT NOT NULL DEFAULT '',
                  `source_system` TEXT NOT NULL DEFAULT '',
                  `target_object` TEXT NOT NULL DEFAULT '',
                  `target_system` TEXT NOT NULL DEFAULT '',
                  `source_type` TEXT NOT NULL DEFAULT '',
                  `target_type` TEXT NOT NULL DEFAULT '',
                  `source_subtype` TEXT NOT NULL DEFAULT '',
                  `target_subtype` TEXT NOT NULL DEFAULT '',
                  `source_field` TEXT NOT NULL DEFAULT '',
                  `source_match_field` TEXT NOT NULL DEFAULT '',
                  `source_iobj_name` TEXT NOT NULL DEFAULT '',
                  `source_fieldtype` TEXT NOT NULL DEFAULT '',
                  `target_field` TEXT NOT NULL DEFAULT '',
                  `target_match_field` TEXT NOT NULL DEFAULT '',
                  `target_iobj_name` TEXT NOT NULL DEFAULT '',
                  `target_fieldtype` TEXT NOT NULL DEFAULT '',
                  `source_field_origin` TEXT NOT NULL DEFAULT '',
                  `target_field_origin` TEXT NOT NULL DEFAULT '',
                  `source_field_matched` INTEGER NOT NULL DEFAULT 0,
                  `target_field_matched` INTEGER NOT NULL DEFAULT 0,
                  `source_field_is_key` TEXT NULL,
                  `target_field_is_key` TEXT NULL,
                  `rule_type` TEXT NULL,
                  `group_type` TEXT NULL,
                  `aggr` TEXT NULL,
                  `ruleposit` TEXT NOT NULL DEFAULT '',
                  `row_kind` TEXT NOT NULL DEFAULT 'mapped',
                  `updated_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`tran_id`, `rule_id`, `step_id`, `seg_id`, `pair_index`, `source_field`, `target_field`)
                )
                """
            )
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_full_tran` ON `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (`tran_id`)")
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_full_source_obj` ON `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (`source_object`, `source_type`)")
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_full_target_obj` ON `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (`target_object`, `target_type`)")
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_full_source_field` ON `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (`tran_id`, `source_field`)")
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_mapping_rule_full_target_field` ON `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (`tran_id`, `target_field`)")
        else:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (
                  `tran_id` VARCHAR(64) NOT NULL,
                  `rule_id` INT NOT NULL,
                  `step_id` INT NOT NULL,
                  `seg_id` VARCHAR(64) NOT NULL DEFAULT '',
                  `pair_index` INT NOT NULL DEFAULT 0,
                  `display_order` INT NOT NULL DEFAULT 0,
                  `source_object` VARCHAR(255) NOT NULL DEFAULT '',
                  `source_system` VARCHAR(64) NOT NULL DEFAULT '',
                  `target_object` VARCHAR(255) NOT NULL DEFAULT '',
                  `target_system` VARCHAR(64) NOT NULL DEFAULT '',
                  `source_type` VARCHAR(32) NOT NULL DEFAULT '',
                  `target_type` VARCHAR(32) NOT NULL DEFAULT '',
                  `source_subtype` VARCHAR(32) NOT NULL DEFAULT '',
                  `target_subtype` VARCHAR(32) NOT NULL DEFAULT '',
                  `source_field` VARCHAR(255) NOT NULL DEFAULT '',
                  `source_match_field` VARCHAR(255) NOT NULL DEFAULT '',
                  `source_iobj_name` VARCHAR(255) NOT NULL DEFAULT '',
                  `source_fieldtype` VARCHAR(8) NOT NULL DEFAULT '',
                  `target_field` VARCHAR(255) NOT NULL DEFAULT '',
                  `target_match_field` VARCHAR(255) NOT NULL DEFAULT '',
                  `target_iobj_name` VARCHAR(255) NOT NULL DEFAULT '',
                  `target_fieldtype` VARCHAR(8) NOT NULL DEFAULT '',
                  `source_field_origin` VARCHAR(32) NOT NULL DEFAULT '',
                  `target_field_origin` VARCHAR(32) NOT NULL DEFAULT '',
                  `source_field_matched` TINYINT(1) NOT NULL DEFAULT 0,
                  `target_field_matched` TINYINT(1) NOT NULL DEFAULT 0,
                  `source_field_is_key` VARCHAR(8) NULL,
                  `target_field_is_key` VARCHAR(8) NULL,
                  `rule_type` VARCHAR(32) NULL,
                  `group_type` VARCHAR(32) NULL,
                  `aggr` VARCHAR(64) NULL,
                  `ruleposit` VARCHAR(64) NOT NULL DEFAULT '',
                  `row_kind` VARCHAR(24) NOT NULL DEFAULT 'mapped',
                  `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  PRIMARY KEY (`tran_id`, `rule_id`, `step_id`, `seg_id`, `pair_index`, `source_field`, `target_field`),
                  KEY `idx_mapping_rule_full_tran` (`tran_id`),
                  KEY `idx_mapping_rule_full_source_obj` (`source_object`, `source_type`),
                  KEY `idx_mapping_rule_full_target_obj` (`target_object`, `target_type`),
                  KEY `idx_mapping_rule_full_source_field` (`tran_id`, `source_field`),
                  KEY `idx_mapping_rule_full_target_field` (`tran_id`, `target_field`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='Materialized full field coverage by transformation step';
                """
            )

        existing_columns = set(get_table_columns(RSTRAN_MAPPING_RULE_FULL_TABLE))
        required_columns = {
            "display_order": "`display_order` INT NOT NULL DEFAULT 0 AFTER `pair_index`",
            "source_system": "`source_system` VARCHAR(64) NOT NULL DEFAULT '' AFTER `source_object`",
            "target_system": "`target_system` VARCHAR(64) NOT NULL DEFAULT '' AFTER `target_object`",
            "source_subtype": "`source_subtype` VARCHAR(32) NOT NULL DEFAULT '' AFTER `target_type`",
            "target_subtype": "`target_subtype` VARCHAR(32) NOT NULL DEFAULT '' AFTER `source_subtype`",
            "source_match_field": "`source_match_field` VARCHAR(255) NOT NULL DEFAULT '' AFTER `source_field`",
            "source_iobj_name": "`source_iobj_name` VARCHAR(255) NOT NULL DEFAULT '' AFTER `source_match_field`",
            "source_fieldtype": "`source_fieldtype` VARCHAR(8) NOT NULL DEFAULT '' AFTER `source_field`",
            "target_match_field": "`target_match_field` VARCHAR(255) NOT NULL DEFAULT '' AFTER `target_field`",
            "target_iobj_name": "`target_iobj_name` VARCHAR(255) NOT NULL DEFAULT '' AFTER `target_match_field`",
            "target_fieldtype": "`target_fieldtype` VARCHAR(8) NOT NULL DEFAULT '' AFTER `target_field`",
            "group_type": "`group_type` VARCHAR(32) NULL AFTER `rule_type`",
            "row_kind": "`row_kind` VARCHAR(24) NOT NULL DEFAULT 'mapped' AFTER `ruleposit`",
        }
        for column_name, column_sql in required_columns.items():
            if column_name not in existing_columns:
                if IS_SQLITE:
                    sqlite_column_sql = {
                        "display_order": "`display_order` INTEGER NOT NULL DEFAULT 0",
                        "source_system": "`source_system` TEXT NOT NULL DEFAULT ''",
                        "target_system": "`target_system` TEXT NOT NULL DEFAULT ''",
                        "source_subtype": "`source_subtype` TEXT NOT NULL DEFAULT ''",
                        "target_subtype": "`target_subtype` TEXT NOT NULL DEFAULT ''",
                        "source_match_field": "`source_match_field` TEXT NOT NULL DEFAULT ''",
                        "source_iobj_name": "`source_iobj_name` TEXT NOT NULL DEFAULT ''",
                        "source_fieldtype": "`source_fieldtype` TEXT NOT NULL DEFAULT ''",
                        "target_match_field": "`target_match_field` TEXT NOT NULL DEFAULT ''",
                        "target_iobj_name": "`target_iobj_name` TEXT NOT NULL DEFAULT ''",
                        "target_fieldtype": "`target_fieldtype` TEXT NOT NULL DEFAULT ''",
                        "group_type": "`group_type` TEXT NULL",
                        "row_kind": "`row_kind` TEXT NOT NULL DEFAULT 'mapped'",
                    }[column_name]
                    cur.execute(f"ALTER TABLE `{RSTRAN_MAPPING_RULE_FULL_TABLE}` ADD COLUMN {sqlite_column_sql}")
                else:
                    cur.execute(f"ALTER TABLE `{RSTRAN_MAPPING_RULE_FULL_TABLE}` ADD COLUMN {column_sql}")

        conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_bw_object_field_inventory_table() -> None:
    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{BW_OBJECT_FIELD_INVENTORY_TABLE}` (
                  `object_type` TEXT NOT NULL,
                  `object_name` TEXT NOT NULL,
                  `object_subtype` TEXT NOT NULL DEFAULT '',
                  `seg_id` TEXT NOT NULL DEFAULT '',
                  `field_name` TEXT NOT NULL,
                  `field_type` TEXT NOT NULL DEFAULT 'F',
                  `keyflag` TEXT NULL,
                  `datatype` TEXT NULL,
                  `length` INTEGER NULL,
                  `decimals` INTEGER NULL,
                  `field_order` INTEGER NOT NULL DEFAULT 0,
                  `source_origin` TEXT NOT NULL DEFAULT '',
                  `updated_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`object_type`, `object_name`, `object_subtype`, `seg_id`, `field_name`)
                )
                """
            )
            cur.execute(f"CREATE INDEX IF NOT EXISTS `idx_bw_object_field_inventory_object` ON `{BW_OBJECT_FIELD_INVENTORY_TABLE}` (`object_type`, `object_name`, `object_subtype`)")
        else:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS `{BW_OBJECT_FIELD_INVENTORY_TABLE}` (
                  `object_type` VARCHAR(32) NOT NULL,
                  `object_name` VARCHAR(255) NOT NULL,
                  `object_subtype` VARCHAR(32) NOT NULL DEFAULT '',
                  `seg_id` VARCHAR(64) NOT NULL DEFAULT '',
                  `field_name` VARCHAR(255) NOT NULL,
                  `field_type` VARCHAR(8) NOT NULL DEFAULT 'F',
                  `keyflag` VARCHAR(8) NULL,
                  `datatype` VARCHAR(32) NULL,
                  `length` INT NULL,
                  `decimals` INT NULL,
                  `field_order` INT NOT NULL DEFAULT 0,
                  `source_origin` VARCHAR(64) NOT NULL DEFAULT '',
                  `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  PRIMARY KEY (`object_type`, `object_name`, `object_subtype`, `seg_id`, `field_name`),
                  KEY `idx_bw_object_field_inventory_object` (`object_type`, `object_name`, `object_subtype`),
                  KEY `idx_bw_object_field_inventory_field` (`object_type`, `object_name`, `field_name`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='Materialized BW object field inventory';
                """
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def normalize_field_name(value: object) -> str:
    return str(value or "").strip().upper()


def build_empty_segment_inventory() -> Dict[str, object]:
    return {
        "ordered": [],
        "field_set": set(),
        "key_by_field": {},
        "type_by_field": {},
        "metadata_by_field": {},
    }


def parse_rsds_identity(raw_name: object, raw_system: object = None) -> Tuple[str, str]:
    datasource = str(raw_name or "").strip()
    logsys = str(raw_system or "").strip()
    composite = str(raw_name or "").strip()
    if composite:
        parts = composite.rsplit(None, 1)
        if len(parts) == 2:
            left, right = parts
            candidate_datasource = left.strip()
            candidate_logsys = right.strip()
            if candidate_datasource and (not logsys or normalize_bw_object_lookup(logsys) == normalize_bw_object_lookup(candidate_logsys)):
                datasource = candidate_datasource
                logsys = logsys or candidate_logsys
    return datasource, logsys


def get_field_text_language_codes(language: str = "EN") -> Tuple[str, ...]:
    normalized = str(language or "").strip().upper()
    if not normalized:
        return ("EN", "E")
    if normalized == "EN":
        return ("EN", "E")
    if len(normalized) == 1:
        return (normalized,)
    return (normalized, normalized[:1])


def get_object_text_language_codes(language: str = "EN") -> Tuple[str, ...]:
    ordered_codes: List[str] = []
    for candidate in (*get_field_text_language_codes(language), "DE", "D"):
        normalized = str(candidate or "").strip().upper()
        if normalized and normalized not in ordered_codes:
            ordered_codes.append(normalized)
    return tuple(ordered_codes)


def choose_preferred_object_text(candidates: List[Tuple[str, str]], language: str = "EN") -> str:
    if not candidates:
        return ""

    lang_priority = {code: index for index, code in enumerate(get_object_text_language_codes(language))}
    best_text = ""
    best_key: Tuple[int, int] | None = None
    for lang_code, text_value in candidates:
        normalized_text = str(text_value or "").strip()
        if not normalized_text:
            continue
        normalized_lang = str(lang_code or "").strip().upper()
        candidate_key = (lang_priority.get(normalized_lang, len(lang_priority) + 1), -len(normalized_text))
        if best_key is None or candidate_key < best_key:
            best_key = candidate_key
            best_text = normalized_text
    return best_text


def fetch_rsds_object_text_lookup(datasource_pairs: List[Tuple[str, str]], language: str = "EN") -> Dict[Tuple[str, str], str]:
    normalized_pairs: List[Tuple[str, str]] = []
    seen_pairs: Set[Tuple[str, str]] = set()
    for datasource, logsys in datasource_pairs or []:
        normalized_pair = (normalize_bw_object_lookup(datasource), normalize_bw_object_lookup(logsys))
        if not normalized_pair[0] or not normalized_pair[1] or normalized_pair in seen_pairs:
            continue
        seen_pairs.add(normalized_pair)
        normalized_pairs.append(normalized_pair)

    if not normalized_pairs or not table_exists("rsdst"):
        return {}

    available_columns = get_table_column_names("rsdst")
    text_columns = [column_name for column_name in ("TXTLG", "TXTMD", "TXTSH") if column_name in available_columns]
    if not text_columns:
        return {}

    pair_filters = []
    params: List[object] = []
    for datasource, logsys in normalized_pairs:
        pair_filters.append("(UPPER(TRIM(DATASOURCE)) = UPPER(%s) AND UPPER(TRIM(LOGSYS)) = UPPER(%s))")
        params.extend([datasource, logsys])

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT DATASOURCE, LOGSYS, LANGU, {', '.join(text_columns)}
            FROM rsdst
            WHERE OBJVERS = 'A'
              AND ({' OR '.join(pair_filters)})
            """,
            tuple(params),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    candidates_by_pair: Dict[Tuple[str, str], List[Tuple[str, str]]] = defaultdict(list)
    for row in rows:
        lookup_pair = (
            normalize_bw_object_lookup(row.get("DATASOURCE")),
            normalize_bw_object_lookup(row.get("LOGSYS")),
        )
        text_value = next((str(row.get(column_name) or "").strip() for column_name in text_columns if str(row.get(column_name) or "").strip()), "")
        if lookup_pair[0] and lookup_pair[1] and text_value:
            candidates_by_pair[lookup_pair].append((str(row.get("LANGU") or "").strip(), text_value))

    return {
        pair: choose_preferred_object_text(candidates, language)
        for pair, candidates in candidates_by_pair.items()
        if choose_preferred_object_text(candidates, language)
    }


def fetch_adso_object_text_lookup(object_names: List[str], language: str = "EN") -> Dict[str, str]:
    normalized_objects = sorted({normalize_bw_object_lookup(object_name) for object_name in (object_names or []) if normalize_bw_object_lookup(object_name)})
    if not normalized_objects or not table_exists("rsoadsot"):
        return {}

    available_columns = get_table_column_names("rsoadsot")
    text_columns = [column_name for column_name in ("DESCRIPTION", "QUICK_INFO") if column_name in available_columns]
    if not text_columns:
        return {}

    placeholders = ", ".join(["%s"] * len(normalized_objects))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT LANGU, ADSONM, OBJVERS, TTYP, COLNAME, {', '.join(text_columns)}
            FROM rsoadsot
            WHERE OBJVERS = 'A'
              AND UPPER(TRIM(ADSONM)) IN ({placeholders})
              AND COALESCE(TRIM(COLNAME), '') = ''
            """,
            tuple(normalized_objects),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    lang_priority = {code: index for index, code in enumerate(get_object_text_language_codes(language))}
    ttyp_priority = {"EUSR": 0, "": 1}
    best_by_object: Dict[str, Tuple[Tuple[int, int, int], str]] = {}
    for row in rows:
        object_name = normalize_bw_object_lookup(row.get("ADSONM"))
        if not object_name:
            continue
        text_value = next((str(row.get(column_name) or "").strip() for column_name in text_columns if str(row.get(column_name) or "").strip()), "")
        if not text_value:
            continue
        lang_code = str(row.get("LANGU") or "").strip().upper()
        ttyp = str(row.get("TTYP") or "").strip().upper()
        candidate_key = (
            ttyp_priority.get(ttyp, len(ttyp_priority) + 1),
            lang_priority.get(lang_code, len(lang_priority) + 1),
            -len(text_value),
        )
        existing = best_by_object.get(object_name)
        if existing is None or candidate_key < existing[0]:
            best_by_object[object_name] = (candidate_key, text_value)

    return {object_name: text_value for object_name, (_key, text_value) in best_by_object.items()}


def fetch_trcs_object_text_lookup(object_names: List[str], language: str = "EN") -> Dict[str, str]:
    normalized_objects = sorted({normalize_bw_object_lookup(object_name) for object_name in (object_names or []) if normalize_bw_object_lookup(object_name)})
    if not normalized_objects or not table_exists("rsksnewt"):
        return {}

    placeholders = ", ".join(["%s"] * len(normalized_objects))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT LANGU, ISOURCE, TXTLG
            FROM rsksnewt
            WHERE OBJVERS = 'A'
              AND UPPER(TRIM(ISOURCE)) IN ({placeholders})
            """,
            tuple(normalized_objects),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    candidates_by_object: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for row in rows:
        object_name = normalize_bw_object_lookup(row.get("ISOURCE"))
        text_value = str(row.get("TXTLG") or "").strip()
        if object_name and text_value:
            candidates_by_object[object_name].append((str(row.get("LANGU") or "").strip(), text_value))

    return {
        object_name: choose_preferred_object_text(candidates, language)
        for object_name, candidates in candidates_by_object.items()
        if choose_preferred_object_text(candidates, language)
    }


def fetch_iobj_object_text_lookup(object_names: List[str], language: str = "EN") -> Dict[str, str]:
    return fetch_rsiobj_text_lookup(object_names, language=language)


def resolve_path_object_display_name(
    object_name: str,
    object_system: str,
    object_type: str,
    *,
    rsds_name_lookup: Dict[Tuple[str, str], str],
    adso_name_lookup: Dict[str, str],
    trcs_name_lookup: Dict[str, str],
    iobj_name_lookup: Dict[str, str],
) -> str:
    normalized_object = normalize_bw_object_lookup(object_name)
    normalized_system = normalize_bw_object_lookup(object_system)
    normalized_type = normalize_type_code(object_type)
    if not normalized_object:
        return ""
    if normalized_type in {"RSDS", "DATASOURCE", "SOURCE"}:
        return str(rsds_name_lookup.get((normalized_object, normalized_system)) or "").strip()
    if normalized_type == "ADSO":
        return str(adso_name_lookup.get(normalized_object) or "").strip()
    if normalized_type == "TRCS":
        return str(trcs_name_lookup.get(normalized_object) or "").strip()
    if normalized_type == "IOBJ":
        return str(iobj_name_lookup.get(normalized_object) or "").strip()
    return ""


def fetch_rsds_field_text_lookup(
    datasource: str,
    logsys: str,
    language: str = "EN",
    field_names: Optional[Set[str]] = None,
    seg_ids: Optional[Set[str]] = None,
) -> Dict[Tuple[str, str], str]:
    normalized_ds = normalize_bw_object_lookup(datasource)
    normalized_logsys = normalize_bw_object_lookup(logsys)
    if not normalized_ds or not normalized_logsys or not table_exists("rsdssegfdt"):
        return {}

    lang_codes = get_field_text_language_codes(language)
    placeholders = ", ".join(["%s"] * len(lang_codes))
    normalized_field_names = sorted({normalize_field_name(name) for name in (field_names or set()) if normalize_field_name(name)})
    normalized_seg_ids = sorted({str(seg_id or "").strip() for seg_id in (seg_ids or set()) if str(seg_id or "").strip()})
    field_filter_sql = ""
    seg_filter_sql = ""
    params: List[object] = [normalized_ds, normalized_logsys, *lang_codes]
    if normalized_field_names:
        field_placeholders = ", ".join(["%s"] * len(normalized_field_names))
        field_filter_sql = f"\n              AND UPPER(TRIM(FIELDNM)) IN ({field_placeholders})"
        params.extend(normalized_field_names)
    if normalized_seg_ids:
        seg_placeholders = ", ".join(["%s"] * len(normalized_seg_ids))
        seg_filter_sql = f"\n              AND COALESCE(TRIM(SEGID), '') IN ({seg_placeholders})"
        params.extend(normalized_seg_ids)

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                SELECT SEGID, FIELDNM, TXTLG
                FROM rsdssegfdt
                WHERE DATASOURCE = %s
                  AND LOGSYS = %s
                  AND LANGU IN ({placeholders})
                  {field_filter_sql}
                  {seg_filter_sql}
                ORDER BY SEGID, POSIT
                """,
                tuple(params),
            )
        else:
            cur.execute(
                f"""
                SELECT SEGID, FIELDNM, TXTLG
                FROM rsdssegfdt
                WHERE UPPER(TRIM(DATASOURCE)) = UPPER(%s)
                  AND UPPER(TRIM(LOGSYS)) = UPPER(%s)
                  AND UPPER(TRIM(LANGU)) IN ({placeholders})
                  {field_filter_sql}
                  {seg_filter_sql}
                ORDER BY CAST(COALESCE(NULLIF(TRIM(SEGID), ''), '0') AS UNSIGNED),
                         CAST(COALESCE(NULLIF(TRIM(POSIT), ''), '0') AS UNSIGNED)
                """,
                tuple(params),
            )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    lookup: Dict[Tuple[str, str], str] = {}
    for row in rows:
        seg_id = str(row.get("SEGID") or "").strip()
        field_name = normalize_field_name(row.get("FIELDNM"))
        text_value = str(row.get("TXTLG") or "").strip()
        if not field_name or not text_value:
            continue
        seg_key = (seg_id, field_name)
        fallback_key = ("", field_name)
        if seg_key not in lookup:
            lookup[seg_key] = text_value
        if fallback_key not in lookup:
            lookup[fallback_key] = text_value
    return lookup


def fetch_adso_field_text_lookup(object_name: str, language: str = "EN") -> Dict[Tuple[str, str], str]:
    normalized_object = normalize_bw_object_lookup(object_name)
    if not normalized_object or not table_exists("dd03t"):
        return {}

    activate_data = fetch_rsoadso_activate_data_by_object([normalized_object]).get(normalized_object, "")
    expected_suffix = "2" if activate_data == "X" else "1"
    cache_key = (normalized_object, str(language or "EN").strip().upper())
    cached = _ADSO_FIELD_TEXT_LOOKUP_CACHE.get(cache_key)
    if cached is not None:
        return dict(cached)

    tabname = f"/BIC/A{normalized_object}{expected_suffix}"
    lang_codes = get_field_text_language_codes(language)
    placeholders = ", ".join(["%s"] * len(lang_codes))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT TABNAME, FIELDNAME, DDTEXT
            FROM dd03t
            WHERE UPPER(TRIM(TABNAME)) = UPPER(%s)
              AND UPPER(TRIM(DDLANGUAGE)) IN ({placeholders})
              AND UPPER(TRIM(COALESCE(AS4LOCAL, ''))) = 'A'
            ORDER BY TABNAME, FIELDNAME
            """,
            (tabname, *lang_codes),
        )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    rsiobj_names = fetch_rsiobj_name_set()
    lookup: Dict[Tuple[str, str], str] = {}
    for row in rows:
        field_name, _field_type = normalize_adso_field_name_with_rsiobj(row.get("FIELDNAME"), rsiobj_names)
        text_value = str(row.get("DDTEXT") or "").strip()
        if not field_name or not text_value:
            continue
        key = ("", field_name)
        if key not in lookup:
            lookup[key] = text_value
    _ADSO_FIELD_TEXT_LOOKUP_CACHE[cache_key] = dict(lookup)
    return lookup


def fetch_dd03t_table_field_text_lookup(table_names: List[str], language: str = "EN") -> Dict[Tuple[str, str], str]:
    normalized_table_names = sorted({str(table_name or "").strip().upper() for table_name in (table_names or []) if str(table_name or "").strip()})
    if not normalized_table_names or not table_exists("dd03t"):
        return {}

    lang_codes = get_field_text_language_codes(language)
    table_placeholders = ", ".join(["%s"] * len(normalized_table_names))
    lang_placeholders = ", ".join(["%s"] * len(lang_codes))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT TABNAME, FIELDNAME, DDTEXT
            FROM dd03t
            WHERE UPPER(TRIM(TABNAME)) IN ({table_placeholders})
              AND UPPER(TRIM(DDLANGUAGE)) IN ({lang_placeholders})
              AND UPPER(TRIM(COALESCE(AS4LOCAL, ''))) = 'A'
            ORDER BY TABNAME, FIELDNAME
            """,
            tuple(normalized_table_names + list(lang_codes)),
        )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    lookup: Dict[Tuple[str, str], str] = {}
    for row in rows:
        tabname = str(row.get("TABNAME") or "").strip().upper()
        field_name = str(row.get("FIELDNAME") or "").strip().upper()
        text_value = str(row.get("DDTEXT") or "").strip()
        if not tabname or not field_name or not text_value:
            continue
        lookup.setdefault((tabname, field_name), text_value)
    return lookup


def fetch_dd03t_field_text_fallback_lookup(field_names: List[str], language: str = "EN") -> Dict[str, str]:
    normalized_field_names = sorted({str(field_name or "").strip().upper() for field_name in (field_names or []) if str(field_name or "").strip()})
    if not normalized_field_names or not table_exists("dd03t"):
        return {}

    lang_codes = get_field_text_language_codes(language)
    field_placeholders = ", ".join(["%s"] * len(normalized_field_names))
    lang_placeholders = ", ".join(["%s"] * len(lang_codes))
    lang_order_cases = " ".join([f"WHEN '{code}' THEN {idx}" for idx, code in enumerate(lang_codes)])

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT FIELDNAME, DDLANGUAGE, DDTEXT
            FROM dd03t
            WHERE UPPER(TRIM(FIELDNAME)) IN ({field_placeholders})
              AND UPPER(TRIM(DDLANGUAGE)) IN ({lang_placeholders})
              AND UPPER(TRIM(COALESCE(AS4LOCAL, ''))) = 'A'
            ORDER BY FIELDNAME,
              CASE UPPER(TRIM(DDLANGUAGE)) {lang_order_cases} ELSE 999 END,
              TABNAME
            """,
            tuple(normalized_field_names + list(lang_codes)),
        )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    lookup: Dict[str, str] = {}
    for row in rows:
        field_name = str(row.get("FIELDNAME") or "").strip().upper()
        text_value = str(row.get("DDTEXT") or "").strip()
        if not field_name or not text_value:
            continue
        lookup.setdefault(field_name, text_value)
    return lookup


def fetch_rs_table_field_text_lookup_by_rollname(table_name: str, language: str = "EN") -> Dict[str, str]:
    normalized_table_name = str(table_name or "").strip().upper()
    if not normalized_table_name.startswith("RS"):
        return {}
    if not table_exists("dd03l") or not table_exists("dd04t"):
        return {}

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT FIELDNAME, ROLLNAME
            FROM dd03l
            WHERE UPPER(TRIM(TABNAME)) = %s
              AND UPPER(TRIM(COALESCE(AS4LOCAL, ''))) = 'A'
            ORDER BY POSITION
            """,
            (normalized_table_name,),
        )
        dd03l_rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    field_to_rollname: Dict[str, str] = {}
    rollnames: List[str] = []
    for row in dd03l_rows:
        field_name = str(row.get("FIELDNAME") or "").strip().upper()
        rollname = str(row.get("ROLLNAME") or "").strip().upper()
        if not field_name or not rollname:
            continue
        if field_name not in field_to_rollname:
            field_to_rollname[field_name] = rollname
            rollnames.append(rollname)

    if not rollnames:
        return {}

    lang_codes = get_field_text_language_codes(language)
    rollname_placeholders = ", ".join(["%s"] * len(rollnames))
    lang_placeholders = ", ".join(["%s"] * len(lang_codes))
    lang_order_cases = " ".join([f"WHEN '{code}' THEN {idx}" for idx, code in enumerate(lang_codes)])

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT ROLLNAME, DDTEXT, DDLANGUAGE
            FROM dd04t
            WHERE UPPER(TRIM(ROLLNAME)) IN ({rollname_placeholders})
              AND UPPER(TRIM(DDLANGUAGE)) IN ({lang_placeholders})
              AND UPPER(TRIM(COALESCE(AS4LOCAL, ''))) = 'A'
            ORDER BY ROLLNAME,
              CASE UPPER(TRIM(DDLANGUAGE)) {lang_order_cases} ELSE 999 END
            """,
            tuple(rollnames + list(lang_codes)),
        )
        dd04t_rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    rollname_text_lookup: Dict[str, str] = {}
    for row in dd04t_rows:
        rollname = str(row.get("ROLLNAME") or "").strip().upper()
        text_value = str(row.get("DDTEXT") or "").strip()
        if not rollname or not text_value:
            continue
        rollname_text_lookup.setdefault(rollname, text_value)

    field_text_lookup: Dict[str, str] = {}
    for field_name, rollname in field_to_rollname.items():
        text_value = rollname_text_lookup.get(rollname, "")
        if text_value:
            field_text_lookup[field_name] = text_value
    return field_text_lookup


def fetch_trcs_field_text_lookup(
    isource: str,
    language: str = "EN",
    field_names: Optional[Set[str]] = None,
    seg_ids: Optional[Set[str]] = None,
) -> Dict[Tuple[str, str], str]:
    normalized_isource = normalize_bw_object_lookup(isource)
    if not normalized_isource or not table_exists("rsksfieldnew") or not table_exists("rsksfieldnewt"):
        return {}

    lang_codes = get_field_text_language_codes(language)
    placeholders = ", ".join(["%s"] * len(lang_codes))
    normalized_seg_ids = sorted({str(seg_id or "").strip() for seg_id in (seg_ids or set()) if str(seg_id or "").strip()})
    seg_filter_sql = ""
    params: List[object] = [*lang_codes, normalized_isource]
    if normalized_seg_ids:
        seg_placeholders = ", ".join(["%s"] * len(normalized_seg_ids))
        seg_filter_sql = f"\n              AND COALESCE(TRIM(a.SEGID), '') IN ({seg_placeholders})"
        params.extend(normalized_seg_ids)

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                SELECT a.SEGID, a.POSIT, a.TYPE, a.IOBJNM, a.FIELDNM, b.TXTLG
                FROM rsksfieldnew AS a
                LEFT JOIN rsksfieldnewt AS b
                  ON b.ISOURCE = a.ISOURCE
                 AND COALESCE(b.SEGID, '') = COALESCE(a.SEGID, '')
                 AND COALESCE(b.POSIT, '') = COALESCE(a.POSIT, '')
                 AND b.LANGU IN ({placeholders})
                WHERE a.OBJVERS = 'A'
                  AND a.ISOURCE = %s
                  {seg_filter_sql}
                ORDER BY a.SEGID, a.POSIT
                """,
                tuple(params),
            )
        else:
            cur.execute(
                f"""
                SELECT a.SEGID, a.POSIT, a.TYPE, a.IOBJNM, a.FIELDNM, b.TXTLG
                FROM rsksfieldnew AS a
                LEFT JOIN rsksfieldnewt AS b
                  ON UPPER(TRIM(b.ISOURCE)) = UPPER(TRIM(a.ISOURCE))
                 AND COALESCE(TRIM(b.SEGID), '') = COALESCE(TRIM(a.SEGID), '')
                 AND COALESCE(TRIM(b.POSIT), '') = COALESCE(TRIM(a.POSIT), '')
                 AND UPPER(TRIM(b.LANGU)) IN ({placeholders})
                WHERE a.OBJVERS = 'A'
                  AND UPPER(TRIM(a.ISOURCE)) = UPPER(%s)
                  {seg_filter_sql}
                ORDER BY CAST(COALESCE(NULLIF(TRIM(a.SEGID), ''), '0') AS UNSIGNED),
                         CAST(COALESCE(NULLIF(TRIM(a.POSIT), ''), '0') AS UNSIGNED)
                """,
                tuple(params),
            )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    lookup: Dict[Tuple[str, str], str] = {}
    for row in rows:
        seg_id = str(row.get("SEGID") or "").strip()
        field_type = str(row.get("TYPE") or "").strip().upper()
        iobj_name = normalize_field_name(row.get("IOBJNM"))
        field_name_raw = row.get("FIELDNM")
        # TRCS rule: TYPE=F should use FIELDNM to match mapping field names.
        # Some datasets also leave IOBJNM empty for field-style rows; use FIELDNM as fallback.
        if field_type == "F" or not iobj_name:
            field_name = normalize_field_name(field_name_raw)
        else:
            field_name = iobj_name
        text_value = str(row.get("TXTLG") or "").strip()
        if not field_name or not text_value:
            continue
        seg_key = (seg_id, field_name)
        fallback_key = ("", field_name)
        if seg_key not in lookup:
            lookup[seg_key] = text_value
        if fallback_key not in lookup:
            lookup[fallback_key] = text_value
    return lookup


def fetch_rsiobj_text_lookup(iobj_names: List[str], language: str = "EN") -> Dict[str, str]:
    normalized_names = sorted({normalize_field_name(name) for name in (iobj_names or []) if normalize_field_name(name)})
    if not normalized_names or not table_exists("rsdiobjt"):
        return {}

    available_columns = get_table_column_names("rsdiobjt")
    text_priority = [column_name for column_name in ("TXTLG", "TXTMD", "TXTSH") if column_name in available_columns]
    if not text_priority:
        return {}

    lang_codes = get_field_text_language_codes(language)
    name_placeholders = ", ".join(["%s"] * len(normalized_names))
    lang_placeholders = ", ".join(["%s"] * len(lang_codes))
    text_expr = ", ".join([f"NULLIF(TRIM({column_name}), '')" for column_name in text_priority])
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT IOBJNM,
                   COALESCE({text_expr}) AS TEXT_VALUE
                        FROM rsdiobjt
            WHERE UPPER(TRIM(IOBJNM)) IN ({name_placeholders})
              AND UPPER(TRIM(LANGU)) IN ({lang_placeholders})
            ORDER BY IOBJNM
            """,
            (*normalized_names, *lang_codes),
        )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    lookup: Dict[str, str] = {}
    for row in rows:
        field_name = normalize_field_name(row.get("IOBJNM"))
        text_value = str(row.get("TEXT_VALUE") or "").strip()
        if not field_name or not text_value or field_name in lookup:
            continue
        lookup[field_name] = text_value
    return lookup


def fetch_rstrant_text_lookup(tran_ids: List[str], language: str = "EN") -> Dict[str, str]:
    normalized_tran_ids = sorted({str(tran_id or "").strip() for tran_id in (tran_ids or []) if str(tran_id or "").strip()})
    if not normalized_tran_ids or not table_exists("rstrant"):
        return {}

    available_columns = get_table_column_names("rstrant")
    if "TRANID" not in available_columns or "LANGU" not in available_columns:
        return {}

    text_priority = [column_name for column_name in ("TXTLG", "TXTMD", "TXTSH") if column_name in available_columns]
    if not text_priority:
        return {}

    tran_placeholders = ", ".join(["%s"] * len(normalized_tran_ids))
    lang_codes = get_object_text_language_codes(language)
    lang_placeholders = ", ".join(["%s"] * len(lang_codes))
    text_expr = ", ".join([f"NULLIF(TRIM({column_name}), '')" for column_name in text_priority])
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT TRANID,
                   LANGU,
                   COALESCE({text_expr}) AS TEXT_VALUE
            FROM rstrant
            WHERE UPPER(TRIM(TRANID)) IN ({tran_placeholders})
              AND UPPER(TRIM(LANGU)) IN ({lang_placeholders})
            ORDER BY TRANID
            """,
            (*normalized_tran_ids, *lang_codes),
        )
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    grouped_candidates: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for row in rows:
        tran_id = str(row.get("TRANID") or "").strip()
        text_value = str(row.get("TEXT_VALUE") or "").strip()
        if not tran_id or not text_value:
            continue
        grouped_candidates[tran_id].append((str(row.get("LANGU") or "").strip(), text_value))

    return {
        tran_id: choose_preferred_object_text(candidates, language)
        for tran_id, candidates in grouped_candidates.items()
    }


def build_field_text_lookup_for_object(
    object_name: str,
    object_system: str,
    object_type: str,
    field_names: Optional[Set[str]] = None,
    seg_ids: Optional[Set[str]] = None,
) -> Dict[Tuple[str, str], str]:
    normalized_type = normalize_type_code(object_type)
    if normalized_type in {"RSDS", "DATASOURCE", "SOURCE"}:
        datasource, logsys = parse_rsds_identity(object_name, object_system)
        return fetch_rsds_field_text_lookup(datasource, logsys, field_names=field_names, seg_ids=seg_ids)
    if normalized_type in {"ADSO", "ODSO"}:
        return fetch_adso_field_text_lookup(object_name)
    if normalized_type == "TRCS":
        return fetch_trcs_field_text_lookup(object_name, field_names=field_names, seg_ids=seg_ids)
    return {}


def resolve_field_text_from_lookup(lookup: Optional[Dict[Tuple[str, str], str]], seg_id: object, field_name: object) -> str:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return ""
    if not isinstance(lookup, dict):
        return ""
    seg_key = str(seg_id or "").strip()
    return str(lookup.get((seg_key, normalized_field)) or lookup.get(("", normalized_field)) or "").strip()


def enrich_mapping_rows_with_field_text(
    rows: List[Dict[str, object]],
    source_object: str,
    source_system: str,
    source_type: str,
    target_object: str,
    target_system: str,
    target_type: str,
    source_lookup: Optional[Dict[Tuple[str, str], str]] = None,
    target_lookup: Optional[Dict[Tuple[str, str], str]] = None,
    rsiobj_type_lookup: Optional[Dict[str, str]] = None,
) -> List[Dict[str, object]]:
    source_field_names = {
        normalize_field_name(row.get("source_field"))
        for row in rows
        if normalize_field_name(row.get("source_field"))
    }
    target_field_names = {
        normalize_field_name(row.get("target_field"))
        for row in rows
        if normalize_field_name(row.get("target_field"))
    }
    seg_ids = {
        str(row.get("seg_id") or "").strip()
        for row in rows
        if str(row.get("seg_id") or "").strip()
    }

    resolved_source_lookup = source_lookup if source_lookup is not None else (
        build_field_text_lookup_for_object(
            source_object,
            source_system,
            source_type,
            field_names=source_field_names,
            seg_ids=seg_ids,
        ) or {}
    )
    resolved_target_lookup = target_lookup if target_lookup is not None else (
        build_field_text_lookup_for_object(
            target_object,
            target_system,
            target_type,
            field_names=target_field_names,
            seg_ids=seg_ids,
        ) or {}
    )
    needed_iobj_names: Set[str] = set()
    for row in rows:
        source_iobj_name = normalize_field_name(row.get("source_iobj_name"))
        target_iobj_name = normalize_field_name(row.get("target_iobj_name"))
        if source_iobj_name:
            needed_iobj_names.add(source_iobj_name)
        if target_iobj_name:
            needed_iobj_names.add(target_iobj_name)

    iobj_text_lookup = fetch_rsiobj_text_lookup(sorted(needed_iobj_names)) if needed_iobj_names else {}
    if not resolved_source_lookup and not resolved_target_lookup and not iobj_text_lookup:
        return rows

    for row in rows:
        seg_id = row.get("seg_id")
        row["source_text"] = resolve_field_text_from_lookup(resolved_source_lookup, seg_id, row.get("source_field"))
        row["target_text"] = resolve_field_text_from_lookup(resolved_target_lookup, seg_id, row.get("target_field"))
        source_iobj_name = normalize_field_name(row.get("source_iobj_name"))
        target_iobj_name = normalize_field_name(row.get("target_iobj_name"))
        if not str(row.get("source_text") or "").strip() and source_iobj_name:
            row["source_text"] = str(iobj_text_lookup.get(source_iobj_name) or "").strip()
        if not str(row.get("target_text") or "").strip() and target_iobj_name:
            row["target_text"] = str(iobj_text_lookup.get(target_iobj_name) or "").strip()

    return rows


def mapping_row_match_field(row: Dict[str, object], field_key: str) -> str:
    if field_key == "source_field":
        return normalize_field_name(row.get("source_match_field")) or normalize_field_name(row.get("source_field"))
    if field_key == "target_field":
        return normalize_field_name(row.get("target_match_field")) or normalize_field_name(row.get("target_field"))
    return normalize_field_name(row.get(field_key))


def fetch_all_materialized_tran_mapping_rows() -> List[Dict[str, object]]:
    if not table_exists(RSTRAN_MAPPING_RULE_TABLE):
        return []

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT tran_id, rule_id, step_id, seg_id, pair_index, ruleposit,
                   source_field, target_field, target_keyflag, rule_type, group_type, aggr
            FROM `{RSTRAN_MAPPING_RULE_TABLE}`
            ORDER BY tran_id, rule_id, step_id, seg_id, pair_index
            """
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return normalize_tran_mapping_rows(rows)


def fetch_active_tran_metadata() -> Dict[str, Dict[str, str]]:
    if not table_exists("rstran"):
        return {}

    rstran_columns = set(get_table_columns("rstran"))
    source_col = "SOURCE" if "SOURCE" in rstran_columns else ("DATASOURCE" if "DATASOURCE" in rstran_columns else "")
    if not source_col:
        return {}
    source_subtype_sql = ", SOURCESUBTYPE" if "SOURCESUBTYPE" in rstran_columns else ", '' AS SOURCESUBTYPE"
    target_subtype_sql = ", TARGETSUBTYPE" if "TARGETSUBTYPE" in rstran_columns else ", '' AS TARGETSUBTYPE"

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT TRANID, SOURCETYPE, {source_col} AS SOURCE, SOURCESYS, TARGETTYPE, TARGETNAME{source_subtype_sql}{target_subtype_sql}
            FROM rstran
            WHERE OBJVERS = 'A'
            """
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    result: Dict[str, Dict[str, str]] = {}
    for row in rows:
        tran_id = str(row.get("TRANID") or "").strip()
        if not tran_id:
            continue

        source_type = normalize_type_code(row.get("SOURCETYPE"))
        target_type = normalize_type_code(row.get("TARGETTYPE"))

        source_name = str(row.get("SOURCE") or "").strip()
        source_system = ""
        if source_type == "RSDS":
            source_name, source_system = parse_rsds_identity(source_name, row.get("SOURCESYS"))

        target_name = str(row.get("TARGETNAME") or "").strip()
        target_system = ""
        if target_type == "RSDS":
            target_name, target_system = parse_rsds_identity(row.get("TARGETNAME"), None)

        result[tran_id] = {
            "tran_id": tran_id,
            "source_type": source_type,
            "target_type": target_type,
            "source_subtype": normalize_iobj_subtype(row.get("SOURCESUBTYPE")),
            "target_subtype": normalize_iobj_subtype(row.get("TARGETSUBTYPE")),
            "source_object": source_name,
            "source_system": source_system,
            "target_object": target_name,
            "target_system": target_system,
        }
    return result


def fetch_rsds_field_inventory() -> Dict[Tuple[str, str], Dict[str, Dict[str, object]]]:
    return fetch_rsds_field_inventory_for_pairs(None)


def fetch_rsds_field_inventory_for_pairs(
    datasource_pairs: List[Tuple[str, str]] | None,
) -> Dict[Tuple[str, str], Dict[str, Dict[str, object]]]:
    if not table_exists("rsdssegfd"):
        return {}

    normalized_pairs: List[Tuple[str, str]] = []
    if datasource_pairs is not None:
        seen_pairs: Set[Tuple[str, str]] = set()
        for datasource, logsys in datasource_pairs:
            normalized_datasource = normalize_bw_object_lookup(datasource)
            normalized_logsys = normalize_bw_object_lookup(logsys)
            if not normalized_datasource or not normalized_logsys:
                continue
            pair = (normalized_datasource, normalized_logsys)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            normalized_pairs.append(pair)
        if not normalized_pairs:
            return {}

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        if normalized_pairs:
            pair_conditions = " OR ".join(["(DATASOURCE = %s AND LOGSYS = %s)"] * len(normalized_pairs))
            params: List[str] = []
            for datasource, logsys in normalized_pairs:
                params.extend([datasource, logsys])
            cur.execute(
                f"""
                                SELECT DATASOURCE, LOGSYS, SEGID, POSIT, FIELDNM, KEYFIELD, DATATYPE, LENG, DECIMALS
                FROM rsdssegfd
                WHERE OBJVERS = 'A'
                  AND TRANSFER = 'X'
                  AND ({pair_conditions})
                ORDER BY DATASOURCE, LOGSYS,
                         CAST(COALESCE(NULLIF(TRIM(SEGID), ''), '0') AS UNSIGNED),
                         CAST(COALESCE(NULLIF(TRIM(POSIT), ''), '0') AS UNSIGNED),
                         FIELDNM
                """,
                tuple(params),
            )
        else:
            cur.execute(
                """
                                SELECT DATASOURCE, LOGSYS, SEGID, POSIT, FIELDNM, KEYFIELD, DATATYPE, LENG, DECIMALS
                FROM rsdssegfd
                WHERE OBJVERS = 'A'
                  AND TRANSFER = 'X'
                ORDER BY DATASOURCE, LOGSYS,
                         CAST(COALESCE(NULLIF(TRIM(SEGID), ''), '0') AS UNSIGNED),
                         CAST(COALESCE(NULLIF(TRIM(POSIT), ''), '0') AS UNSIGNED),
                         FIELDNM
                """
            )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]] = {}
    seen: Set[Tuple[str, str, str, str]] = set()
    for row in rows:
        datasource = normalize_bw_object_lookup(row.get("DATASOURCE"))
        logsys = normalize_bw_object_lookup(row.get("LOGSYS"))
        seg_id = str(row.get("SEGID") or "").strip()
        field_name = normalize_field_name(row.get("FIELDNM"))
        if not datasource or not logsys or not field_name:
            continue
        dedup_key = (datasource, logsys, seg_id, field_name)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        object_inventory = inventory.setdefault((datasource, logsys), {})
        seg_inventory = object_inventory.setdefault(
            seg_id,
            build_empty_segment_inventory(),
        )
        keyflag = str(row.get("KEYFIELD") or "").strip()
        seg_inventory["ordered"].append((field_name, keyflag))
        seg_inventory["field_set"].add(field_name)
        seg_inventory["key_by_field"][field_name] = keyflag
        seg_inventory["type_by_field"][field_name] = "F"
        seg_inventory["metadata_by_field"][field_name] = {
            "datatype": str(row.get("DATATYPE") or "").strip().upper(),
            "length": _normalize_inventory_numeric_value(row.get("LENG")),
            "decimals": _normalize_inventory_numeric_value(row.get("DECIMALS")),
        }

    return inventory


def derive_adso_name_from_dd03l_tabname(tabname: object) -> str:
    normalized_name = str(tabname or "").strip().upper()
    if not normalized_name.startswith("/BIC/A") or len(normalized_name) <= 7:
        return ""
    derived_name = normalized_name[6:-1].strip().upper()
    if derived_name == "TD010":
        return "TD01"
    if derived_name == "TD020":
        return "TD02"
    return derived_name


def derive_adso_field_name_from_dd03l(field_name: object) -> str:
    normalized_field = str(field_name or "").strip().upper()
    if not normalized_field:
        return ""
    if normalized_field.startswith("/BIC/"):
        return normalized_field[5:]
    return f"0{normalized_field}"


def is_dd03l_adso_business_field(field_name: object) -> bool:
    normalized_field = derive_adso_field_name_from_dd03l(field_name)
    return bool(normalized_field)


def dd03l_tabname_suffix_digit(tabname: object) -> str:
    normalized_name = str(tabname or "").strip().upper()
    if not normalized_name:
        return ""
    suffix = normalized_name[-1]
    return suffix if suffix.isdigit() else ""


def fetch_rsiobj_name_set() -> Set[str]:
    return set(fetch_rsiobj_type_lookup().keys())


def normalize_rsiobj_field_type_from_iobjtp(iobjtp: object) -> str:
    normalized_iobjtp = str(iobjtp or "").strip().upper()
    return RSDIOBJ_IOBJTP_TO_FIELD_TYPE.get(normalized_iobjtp, "F")


def fetch_rsiobj_type_lookup() -> Dict[str, str]:
    global _RSIOBJ_TYPE_LOOKUP_CACHE
    if _RSIOBJ_TYPE_LOOKUP_CACHE is not None:
        return _RSIOBJ_TYPE_LOOKUP_CACHE

    if not table_exists("rsdiobj"):
        return {}

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        try:
            cur.execute(
                """
                SELECT IOBJNM, IOBJTP
                FROM rsdiobj
                WHERE COALESCE(NULLIF(TRIM(OBJVERS), ''), 'A') = 'A'
                """
            )
        except mysql.connector.Error:
            cur.execute("SELECT IOBJNM, IOBJTP FROM rsdiobj")
        rows = cur.fetchall()
    except mysql.connector.Error:
        return {}
    finally:
        cur.close()
        conn.close()

    lookup: Dict[str, str] = {}
    for row in rows:
        name = normalize_field_name(row.get("IOBJNM"))
        if not name:
            continue
        lookup[name] = normalize_rsiobj_field_type_from_iobjtp(row.get("IOBJTP"))
    _RSIOBJ_TYPE_LOOKUP_CACHE = lookup
    return lookup


def is_rsiobj_backed_field_type(value: object) -> bool:
    return str(value or "").strip().upper() in RSDIOBJ_BACKED_FIELD_TYPES


def resolve_rsiobj_field_type(iobj_name: object, rsiobj_type_lookup: Dict[str, str]) -> str:
    normalized_iobj_name = normalize_field_name(iobj_name)
    if not normalized_iobj_name:
        return "F"
    return str(rsiobj_type_lookup.get(normalized_iobj_name) or "F").strip().upper() or "F"


def resolve_rsiobj_field_type_with_default_i(iobj_name: object, rsiobj_type_lookup: Dict[str, str]) -> str:
    normalized_iobj_name = normalize_field_name(iobj_name)
    if not normalized_iobj_name:
        return "F"
    resolved_field_type = resolve_rsiobj_field_type(normalized_iobj_name, rsiobj_type_lookup)
    return resolved_field_type if is_rsiobj_backed_field_type(resolved_field_type) else "I"


def build_bw_field_aliases(field_name: object) -> Set[str]:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return set()

    aliases = {normalized_field}
    if normalized_field.startswith("/BIC/"):
        aliases.add(normalized_field[5:])
    else:
        aliases.add(f"/BIC/{normalized_field}")
    if normalized_field.startswith("0") and len(normalized_field) > 1:
        aliases.add(normalized_field[1:])
    return {alias for alias in aliases if alias}


def merge_object_field_lookup_entry(
    lookup: Dict[str, Tuple[str, str, str]],
    alias: str,
    entry: Tuple[str, str, str],
    exact_aliases: Set[str],
) -> None:
    existing_entry = lookup.get(alias)
    if existing_entry is None:
        lookup[alias] = entry
        return

    existing_is_iobj = is_rsiobj_backed_field_type(existing_entry[2])
    new_is_iobj = is_rsiobj_backed_field_type(entry[2])
    if new_is_iobj and not existing_is_iobj:
        lookup[alias] = entry
        return
    if existing_is_iobj and not new_is_iobj:
        return
    if alias in exact_aliases:
        lookup[alias] = entry


def parse_adso_xml_field_lookup(
    xml_ui: object,
    rsiobj_type_lookup: Dict[str, str],
) -> Dict[str, Tuple[str, str, str]]:
    raw_xml = str(xml_ui or "").strip()
    if not raw_xml:
        return {}

    try:
        xml_bytes = bytes.fromhex(raw_xml)
    except (ValueError, binascii.Error):
        return {}

    try:
        root = ET.fromstring(xml_bytes.decode("utf-8", errors="replace"))
    except ET.ParseError:
        return {}

    lookup: Dict[str, Tuple[str, str, str]] = {}
    for element in root.iter():
        if not str(element.tag or "").endswith("element"):
            continue
        canonical_field = normalize_field_name(element.attrib.get("name"))
        if not canonical_field:
            continue
        iobj_name = normalize_field_name(element.attrib.get("infoObjectName"))
        field_type = resolve_rsiobj_field_type_with_default_i(iobj_name, rsiobj_type_lookup) if iobj_name else "F"
        entry = (canonical_field, iobj_name, field_type or "F")
        aliases = build_bw_field_aliases(canonical_field)
        if iobj_name:
            aliases.update(build_bw_field_aliases(iobj_name))
        exact_aliases = {alias for alias in {canonical_field, iobj_name} if alias}
        for alias in aliases:
            merge_object_field_lookup_entry(lookup, alias, entry, exact_aliases)
    return lookup


def fetch_adso_xml_field_lookup_for_objects(
    object_names: List[object],
    rsiobj_type_lookup: Dict[str, str] | None = None,
) -> Dict[str, Dict[str, Tuple[str, str, str]]]:
    normalized_objects: List[str] = []
    seen_objects: Set[str] = set()
    for object_name in object_names:
        normalized_object = normalize_bw_object_lookup(object_name)
        if not normalized_object or normalized_object in seen_objects:
            continue
        seen_objects.add(normalized_object)
        normalized_objects.append(normalized_object)

    if not normalized_objects:
        return {}

    resolved_rsiobj_type_lookup = rsiobj_type_lookup or fetch_rsiobj_type_lookup()
    lookup_by_object: Dict[str, Dict[str, Tuple[str, str, str]]] = {}
    uncached_objects: List[str] = []
    for normalized_object in normalized_objects:
        cached_lookup = _ADSO_XML_FIELD_LOOKUP_CACHE.get(normalized_object)
        if cached_lookup is None:
            uncached_objects.append(normalized_object)
            continue
        lookup_by_object[normalized_object] = cached_lookup

    if uncached_objects:
        raw_xml_by_object: Dict[str, object] = {}
        if table_exists("rsoadso"):
            placeholders = ", ".join(["%s"] * len(uncached_objects))
            conn = get_conn()
            cur = conn.cursor(dictionary=True)
            try:
                cur.execute(
                    f"""
                    SELECT ADSONM, XML_UI
                    FROM rsoadso
                    WHERE COALESCE(NULLIF(TRIM(OBJVERS), ''), 'A') = 'A'
                      AND UPPER(TRIM(ADSONM)) IN ({placeholders})
                    """,
                    tuple(uncached_objects),
                )
                rows = cur.fetchall()
            finally:
                cur.close()
                conn.close()

            for row in rows:
                normalized_object = normalize_bw_object_lookup(row.get("ADSONM"))
                if not normalized_object or normalized_object in raw_xml_by_object:
                    continue
                raw_xml_by_object[normalized_object] = row.get("XML_UI")

        for normalized_object in uncached_objects:
            lookup = parse_adso_xml_field_lookup(
                raw_xml_by_object.get(normalized_object),
                resolved_rsiobj_type_lookup,
            )
            _ADSO_XML_FIELD_LOOKUP_CACHE[normalized_object] = lookup
            lookup_by_object[normalized_object] = lookup

    return lookup_by_object


def fetch_adso_xml_field_lookup_for_object(
    object_name: object,
    rsiobj_type_lookup: Dict[str, str] | None = None,
) -> Dict[str, Tuple[str, str, str]]:
    normalized_object = normalize_bw_object_lookup(object_name)
    if not normalized_object:
        return {}
    return fetch_adso_xml_field_lookup_for_objects([normalized_object], rsiobj_type_lookup).get(normalized_object, {})


def resolve_adso_field_metadata(
    field_name: object,
    object_name: object,
    rsiobj_type_lookup: Dict[str, str] | None = None,
) -> Tuple[str, str, str] | None:
    normalized_field = normalize_field_name(field_name)
    normalized_object = normalize_bw_object_lookup(object_name)
    if not normalized_field or not normalized_object:
        return None
    field_lookup = fetch_adso_xml_field_lookup_for_object(normalized_object, rsiobj_type_lookup)
    return field_lookup.get(normalized_field)


def fetch_trcs_field_lookup_for_object(
    object_name: object,
    rsiobj_type_lookup: Dict[str, str] | None = None,
) -> Dict[str, Tuple[str, str, str]]:
    normalized_object = normalize_bw_object_lookup(object_name)
    if not normalized_object:
        return {}
    if normalized_object in _TRCS_FIELD_LOOKUP_CACHE:
        return _TRCS_FIELD_LOOKUP_CACHE[normalized_object]
    if not table_exists("rsksfieldnew"):
        _TRCS_FIELD_LOOKUP_CACHE[normalized_object] = {}
        return {}

    resolved_rsiobj_type_lookup = rsiobj_type_lookup or fetch_rsiobj_type_lookup()
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT FIELDNM, IOBJNM, SEGID, POSIT
            FROM rsksfieldnew
            WHERE COALESCE(NULLIF(TRIM(OBJVERS), ''), 'A') = 'A'
              AND UPPER(TRIM(ISOURCE)) = UPPER(TRIM(%s))
            ORDER BY CAST(COALESCE(NULLIF(TRIM(SEGID), ''), '0') AS UNSIGNED),
                     CAST(COALESCE(NULLIF(TRIM(POSIT), ''), '0') AS UNSIGNED),
                     FIELDNM,
                     IOBJNM
            """,
            (normalized_object,),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    lookup: Dict[str, Tuple[str, str, str]] = {}
    for row in rows:
        field_name = normalize_field_name(row.get("FIELDNM"))
        iobj_name = normalize_field_name(row.get("IOBJNM"))
        canonical_field = iobj_name or field_name
        if not canonical_field:
            continue
        field_type = resolve_rsiobj_field_type_with_default_i(iobj_name, resolved_rsiobj_type_lookup) if iobj_name else "F"
        entry = (canonical_field, iobj_name, field_type or "F")
        aliases = build_bw_field_aliases(field_name)
        if iobj_name:
            aliases.update(build_bw_field_aliases(iobj_name))
        exact_aliases = {alias for alias in {canonical_field, iobj_name} if alias}
        for alias in aliases:
            merge_object_field_lookup_entry(lookup, alias, entry, exact_aliases)

    _TRCS_FIELD_LOOKUP_CACHE[normalized_object] = lookup
    return lookup


def resolve_trcs_field_metadata(
    field_name: object,
    object_name: object,
    rsiobj_type_lookup: Dict[str, str] | None = None,
) -> Tuple[str, str, str] | None:
    normalized_field = normalize_field_name(field_name)
    normalized_object = normalize_bw_object_lookup(object_name)
    if not normalized_field or not normalized_object:
        return None
    field_lookup = fetch_trcs_field_lookup_for_object(normalized_object, rsiobj_type_lookup)
    return field_lookup.get(normalized_field)


def normalize_adso_field_name_with_rsiobj(
    field_name: object,
    rsiobj_names: Set[str],
    adso_field_lookup: Optional[Dict[str, Tuple[str, str, str]]] = None,
) -> Tuple[str, str]:
    normalized_field = str(field_name or "").strip().upper()
    if not normalized_field:
        return "", ""

    if adso_field_lookup:
        matched_entry = adso_field_lookup.get(normalized_field)
        if matched_entry:
            canonical_field, _iobj_name, field_type = matched_entry
            return canonical_field, field_type or "F"

    if normalized_field.startswith("/BIC/"):
        candidate = normalized_field[5:]
        if candidate in rsiobj_names:
            return candidate, resolve_rsiobj_field_type_with_default_i(candidate, fetch_rsiobj_type_lookup())
        # Not an InfoObject: return original DD03L field name (without /BIC/ prefix)
        return candidate, "F"

    iobj_candidate = f"0{normalized_field}"
    if iobj_candidate in rsiobj_names:
        return iobj_candidate, resolve_rsiobj_field_type_with_default_i(iobj_candidate, fetch_rsiobj_type_lookup())
    if normalized_field in rsiobj_names:
        return normalized_field, resolve_rsiobj_field_type_with_default_i(normalized_field, fetch_rsiobj_type_lookup())
    # Not an InfoObject: return original DD03L field name
    return normalized_field, "F"


def infer_field_metadata_with_rsiobj(
    field_name: object,
    object_type: object,
    rsiobj_names: Set[str],
    rsiobj_type_lookup: Dict[str, str] | None = None,
    object_name: object = "",
) -> Tuple[str, str, str]:
    normalized_field = normalize_field_name(field_name)
    normalized_type = normalize_type_code(object_type)
    resolved_rsiobj_type_lookup = rsiobj_type_lookup or {}
    if not normalized_field:
        return "", "", ""

    if normalized_type == "ADSO":
        matched_adso_field = resolve_adso_field_metadata(normalized_field, object_name, resolved_rsiobj_type_lookup)
        if matched_adso_field:
            return matched_adso_field

    if normalized_type == "TRCS":
        matched_trcs_field = resolve_trcs_field_metadata(normalized_field, object_name, resolved_rsiobj_type_lookup)
        if matched_trcs_field:
            return matched_trcs_field

    if normalized_type in {"ADSO", "ODSO", "IOBJ"}:
        adso_field_lookup = (
            fetch_adso_xml_field_lookup_for_object(object_name, resolved_rsiobj_type_lookup)
            if normalized_type == "ADSO" and normalize_bw_object_lookup(object_name)
            else None
        )
        canonical_field, matched_field_type = normalize_adso_field_name_with_rsiobj(
            normalized_field,
            rsiobj_names,
            adso_field_lookup,
        )
        iobj_name = canonical_field if is_rsiobj_backed_field_type(matched_field_type) and canonical_field in rsiobj_names else ""
        field_type = matched_field_type or (resolve_rsiobj_field_type(iobj_name, resolved_rsiobj_type_lookup) if iobj_name else "F")
        return canonical_field, iobj_name, field_type or "F"

    if normalized_field in rsiobj_names:
        return normalized_field, normalized_field, resolve_rsiobj_field_type(normalized_field, resolved_rsiobj_type_lookup)

    if not normalized_field.startswith("0"):
        iobj_candidate = f"0{normalized_field}"
        if iobj_candidate in rsiobj_names:
            return normalized_field, iobj_candidate, resolve_rsiobj_field_type(iobj_candidate, resolved_rsiobj_type_lookup)

    if normalized_field.startswith("/BIC/"):
        candidate = normalized_field[5:]
        if candidate in rsiobj_names:
            return candidate, candidate, resolve_rsiobj_field_type(candidate, resolved_rsiobj_type_lookup)
        return candidate, "", "F"

    return normalized_field, "", "F"


def canonicalize_object_field_for_matching(
    field_name: object,
    object_type: object,
    rsiobj_names: Set[str] | None = None,
    object_name: object = "",
) -> str:
    normalized_field = normalize_field_name(field_name)
    normalized_type = normalize_type_code(object_type)
    if not normalized_field:
        return ""
    if normalized_type == "ADSO":
        matched_adso_field = resolve_adso_field_metadata(normalized_field, object_name)
        if matched_adso_field:
            return matched_adso_field[0]
    if normalized_type == "TRCS":
        matched_trcs_field = resolve_trcs_field_metadata(normalized_field, object_name)
        if matched_trcs_field:
            return matched_trcs_field[0]
    if normalized_type in {"ADSO", "ODSO"} and rsiobj_names is not None:
        canonical_field, _field_type = normalize_adso_field_name_with_rsiobj(normalized_field, rsiobj_names)
        return canonical_field
    return normalized_field


def build_dd03l_object_inventory(
    rows: List[Dict[str, object]],
    rsiobj_names: Set[str] | None = None,
    adso_field_lookup: Optional[Dict[str, Tuple[str, str, str]]] = None,
) -> Dict[str, Dict[str, object]]:
    seg_inventory = build_empty_segment_inventory()
    resolved_rsiobj_names = rsiobj_names or set()
    use_rsiobj_matching = rsiobj_names is not None
    seen_fields: Set[str] = set()
    for row in rows:
        if use_rsiobj_matching:
            field_name, field_type = normalize_adso_field_name_with_rsiobj(
                row.get("FIELDNAME"),
                resolved_rsiobj_names,
                adso_field_lookup,
            )
            if not field_name:
                continue
        else:
            field_name = derive_adso_field_name_from_dd03l(row.get("FIELDNAME"))
            field_type = "F"
            if not is_dd03l_adso_business_field(field_name):
                continue
        if field_name in seen_fields:
            continue
        seen_fields.add(field_name)
        keyflag = str(row.get("KEYFLAG") or "").strip()
        seg_inventory["ordered"].append((field_name, keyflag))
        seg_inventory["field_set"].add(field_name)
        seg_inventory["key_by_field"][field_name] = keyflag
        seg_inventory["type_by_field"][field_name] = field_type or "F"
        seg_inventory["metadata_by_field"][field_name] = {
            "datatype": str(row.get("DATATYPE") or "").strip().upper(),
            "length": _normalize_inventory_numeric_value(row.get("LENG")),
            "decimals": _normalize_inventory_numeric_value(row.get("DECIMALS")),
        }
    return {"": seg_inventory} if seg_inventory["ordered"] else {}


def _normalize_inventory_numeric_value(value: object) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def fetch_rsoadso_activate_data_by_object(object_names: List[str]) -> Dict[str, str]:
    normalized_objects = []
    seen_objects: Set[str] = set()
    for object_name in object_names:
        normalized_name = normalize_bw_object_lookup(object_name)
        if not normalized_name or normalized_name in seen_objects:
            continue
        seen_objects.add(normalized_name)
        normalized_objects.append(normalized_name)

    if not normalized_objects or not table_exists("rsoadso"):
        return {}

    placeholders = ", ".join(["%s"] * len(normalized_objects))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT ADSONM, ACTIVATE_DATA
            FROM rsoadso
            WHERE OBJVERS = 'A'
                            AND ADSONM IN ({placeholders})
            """,
                        tuple(normalized_objects),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    activate_data_by_object: Dict[str, str] = {}
    for row in rows:
        object_name = normalize_bw_object_lookup(row.get("ADSONM"))
        if not object_name:
            continue
        activate_data_by_object[object_name] = str(row.get("ACTIVATE_DATA") or "").strip().upper()
    return activate_data_by_object


def build_dd03l_object_tabname_candidates(
    object_names: List[str],
    suffixes_by_object: Dict[str, Set[str]] | None = None,
) -> Dict[str, List[str]]:
    candidates_by_object: Dict[str, List[str]] = {}
    resolved_suffixes_by_object = suffixes_by_object or {}
    for object_name in object_names:
        normalized_name = normalize_bw_object_lookup(object_name)
        if not normalized_name:
            continue
        suffixes = resolved_suffixes_by_object.get(normalized_name) or {"1", "2"}
        table_names: List[str] = []
        seen: Set[str] = set()
        for suffix in suffixes:
            normalized_suffix = str(suffix or "").strip()
            if normalized_suffix not in {"1", "2"}:
                continue
            table_name = f"/BIC/A{normalized_name}{normalized_suffix}"
            if table_name in seen:
                continue
            seen.add(table_name)
            table_names.append(table_name)
        if table_names:
            candidates_by_object[normalized_name] = table_names
    return candidates_by_object


def fetch_dd03l_rows_for_tabnames(tabnames: List[str]) -> List[Dict[str, object]]:
    normalized_tabnames = []
    seen_tabnames: Set[str] = set()
    for tabname in tabnames:
        normalized_tabname = str(tabname or "").strip().upper()
        if not normalized_tabname or normalized_tabname in seen_tabnames:
            continue
        seen_tabnames.add(normalized_tabname)
        normalized_tabnames.append(normalized_tabname)

    if not normalized_tabnames or not table_exists("dd03l"):
        return []

    placeholders = ", ".join(["%s"] * len(normalized_tabnames))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
                        SELECT TABNAME, POSITION, FIELDNAME, KEYFLAG, DATATYPE, LENG, DECIMALS
            FROM dd03l
            WHERE AS4LOCAL = 'A'
              AND TABNAME IN ({placeholders})
            ORDER BY TABNAME,
                     CAST(COALESCE(NULLIF(TRIM(POSITION), ''), '0') AS UNSIGNED),
                     FIELDNAME
            """,
            tuple(normalized_tabnames),
        )
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()


def fetch_adso_field_inventory_from_dd03l(
    object_names: List[str],
    activate_data_by_object: Dict[str, str] | None = None,
    rsiobj_names: Set[str] | None = None,
) -> Dict[str, Dict[str, Dict[str, object]]]:
    if not table_exists("dd03l"):
        return {}

    normalized_objects = []
    seen_objects: Set[str] = set()

    for object_name in object_names:
        normalized_name = normalize_bw_object_lookup(object_name)
        if not normalized_name or normalized_name in seen_objects:
            continue
        seen_objects.add(normalized_name)
        normalized_objects.append(normalized_name)
    if not normalized_objects:
        return {}

    resolved_activate_data = activate_data_by_object or fetch_rsoadso_activate_data_by_object(normalized_objects)
    resolved_rsiobj_type_lookup = fetch_rsiobj_type_lookup()
    resolved_rsiobj_names = rsiobj_names if rsiobj_names is not None else set(resolved_rsiobj_type_lookup.keys())
    suffixes_by_object = {
        object_name: {"2" if resolved_activate_data.get(object_name, "") == "X" else "1"}
        for object_name in normalized_objects
    }
    candidate_tables = build_dd03l_object_tabname_candidates(normalized_objects, suffixes_by_object)
    rows = fetch_dd03l_rows_for_tabnames([table for table_names in candidate_tables.values() for table in table_names])

    adso_field_lookup_by_object = fetch_adso_xml_field_lookup_for_objects(
        normalized_objects,
        resolved_rsiobj_type_lookup,
    )

    inventory: Dict[str, Dict[str, Dict[str, object]]] = {}
    table_to_object = {
        table_name: object_name
        for object_name, table_names in candidate_tables.items()
        for table_name in table_names
    }
    inventory_by_object: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        table_name = str(row.get("TABNAME") or "").strip().upper()
        object_name = table_to_object.get(table_name)
        if not object_name:
            continue
        inventory_by_object[object_name].append(row)

    for object_name in normalized_objects:
        object_inventory = build_dd03l_object_inventory(
            inventory_by_object.get(object_name, []),
            resolved_rsiobj_names,
            adso_field_lookup_by_object.get(object_name, {}),
        )
        if not object_inventory:
            continue
        inventory[object_name] = object_inventory

    return inventory


def fetch_odso_field_inventory_from_dd03l(
    object_names: List[str],
    rsiobj_names: Set[str] | None = None,
) -> Dict[str, Dict[str, Dict[str, object]]]:
    if not table_exists("dd03l"):
        return {}

    normalized_objects = []
    seen_objects: Set[str] = set()

    for object_name in object_names:
        normalized_name = normalize_bw_object_lookup(object_name)
        if not normalized_name or normalized_name in seen_objects:
            continue
        seen_objects.add(normalized_name)
        normalized_objects.append(normalized_name)
    if not normalized_objects:
        return {}

    resolved_rsiobj_names = rsiobj_names if rsiobj_names is not None else fetch_rsiobj_name_set()
    candidate_tables = build_dd03l_object_tabname_candidates(normalized_objects)
    rows = fetch_dd03l_rows_for_tabnames([table for table_names in candidate_tables.values() for table in table_names])

    inventory: Dict[str, Dict[str, Dict[str, object]]] = {}
    table_to_object = {
        table_name: object_name
        for object_name, table_names in candidate_tables.items()
        for table_name in table_names
    }
    inventory_by_object: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        table_name = str(row.get("TABNAME") or "").strip().upper()
        object_name = table_to_object.get(table_name)
        if not object_name:
            continue
        inventory_by_object[object_name].append(row)

    for object_name in normalized_objects:
        object_inventory = build_dd03l_object_inventory(
            inventory_by_object.get(object_name, []),
            resolved_rsiobj_names,
        )
        if not object_inventory:
            continue
        inventory[object_name] = object_inventory

    return inventory


def fetch_trcs_field_inventory(object_names: List[str]) -> Dict[str, Dict[str, Dict[str, object]]]:
    if not table_exists("rsksfieldnew"):
        return {}

    normalized_objects = []
    seen_objects: Set[str] = set()
    for object_name in object_names:
        normalized_name = normalize_bw_object_lookup(object_name)
        if not normalized_name or normalized_name in seen_objects:
            continue
        seen_objects.add(normalized_name)
        normalized_objects.append(normalized_name)

    if not normalized_objects:
        return {}

    resolved_rsiobj_type_lookup = fetch_rsiobj_type_lookup()
    placeholders = ", ".join(["%s"] * len(normalized_objects))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT ISOURCE, SEGID, POSIT, TYPE, FIELDNM, IOBJNM, KEYFLAG, DATATYPE, LENG, DECIMALS
            FROM rsksfieldnew
            WHERE OBJVERS = 'A' AND ISOURCE IN ({placeholders})
            ORDER BY ISOURCE,
                     CAST(COALESCE(NULLIF(TRIM(SEGID), ''), '0') AS UNSIGNED),
                     CAST(COALESCE(NULLIF(TRIM(POSIT), ''), '0') AS UNSIGNED),
                     FIELDNM,
                     IOBJNM
            """,
            tuple(normalized_objects),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    inventory: Dict[str, Dict[str, Dict[str, object]]] = {}
    lookup_by_object: Dict[str, Dict[str, Tuple[str, str, str]]] = defaultdict(dict)
    seen: Set[Tuple[str, str, str]] = set()
    for row in rows:
        isource = normalize_bw_object_lookup(row.get("ISOURCE"))
        seg_id = str(row.get("SEGID") or "").strip()
        iobj_name = normalize_field_name(row.get("IOBJNM"))
        field_name_raw = row.get("FIELDNM")
        field_name = iobj_name or normalize_field_name(field_name_raw)
        if not isource or not field_name:
            continue
        dedup_key = (isource, seg_id, field_name)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        object_inventory = inventory.setdefault(isource, {})
        seg_inventory = object_inventory.setdefault(
            seg_id,
            build_empty_segment_inventory(),
        )
        keyflag = str(row.get("KEYFLAG") or "").strip()
        field_type = "I" if iobj_name else "F"
        seg_inventory["ordered"].append((field_name, keyflag))
        seg_inventory["field_set"].add(field_name)
        seg_inventory["key_by_field"][field_name] = keyflag
        seg_inventory["type_by_field"][field_name] = field_type
        seg_inventory["metadata_by_field"][field_name] = {
            "datatype": str(row.get("DATATYPE") or "").strip().upper(),
            "length": _normalize_inventory_numeric_value(row.get("LENG")),
            "decimals": _normalize_inventory_numeric_value(row.get("DECIMALS")),
        }

        entry = (
            field_name,
            iobj_name,
            resolve_rsiobj_field_type_with_default_i(iobj_name, resolved_rsiobj_type_lookup) if iobj_name else "F",
        )
        aliases = build_bw_field_aliases(row.get("FIELDNM"))
        if iobj_name:
            aliases.update(build_bw_field_aliases(iobj_name))
        exact_aliases = {alias for alias in {field_name, iobj_name} if alias}
        object_lookup = lookup_by_object[isource]
        for alias in aliases:
            merge_object_field_lookup_entry(object_lookup, alias, entry, exact_aliases)

    for object_name in normalized_objects:
        _TRCS_FIELD_LOOKUP_CACHE[object_name] = lookup_by_object.get(object_name, {})

    return inventory


def build_iobj_dd03l_table_candidates(object_name: str) -> List[str]:
    return build_iobj_dd03l_table_candidates_for_subtype(object_name, "")


def build_iobj_dd03l_table_candidates_for_subtype(object_name: str, subtype: object = "") -> List[str]:
    normalized_name = normalize_bw_object_lookup(object_name)
    if not normalized_name:
        return []

    normalized_subtype = normalize_iobj_subtype(subtype)
    if normalized_name.startswith("0"):
        base_name = normalized_name[1:]
        prefix = "/BI0/"
    else:
        base_name = normalized_name
        prefix = "/BIC/"

    if normalized_subtype == "ATTR":
        suffixes = ("P",)
    elif normalized_subtype == "TEXT":
        suffixes = ("T",)
    else:
        suffixes = ("P", "T", "H")

    candidates: List[str] = []
    seen: Set[str] = set()
    for suffix in suffixes:
        table_name = f"{prefix}{suffix}{base_name}"
        if table_name not in seen:
            seen.add(table_name)
            candidates.append(table_name)
    return candidates


def normalize_iobj_dd03l_field_name(field_name: object) -> str:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return ""
    if normalized_field.startswith("/BIC/"):
        return normalized_field[len("/BIC/"):]
    if normalized_field in {"OBJVERS", "CHANGED"}:
        return ""
    if normalized_field == "TXTLG":
        return "0TXTXL"
    return f"0{normalized_field}"


def fetch_iobj_field_inventory_from_dd03l(object_entries: List[object]) -> Dict[Tuple[str, str], Dict[str, Dict[str, object]]]:
    if not table_exists("dd03l"):
        return {}

    normalized_entries: List[Tuple[str, str]] = []
    seen_entries: Set[Tuple[str, str]] = set()
    all_candidate_tables: List[str] = []
    seen_tables: Set[str] = set()
    for entry in object_entries:
        if isinstance(entry, (tuple, list)):
            object_name = entry[0] if len(entry) > 0 else ""
            object_subtype = entry[1] if len(entry) > 1 else ""
        else:
            object_name = entry
            object_subtype = ""
        normalized_name = normalize_bw_object_lookup(object_name)
        normalized_subtype = normalize_iobj_subtype(object_subtype)
        key = (normalized_name, normalized_subtype)
        if not normalized_name or key in seen_entries:
            continue
        seen_entries.add(key)
        normalized_entries.append(key)
        for table_name in build_iobj_dd03l_table_candidates_for_subtype(normalized_name, normalized_subtype):
            if table_name not in seen_tables:
                seen_tables.add(table_name)
                all_candidate_tables.append(table_name)

    if not all_candidate_tables:
        return {}

    placeholders = ", ".join(["%s"] * len(all_candidate_tables))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT TABNAME, POSITION, FIELDNAME, KEYFLAG, DATATYPE, LENG, DECIMALS
            FROM dd03l
            WHERE TABNAME IN ({placeholders})
            ORDER BY TABNAME,
                     CAST(COALESCE(NULLIF(TRIM(POSITION), ''), '0') AS UNSIGNED),
                     FIELDNAME
            """,
            tuple(all_candidate_tables),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    rows_by_table: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        table_name = str(row.get("TABNAME") or "").strip().upper()
        if table_name:
            rows_by_table[table_name].append(row)

    inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]] = {}
    for object_name, object_subtype in normalized_entries:
        seg_inventory = build_empty_segment_inventory()
        seen_fields: Set[str] = set()
        for table_name in build_iobj_dd03l_table_candidates_for_subtype(object_name, object_subtype):
            for row in rows_by_table.get(table_name, []):
                field_name = normalize_iobj_dd03l_field_name(row.get("FIELDNAME"))
                if not field_name or field_name in seen_fields:
                    continue
                seen_fields.add(field_name)
                keyflag = str(row.get("KEYFLAG") or "").strip()
                seg_inventory["ordered"].append((field_name, keyflag))
                seg_inventory["field_set"].add(field_name)
                seg_inventory["key_by_field"][field_name] = keyflag
                seg_inventory["type_by_field"][field_name] = "I"
                seg_inventory["metadata_by_field"][field_name] = {
                    "datatype": str(row.get("DATATYPE") or "").strip().upper(),
                    "length": _normalize_inventory_numeric_value(row.get("LENG")),
                    "decimals": _normalize_inventory_numeric_value(row.get("DECIMALS")),
                }
        if seg_inventory["ordered"]:
            inventory[(object_name, object_subtype)] = {"": seg_inventory}

    return inventory


def get_supported_inventory_for_object(
    object_name: str,
    object_system: str,
    object_type: str,
    object_subtype: str,
    rsds_inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]],
    dd03l_inventory: Dict[str, Dict[str, Dict[str, object]]],
    trcs_inventory: Dict[str, Dict[str, Dict[str, object]]],
    iobj_inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]],
) -> Tuple[Dict[str, Dict[str, object]], bool, str]:
    normalized_type = normalize_type_code(object_type)
    if normalized_type == "RSDS":
        key = (normalize_bw_object_lookup(object_name), normalize_bw_object_lookup(object_system))
        inventory = rsds_inventory.get(key, {})
        return inventory, bool(inventory), "rsdssegfd"
    if normalized_type in {"ADSO", "ODSO"}:
        inventory = dd03l_inventory.get(normalize_bw_object_lookup(object_name), {})
        return inventory, bool(inventory), "rsoadso.xml_ui" if normalized_type == "ADSO" else "dd03l"
    if normalized_type == "TRCS":
        inventory = trcs_inventory.get(normalize_bw_object_lookup(object_name), {})
        return inventory, bool(inventory), "rsksfieldnew"
    if normalized_type == "IOBJ":
        normalized_object = normalize_bw_object_lookup(object_name)
        normalized_subtype = normalize_iobj_subtype(object_subtype)
        inventory = iobj_inventory.get((normalized_object, normalized_subtype), {})
        if not inventory and normalized_subtype:
            inventory = iobj_inventory.get((normalized_object, ""), {})
        return inventory, bool(inventory), "dd03l"
    return {}, False, ""


def inventory_contains_field(inventory_by_seg: Dict[str, Dict[str, object]], seg_id: str, field_name: str) -> bool:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return False
    seg_key = str(seg_id or "").strip()
    seg_inventory = inventory_by_seg.get(seg_key) or inventory_by_seg.get("") or {}
    return normalized_field in (seg_inventory.get("field_set") or set())


def inventory_keyflag(inventory_by_seg: Dict[str, Dict[str, object]], seg_id: str, field_name: str) -> str:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return ""
    seg_key = str(seg_id or "").strip()
    seg_inventory = inventory_by_seg.get(seg_key) or inventory_by_seg.get("") or {}
    return str((seg_inventory.get("key_by_field") or {}).get(normalized_field) or "").strip()


def inventory_field_type(inventory_by_seg: Dict[str, Dict[str, object]], seg_id: str, field_name: str) -> str:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return ""
    seg_key = str(seg_id or "").strip()
    seg_inventory = inventory_by_seg.get(seg_key) or inventory_by_seg.get("") or {}
    return str((seg_inventory.get("type_by_field") or {}).get(normalized_field) or "").strip().upper()


def build_inventory_from_materialized_rows(rows: List[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    inventory: Dict[str, Dict[str, object]] = {}
    for row in rows:
        seg_id = str(row.get("seg_id") or "").strip()
        field_name = normalize_field_name(row.get("field_name"))
        if not field_name:
            continue
        seg_inventory = inventory.setdefault(seg_id, build_empty_segment_inventory())
        if field_name in (seg_inventory.get("field_set") or set()):
            continue
        keyflag = str(row.get("keyflag") or "").strip()
        seg_inventory["ordered"].append((field_name, keyflag))
        seg_inventory["field_set"].add(field_name)
        seg_inventory["key_by_field"][field_name] = keyflag
        seg_inventory["type_by_field"][field_name] = str(row.get("field_type") or "").strip().upper() or "F"
        seg_inventory["metadata_by_field"][field_name] = {
            "datatype": str(row.get("datatype") or "").strip().upper(),
            "length": _normalize_inventory_numeric_value(row.get("length")),
            "decimals": _normalize_inventory_numeric_value(row.get("decimals")),
        }
    return inventory


def fetch_materialized_object_field_inventory_rows(
    object_type: str,
    object_names: List[str] | None = None,
    object_entries: List[Tuple[str, str]] | None = None,
) -> List[Dict[str, object]]:
    normalized_type = normalize_type_code(object_type)
    if not normalized_type or not table_exists(BW_OBJECT_FIELD_INVENTORY_TABLE):
        return []

    filters: List[object] = [normalized_type]
    where_sql = "WHERE object_type = %s"

    if object_entries is not None:
        normalized_entries: List[Tuple[str, str]] = []
        seen_entries: Set[Tuple[str, str]] = set()
        for object_name, object_subtype in object_entries:
            entry = (normalize_bw_object_lookup(object_name), normalize_iobj_subtype(object_subtype))
            if not entry[0] or entry in seen_entries:
                continue
            seen_entries.add(entry)
            normalized_entries.append(entry)
        if not normalized_entries:
            return []
        where_sql += " AND (" + " OR ".join(["(object_name = %s AND object_subtype = %s)"] * len(normalized_entries)) + ")"
        for object_name, object_subtype in normalized_entries:
            filters.extend([object_name, object_subtype])
    else:
        normalized_objects: List[str] = []
        seen_objects: Set[str] = set()
        for object_name in object_names or []:
            normalized_name = normalize_bw_object_lookup(object_name)
            if not normalized_name or normalized_name in seen_objects:
                continue
            seen_objects.add(normalized_name)
            normalized_objects.append(normalized_name)
        if not normalized_objects:
            return []
        placeholders = ", ".join(["%s"] * len(normalized_objects))
        where_sql += f" AND object_name IN ({placeholders})"
        filters.extend(normalized_objects)

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT object_name, object_subtype, seg_id, field_name, field_type, keyflag, datatype, length, decimals, field_order
            FROM `{BW_OBJECT_FIELD_INVENTORY_TABLE}`
            {where_sql}
            ORDER BY object_name, object_subtype, seg_id, field_order, field_name
            """,
            tuple(filters),
        )
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()


def fetch_materialized_dd03l_field_inventory(object_type: str, object_names: List[str]) -> Dict[str, Dict[str, Dict[str, object]]]:
    rows = fetch_materialized_object_field_inventory_rows(object_type, object_names=object_names)
    inventory: Dict[str, Dict[str, Dict[str, object]]] = {}
    rows_by_object: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        object_name = normalize_bw_object_lookup(row.get("object_name"))
        if object_name:
            rows_by_object[object_name].append(row)
    for object_name, object_rows in rows_by_object.items():
        object_inventory = build_inventory_from_materialized_rows(object_rows)
        if object_inventory:
            inventory[object_name] = object_inventory
    return inventory


def fetch_materialized_trcs_field_inventory(object_names: List[str]) -> Dict[str, Dict[str, Dict[str, object]]]:
    return fetch_materialized_dd03l_field_inventory("TRCS", object_names)


def fetch_materialized_iobj_field_inventory(object_entries: List[object]) -> Dict[Tuple[str, str], Dict[str, Dict[str, object]]]:
    normalized_entries: List[Tuple[str, str]] = []
    for entry in object_entries:
        if isinstance(entry, (tuple, list)):
            object_name = entry[0] if len(entry) > 0 else ""
            object_subtype = entry[1] if len(entry) > 1 else ""
        else:
            object_name = entry
            object_subtype = ""
        normalized_entries.append((normalize_bw_object_lookup(object_name), normalize_iobj_subtype(object_subtype)))

    rows = fetch_materialized_object_field_inventory_rows("IOBJ", object_entries=normalized_entries)
    inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]] = {}
    rows_by_object: Dict[Tuple[str, str], List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        object_name = normalize_bw_object_lookup(row.get("object_name"))
        object_subtype = normalize_iobj_subtype(row.get("object_subtype"))
        if object_name:
            rows_by_object[(object_name, object_subtype)].append(row)
    for object_key, object_rows in rows_by_object.items():
        object_inventory = build_inventory_from_materialized_rows(object_rows)
        if object_inventory:
            inventory[object_key] = object_inventory
    return inventory


def append_materialized_inventory_rows(
    insert_rows: List[Tuple[object, ...]],
    object_type: str,
    object_name: str,
    object_subtype: str,
    inventory_by_seg: Dict[str, Dict[str, object]],
    source_origin: str,
) -> None:
    normalized_type = normalize_type_code(object_type)
    normalized_name = normalize_bw_object_lookup(object_name)
    normalized_subtype = normalize_iobj_subtype(object_subtype) if normalized_type == "IOBJ" else ""
    if not normalized_type or not normalized_name:
        return

    for seg_id, seg_inventory in inventory_by_seg.items():
        metadata_by_field = seg_inventory.get("metadata_by_field") or {}
        type_by_field = seg_inventory.get("type_by_field") or {}
        for field_order, (field_name, keyflag) in enumerate(seg_inventory.get("ordered") or [], start=1):
            metadata = metadata_by_field.get(field_name) or {}
            insert_rows.append(
                (
                    normalized_type,
                    normalized_name,
                    normalized_subtype,
                    str(seg_id or "").strip(),
                    normalize_field_name(field_name),
                    str(type_by_field.get(field_name) or "F").strip().upper() or "F",
                    str(keyflag or "").strip(),
                    str(metadata.get("datatype") or "").strip().upper(),
                    metadata.get("length"),
                    metadata.get("decimals"),
                    field_order,
                    source_origin,
                )
            )


def rebuild_bw_object_field_inventory_table(
    tran_metadata: Optional[Dict[str, Dict[str, object]]] = None,
) -> Dict[str, int]:
    ensure_bw_object_field_inventory_table()
    resolved_tran_metadata = tran_metadata or fetch_active_tran_metadata()

    adso_objects = sorted(
        {
            normalize_bw_object_lookup(meta.get("source_object", ""))
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "ADSO"
        }
        | {
            normalize_bw_object_lookup(meta.get("target_object", ""))
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "ADSO"
        }
    )
    odso_objects = sorted(
        {
            normalize_bw_object_lookup(meta.get("source_object", ""))
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "ODSO"
        }
        | {
            normalize_bw_object_lookup(meta.get("target_object", ""))
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "ODSO"
        }
    )
    trcs_objects = sorted(
        {
            normalize_bw_object_lookup(meta.get("source_object", ""))
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "TRCS"
        }
        | {
            normalize_bw_object_lookup(meta.get("target_object", ""))
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "TRCS"
        }
    )
    iobj_objects = sorted(
        {
            (
                normalize_bw_object_lookup(meta.get("source_object", "")),
                normalize_iobj_subtype(meta.get("source_subtype", "")),
            )
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "IOBJ"
            and normalize_bw_object_lookup(meta.get("source_object", ""))
        }
        | {
            (
                normalize_bw_object_lookup(meta.get("target_object", "")),
                normalize_iobj_subtype(meta.get("target_subtype", "")),
            )
            for meta in resolved_tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "IOBJ"
            and normalize_bw_object_lookup(meta.get("target_object", ""))
        }
    )

    activate_data_by_object = fetch_rsoadso_activate_data_by_object(adso_objects) if adso_objects else {}
    rsiobj_names = fetch_rsiobj_name_set() if (adso_objects or odso_objects) else set()
    odso_inventory = fetch_odso_field_inventory_from_dd03l(odso_objects, rsiobj_names=rsiobj_names) if odso_objects else {}
    adso_inventory = fetch_adso_field_inventory_from_dd03l(
        adso_objects,
        activate_data_by_object=activate_data_by_object,
        rsiobj_names=rsiobj_names,
    ) if adso_objects else {}
    trcs_inventory = fetch_trcs_field_inventory(trcs_objects) if trcs_objects else {}
    iobj_inventory = fetch_iobj_field_inventory_from_dd03l(iobj_objects) if iobj_objects else {}

    insert_rows: List[Tuple[object, ...]] = []
    for object_name, inventory in odso_inventory.items():
        append_materialized_inventory_rows(insert_rows, "ODSO", object_name, "", inventory, "dd03l")
    for object_name, inventory in adso_inventory.items():
        append_materialized_inventory_rows(insert_rows, "ADSO", object_name, "", inventory, "rsoadso.xml_ui")
    for object_name, inventory in trcs_inventory.items():
        append_materialized_inventory_rows(insert_rows, "TRCS", object_name, "", inventory, "rsksfieldnew")
    for (object_name, object_subtype), inventory in iobj_inventory.items():
        append_materialized_inventory_rows(insert_rows, "IOBJ", object_name, object_subtype, inventory, "dd03l")

    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(f"DELETE FROM `{BW_OBJECT_FIELD_INVENTORY_TABLE}`")
        else:
            cur.execute(f"TRUNCATE TABLE `{BW_OBJECT_FIELD_INVENTORY_TABLE}`")
        if insert_rows:
            cur.executemany(
                f"""
                INSERT INTO `{BW_OBJECT_FIELD_INVENTORY_TABLE}` (
                    object_type,
                    object_name,
                    object_subtype,
                    seg_id,
                    field_name,
                    field_type,
                    keyflag,
                    datatype,
                    length,
                    decimals,
                    field_order,
                    source_origin
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                insert_rows,
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return {
        "inserted_rows": len(insert_rows),
        "object_count": len(odso_inventory) + len(adso_inventory) + len(trcs_inventory) + len(iobj_inventory),
        "odso_count": len(odso_inventory),
        "adso_count": len(adso_inventory),
        "trcs_count": len(trcs_inventory),
        "iobj_count": len(iobj_inventory),
    }


def inventory_field_metadata(
    inventory_by_seg: Dict[str, Dict[str, object]],
    seg_id: str,
    field_name: str,
) -> Dict[str, object]:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return {"datatype": "", "length": None, "decimals": None}
    seg_key = str(seg_id or "").strip()
    seg_inventory = inventory_by_seg.get(seg_key) or inventory_by_seg.get("") or {}
    metadata = (seg_inventory.get("metadata_by_field") or {}).get(normalized_field) or {}
    return {
        "datatype": str(metadata.get("datatype") or "").strip().upper(),
        "length": metadata.get("length"),
        "decimals": metadata.get("decimals"),
    }


def enrich_mapping_rows_with_inventory_field_metadata(
    rows: List[Dict[str, object]],
    source_inventory: Dict[str, Dict[str, object]],
    source_type: str,
    target_inventory: Dict[str, Dict[str, object]],
    target_type: str,
    rsiobj_names: Set[str] | None = None,
    source_object: str = "",
    target_object: str = "",
) -> List[Dict[str, object]]:
    if not rows:
        return rows

    for row in rows:
        seg_id = str(row.get("seg_id") or "").strip()
        source_meta = inventory_field_metadata(
            source_inventory,
            seg_id,
            mapping_row_match_field(row, "source_field"),
        )
        target_meta = inventory_field_metadata(
            target_inventory,
            seg_id,
            mapping_row_match_field(row, "target_field"),
        )
        row["source_datatype"] = source_meta.get("datatype")
        row["source_length"] = source_meta.get("length")
        row["source_decimals"] = source_meta.get("decimals")
        row["target_datatype"] = target_meta.get("datatype")
        row["target_length"] = target_meta.get("length")
        row["target_decimals"] = target_meta.get("decimals")

    return rows


def inventory_field_already_mapped(mapped_keys: Set[Tuple[str, str]], seg_id: str, field_name: str) -> bool:
    normalized_field = normalize_field_name(field_name)
    if not normalized_field:
        return False
    normalized_seg_id = str(seg_id or "").strip()
    if (normalized_seg_id, normalized_field) in mapped_keys:
        return True
    if not normalized_seg_id:
        return any(mapped_field == normalized_field for _, mapped_field in mapped_keys)
    return False


def inventory_unique_field_name_set(inventory_by_seg: Dict[str, Dict[str, object]]) -> Set[str]:
    field_names: Set[str] = set()
    for seg_inventory in inventory_by_seg.values():
        field_names.update(seg_inventory.get("field_set") or set())
    return field_names


def rows_unique_field_name_set(
    rows: List[Dict[str, object]],
    field_key: str,
) -> Set[str]:
    field_names: Set[str] = set()
    for row in rows:
        field_name = mapping_row_match_field(row, field_key)
        if field_name:
            field_names.add(field_name)
    return field_names


def build_inventory_reference_meta(
    object_name: str,
    object_type: str,
    object_subtype: str,
    inventory_origin: str,
    activate_data_by_object: Dict[str, str],
) -> Dict[str, str]:
    normalized_object = normalize_bw_object_lookup(object_name)
    normalized_type = normalize_type_code(object_type)
    origin = str(inventory_origin or "").strip().lower()
    result = {
        "inventory_origin": origin,
        "inventory_label": origin.upper() if origin else "",
        "adso_table_suffix": "",
        "adso_table_name": "",
    }

    if origin in {"dd03l", "rsoadso.xml_ui"} and normalized_type == "ADSO" and normalized_object:
        suffix = "2" if activate_data_by_object.get(normalized_object, "") == "X" else "1"
        result["inventory_label"] = "RSOADSO XML_UI" if origin == "rsoadso.xml_ui" else f"DD03L A-table {suffix}"
        result["adso_table_suffix"] = suffix
        result["adso_table_name"] = f"/BIC/A{normalized_object}{suffix}"
        return result

    if normalized_type == "IOBJ" and normalized_object:
        normalized_subtype = normalize_iobj_subtype(object_subtype)
        if origin == "dd03l" and normalized_subtype:
            result["inventory_label"] = f"DD03L {normalized_subtype}"
        if normalized_subtype:
            table_candidates = build_iobj_dd03l_table_candidates_for_subtype(normalized_object, normalized_subtype)
            if table_candidates:
                result["adso_table_name"] = table_candidates[0]
        else:
            result["adso_table_name"] = "--"
        return result

    if normalized_type not in {"", "ADSO", "IOBJ"}:
        result["adso_table_name"] = "--"
        return result

    if origin == "dd03l":
        result["inventory_label"] = "DD03L"
    elif origin == "rsksfieldnew":
        result["inventory_label"] = "RSKSFIELDNEW"
    elif origin == "rsdssegfd":
        result["inventory_label"] = "RSDSSEGFD"

    return result


def build_object_field_diagnostics(
    rows: List[Dict[str, object]],
    field_key: str,
    object_name: str,
    object_system: str,
    object_type: str,
    object_subtype: str,
    object_display_name_map: Dict[str, str],
    rsds_inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]],
    dd03l_inventory: Dict[str, Dict[str, Dict[str, object]]],
    trcs_inventory: Dict[str, Dict[str, Dict[str, object]]],
    iobj_inventory: Dict[Tuple[str, str], Dict[str, Dict[str, object]]],
    activate_data_by_object: Dict[str, str],
    rsiobj_names: Set[str] | None = None,
) -> Dict[str, object]:
    normalized_object = normalize_bw_object_lookup(object_name)
    normalized_system = normalize_bw_object_lookup(object_system)
    normalized_type = normalize_type_code(object_type)
    inventory, comparison_available, inventory_origin = get_supported_inventory_for_object(
        normalized_object,
        normalized_system,
        normalized_type,
        object_subtype,
        rsds_inventory,
        dd03l_inventory,
        trcs_inventory,
        iobj_inventory,
    )
    row_fields = rows_unique_field_name_set(rows, field_key)
    inventory_fields = inventory_unique_field_name_set(inventory) if comparison_available else set()
    missing_fields = inventory_fields - row_fields
    extra_fields = row_fields - inventory_fields
    reference_meta = build_inventory_reference_meta(normalized_object, normalized_type, object_subtype, inventory_origin, activate_data_by_object)

    return {
        "object": normalized_object,
        "object_display_name": str(object_display_name_map.get(normalized_object) or "").strip(),
        "object_type": normalized_type,
        "object_system": normalized_system,
        "field_key": field_key,
        "unique_field_count": len(row_fields),
        "comparison_available": comparison_available,
        "inventory_field_count": len(inventory_fields) if comparison_available else None,
        "difference_count": (len(missing_fields) + len(extra_fields)) if comparison_available else None,
        "missing_count": len(missing_fields) if comparison_available else None,
        "extra_count": len(extra_fields) if comparison_available else None,
        "has_difference": bool(missing_fields or extra_fields) if comparison_available else False,
        **reference_meta,
    }


def rebuild_rstran_mapping_rule_full_table() -> Dict[str, int]:
    ensure_rstran_mapping_rule_full_table()
    tran_metadata = fetch_active_tran_metadata()
    base_rows = fetch_normalized_tran_mapping_rows()
    base_result = rebuild_rstran_mapping_rule_table(base_rows)
    inventory_result = rebuild_bw_object_field_inventory_table(tran_metadata)
    adso_objects = sorted(
        {
            normalize_bw_object_lookup(meta.get("source_object", ""))
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "ADSO"
        }
        | {
            normalize_bw_object_lookup(meta.get("target_object", ""))
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "ADSO"
        }
    )
    odso_objects = sorted(
        {
            normalize_bw_object_lookup(meta.get("source_object", ""))
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "ODSO"
        }
        | {
            normalize_bw_object_lookup(meta.get("target_object", ""))
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "ODSO"
        }
    )
    trcs_objects = sorted(
        {
            normalize_bw_object_lookup(meta.get("source_object", ""))
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "TRCS"
        }
        | {
            normalize_bw_object_lookup(meta.get("target_object", ""))
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "TRCS"
        }
    )
    iobj_objects = sorted(
        {
            (
                normalize_bw_object_lookup(meta.get("source_object", "")),
                normalize_iobj_subtype(meta.get("source_subtype", "")),
            )
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("source_type")) == "IOBJ"
            and normalize_bw_object_lookup(meta.get("source_object", ""))
        }
        | {
            (
                normalize_bw_object_lookup(meta.get("target_object", "")),
                normalize_iobj_subtype(meta.get("target_subtype", "")),
            )
            for meta in tran_metadata.values()
            if normalize_type_code(meta.get("target_type")) == "IOBJ"
            and normalize_bw_object_lookup(meta.get("target_object", ""))
        }
    )
    rsds_inventory = fetch_rsds_field_inventory()
    dd03l_inventory = fetch_materialized_dd03l_field_inventory("ODSO", odso_objects)
    dd03l_inventory.update(fetch_materialized_dd03l_field_inventory("ADSO", adso_objects))
    trcs_inventory = fetch_materialized_trcs_field_inventory(trcs_objects)
    iobj_inventory = fetch_materialized_iobj_field_inventory(iobj_objects)
    rsiobj_names = fetch_rsiobj_name_set() if adso_objects else set()

    rows_by_tran: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in base_rows:
        tran_id = str(row.get("tran_id") or "").strip()
        if tran_id:
            rows_by_tran[tran_id].append(row)

    all_tran_ids = sorted(set(tran_metadata.keys()) | set(rows_by_tran.keys()))
    insert_rows: List[Tuple[object, ...]] = []
    insert_batch_size = 10000
    mapped_rows = 0
    source_completed_rows = 0
    target_completed_rows = 0
    touched_transforms = 0
    empty_field_set: Set[str] = set()
    empty_key_lookup: Dict[str, str] = {}
    empty_type_lookup: Dict[str, str] = {}

    for tran_id in all_tran_ids:
        tran_meta = tran_metadata.get(tran_id) or {
            "tran_id": tran_id,
            "source_type": "",
            "target_type": "",
            "source_subtype": "",
            "target_subtype": "",
            "source_object": "",
            "source_system": "",
            "target_object": "",
            "target_system": "",
        }

        tran_rows = rows_by_tran.get(tran_id, [])
        default_step_id = int(tran_rows[0].get("step_id") or 0) if tran_rows else 0
        display_order = 0
        source_object = str(tran_meta.get("source_object") or "").strip()
        source_system = str(tran_meta.get("source_system") or "").strip()
        target_object = str(tran_meta.get("target_object") or "").strip()
        target_system = str(tran_meta.get("target_system") or "").strip()
        source_type = str(tran_meta.get("source_type") or "").strip()
        target_type = str(tran_meta.get("target_type") or "").strip()
        source_subtype = str(tran_meta.get("source_subtype") or "").strip()
        target_subtype = str(tran_meta.get("target_subtype") or "").strip()

        source_inventory, source_supported, source_inventory_origin = get_supported_inventory_for_object(
            source_object,
            source_system,
            source_type,
            source_subtype,
            rsds_inventory,
            dd03l_inventory,
            trcs_inventory,
            iobj_inventory,
        )
        target_inventory, target_supported, target_inventory_origin = get_supported_inventory_for_object(
            target_object,
            target_system,
            target_type,
            target_subtype,
            rsds_inventory,
            dd03l_inventory,
            trcs_inventory,
            iobj_inventory,
        )
        source_default_inventory = source_inventory.get("") or {}
        target_default_inventory = target_inventory.get("") or {}
        source_match_cache: Dict[str, str] = {}
        target_match_cache: Dict[str, str] = {}

        mapped_source_keys: Set[Tuple[str, str]] = set()
        mapped_target_keys: Set[Tuple[str, str]] = set()

        for row in tran_rows:
            seg_id = str(row.get("seg_id") or "").strip()
            source_field = normalize_field_name(row.get("source_field"))
            target_field = normalize_field_name(row.get("target_field"))
            source_match_field = source_match_cache.get(source_field)
            if source_match_field is None:
                source_match_field = canonicalize_object_field_for_matching(
                    source_field,
                    source_type,
                    rsiobj_names,
                    source_object,
                )
                source_match_cache[source_field] = source_match_field
            target_match_field = target_match_cache.get(target_field)
            if target_match_field is None:
                target_match_field = canonicalize_object_field_for_matching(
                    target_field,
                    target_type,
                    rsiobj_names,
                    target_object,
                )
                target_match_cache[target_field] = target_match_field
            source_seg_inventory = source_inventory.get(seg_id) or source_default_inventory
            target_seg_inventory = target_inventory.get(seg_id) or target_default_inventory
            source_field_set = source_seg_inventory.get("field_set") or empty_field_set
            target_field_set = target_seg_inventory.get("field_set") or empty_field_set
            source_key_by_field = source_seg_inventory.get("key_by_field") or empty_key_lookup
            target_key_by_field = target_seg_inventory.get("key_by_field") or empty_key_lookup
            source_type_by_field = source_seg_inventory.get("type_by_field") or empty_type_lookup
            target_type_by_field = target_seg_inventory.get("type_by_field") or empty_type_lookup
            source_field_type = str(source_type_by_field.get(source_match_field) or "").strip().upper() if source_supported and source_match_field else ""
            target_field_type = str(target_type_by_field.get(target_match_field) or "").strip().upper() if target_supported and target_match_field else ""
            source_iobj_name = source_match_field if is_rsiobj_backed_field_type(source_field_type) else ""
            target_iobj_name = target_match_field if is_rsiobj_backed_field_type(target_field_type) else ""

            if source_match_field:
                mapped_source_keys.add((seg_id, source_match_field))
            if target_match_field:
                mapped_target_keys.add((seg_id, target_match_field))

            if source_field and source_supported:
                source_field_matched = 1 if source_match_field in source_field_set else 0
            else:
                source_field_matched = 1 if source_field else 0

            if target_field and target_supported:
                target_field_matched = 1 if target_match_field in target_field_set else 0
            else:
                target_field_matched = 1 if target_field else 0

            row_kind = "mapped" if source_field and target_field else "source_only" if source_field else "target_only"
            display_order += 1
            insert_rows.append(
                (
                    tran_id,
                    int(row.get("rule_id") or 0),
                    int(row.get("step_id") or 0),
                    seg_id,
                    int(row.get("pair_index") or 0),
                    display_order,
                    source_object,
                    source_system,
                    target_object,
                    target_system,
                    source_type,
                    target_type,
                    source_subtype,
                    target_subtype,
                    source_field,
                    source_match_field,
                    source_iobj_name,
                    source_field_type,
                    target_field,
                    target_match_field,
                    target_iobj_name,
                    target_field_type,
                    "rstranfield" if source_field else "",
                    "rstranfield" if target_field else "",
                    source_field_matched,
                    target_field_matched,
                    str(source_key_by_field.get(source_match_field) or "").strip() if source_supported else "",
                    str(row.get("target_key") or "").strip() or (str(target_key_by_field.get(target_match_field) or "").strip() if target_supported else ""),
                    str(row.get("rule") or "").strip(),
                    str(row.get("group_type") or "").strip(),
                    str(row.get("aggr") or "").strip(),
                    str(row.get("ruleposit") or "").strip(),
                    row_kind,
                )
            )
            mapped_rows += 1

        mapped_source_any_fields = {field_name for _seg_id, field_name in mapped_source_keys}
        source_pair_index: Dict[str, int] = defaultdict(int)
        for seg_id, seg_inventory in source_inventory.items():
            type_by_field = seg_inventory.get("type_by_field") or empty_type_lookup
            for field_name, keyflag in seg_inventory.get("ordered") or []:
                if (seg_id, field_name) in mapped_source_keys or (not seg_id and field_name in mapped_source_any_fields):
                    continue
                source_pair_index[seg_id] += 1
                display_order += 1
                field_type = str(type_by_field.get(field_name) or "").strip().upper()
                insert_rows.append(
                    (
                        tran_id,
                        0,
                        default_step_id,
                        seg_id,
                        source_pair_index[seg_id],
                        display_order,
                        source_object,
                        source_system,
                        target_object,
                        target_system,
                        source_type,
                        target_type,
                        source_subtype,
                        target_subtype,
                        field_name,
                        field_name,
                        field_name if is_rsiobj_backed_field_type(field_type) else "",
                        field_type,
                        "",
                        "",
                        "",
                        "",
                        source_inventory_origin,
                        "",
                        1,
                        0,
                        keyflag,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "source_only",
                    )
                )
                source_completed_rows += 1

        mapped_target_any_fields = {field_name for _seg_id, field_name in mapped_target_keys}
        target_pair_index: Dict[str, int] = defaultdict(int)
        for seg_id, seg_inventory in target_inventory.items():
            type_by_field = seg_inventory.get("type_by_field") or empty_type_lookup
            for field_name, keyflag in seg_inventory.get("ordered") or []:
                if (seg_id, field_name) in mapped_target_keys or (not seg_id and field_name in mapped_target_any_fields):
                    continue
                target_pair_index[seg_id] += 1
                display_order += 1
                field_type = str(type_by_field.get(field_name) or "").strip().upper()
                insert_rows.append(
                    (
                        tran_id,
                        0,
                        default_step_id,
                        seg_id,
                        target_pair_index[seg_id],
                        display_order,
                        source_object,
                        source_system,
                        target_object,
                        target_system,
                        source_type,
                        target_type,
                        source_subtype,
                        target_subtype,
                        "",
                        "",
                        "",
                        "",
                        field_name,
                        field_name,
                        field_name if is_rsiobj_backed_field_type(field_type) else "",
                        field_type,
                        "",
                        target_inventory_origin,
                        0,
                        1,
                        "",
                        keyflag,
                        "",
                        "",
                        "",
                        "",
                        "target_only",
                    )
                )
                target_completed_rows += 1

        if tran_rows or source_inventory or target_inventory:
            touched_transforms += 1

    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(f"DELETE FROM `{RSTRAN_MAPPING_RULE_FULL_TABLE}`")
        else:
            cur.execute(f"TRUNCATE TABLE `{RSTRAN_MAPPING_RULE_FULL_TABLE}`")
        if insert_rows:
            insert_sql = f"""
                INSERT INTO `{RSTRAN_MAPPING_RULE_FULL_TABLE}` (
                    tran_id,
                    rule_id,
                    step_id,
                    seg_id,
                    pair_index,
                    display_order,
                    source_object,
                    source_system,
                    target_object,
                    target_system,
                    source_type,
                    target_type,
                    source_subtype,
                    target_subtype,
                    source_field,
                    source_match_field,
                    source_iobj_name,
                    source_fieldtype,
                    target_field,
                    target_match_field,
                    target_iobj_name,
                    target_fieldtype,
                    source_field_origin,
                    target_field_origin,
                    source_field_matched,
                    target_field_matched,
                    source_field_is_key,
                    target_field_is_key,
                    rule_type,
                    group_type,
                    aggr,
                    ruleposit,
                    row_kind
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            for start in range(0, len(insert_rows), insert_batch_size):
                cur.executemany(insert_sql, insert_rows[start:start + insert_batch_size])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        cur.close()
        conn.close()

    return {
        "inserted_rows": len(insert_rows),
        "tran_count": touched_transforms,
        "mapped_rows": mapped_rows,
        "source_completed_rows": source_completed_rows,
        "target_completed_rows": target_completed_rows,
        "base_inserted_rows": int(base_result.get("inserted_rows") or 0),
        "base_tran_count": int(base_result.get("tran_count") or 0),
        "inventory_inserted_rows": int(inventory_result.get("inserted_rows") or 0),
        "inventory_object_count": int(inventory_result.get("object_count") or 0),
    }


def fetch_normalized_tran_mapping_rows(tran_id: Optional[str] = None) -> List[Dict[str, object]]:
    normalized_tran_id = str(tran_id or "").strip()
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        if normalized_tran_id:
            cur.execute(get_tran_mapping_rule_select_sql(filter_by_tran=True), (normalized_tran_id,))
        else:
            cur.execute(get_tran_mapping_rule_select_sql(filter_by_tran=False))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    return normalize_tran_mapping_rows(rows)


def rebuild_rstran_mapping_rule_table(normalized_rows: Optional[List[Dict[str, object]]] = None) -> Dict[str, int]:
    ensure_rstran_mapping_rule_table()

    conn = get_conn()
    inserted_rows = 0
    tran_count = 0
    try:
        cur = conn.cursor()
        try:
            if IS_SQLITE:
                cur.execute(f"DELETE FROM `{RSTRAN_MAPPING_RULE_TABLE}`")
            else:
                cur.execute(f"TRUNCATE TABLE `{RSTRAN_MAPPING_RULE_TABLE}`")

            if normalized_rows is None and IS_SQLITE:
                cur.execute(
                    f"""
                    INSERT INTO `{RSTRAN_MAPPING_RULE_TABLE}` (
                        tran_id,
                        rule_id,
                        step_id,
                        seg_id,
                        pair_index,
                        ruleposit,
                        source_field,
                        target_field,
                        target_keyflag,
                        rule_type,
                        group_type,
                        aggr
                    )
                    SELECT
                        tran_id,
                        rule_id,
                        step_id,
                        seg_id,
                        pair_index,
                        ruleposit,
                        source_field,
                        target_field,
                        target_keyflag,
                        rule_type,
                        group_type,
                        aggr
                    FROM (
                        {get_tran_mapping_rule_select_sql(filter_by_tran=False)}
                    ) AS mapping_rows
                    """
                )
                count_cur = conn.cursor(dictionary=True)
                try:
                    count_cur.execute(
                        f"""
                        SELECT COUNT(*) AS row_count, COUNT(DISTINCT tran_id) AS tran_count
                        FROM `{RSTRAN_MAPPING_RULE_TABLE}`
                        """
                    )
                    counts = count_cur.fetchone() or {}
                    inserted_rows = int(counts.get("row_count") or 0)
                    tran_count = int(counts.get("tran_count") or 0)
                finally:
                    count_cur.close()
            else:
                resolved_rows = normalized_rows if normalized_rows is not None else fetch_normalized_tran_mapping_rows()
                if resolved_rows:
                    cur.executemany(
                        f"""
                        INSERT INTO `{RSTRAN_MAPPING_RULE_TABLE}` (
                            tran_id,
                            rule_id,
                            step_id,
                            seg_id,
                            pair_index,
                            ruleposit,
                            source_field,
                            target_field,
                            target_keyflag,
                            rule_type,
                            group_type,
                            aggr
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        [
                            (
                                row["tran_id"],
                                row["rule_id"],
                                row["step_id"],
                                row["seg_id"],
                                row["pair_index"],
                                row["ruleposit"],
                                row["source_field"],
                                row["target_field"],
                                row["target_key"],
                                row["rule"],
                                row.get("group_type", ""),
                                row["aggr"],
                            )
                            for row in resolved_rows
                        ],
                    )
                inserted_rows = len(resolved_rows)
                tran_count = len({row["tran_id"] for row in resolved_rows if row.get("tran_id")})
        finally:
            cur.close()

        conn.commit()
        return {
            "inserted_rows": inserted_rows,
            "tran_count": tran_count,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def ensure_auth_tables() -> None:
    if IS_SQLITE:
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS `users` (
                  `username` TEXT NOT NULL,
                  `password_hash` TEXT NOT NULL,
                  `role` TEXT NOT NULL DEFAULT 'user',
                  `is_locked` INTEGER NOT NULL DEFAULT 0,
                  `failed_attempts` INTEGER NOT NULL DEFAULT 0,
                  `temp_lock_until` TEXT NULL,
                  `last_login_at` TEXT NULL,
                  `created_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  `updated_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`username`)
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS `user_sessions` (
                  `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                  `username` TEXT NOT NULL,
                  `session_hash` TEXT NOT NULL UNIQUE,
                  `expires_at` TEXT NOT NULL,
                  `revoked` INTEGER NOT NULL DEFAULT 0,
                  `created_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  `last_seen_at` TEXT NULL,
                  FOREIGN KEY (`username`) REFERENCES `users` (`username`) ON DELETE CASCADE
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS `auth_audit_logs` (
                  `id` INTEGER PRIMARY KEY AUTOINCREMENT,
                  `event_type` TEXT NOT NULL,
                  `username` TEXT NULL,
                  `actor` TEXT NULL,
                  `success` INTEGER NOT NULL DEFAULT 1,
                  `detail` TEXT NULL,
                  `created_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.commit()

            cur.execute("SELECT COUNT(*) FROM users")
            has_user = int(cur.fetchone()[0]) > 0
            if not has_user:
                bootstrap_admin_username = normalize_username(DEFAULT_ADMIN_USERNAME) or "admin"
                if not DEFAULT_ADMIN_PASSWORD:
                    raise RuntimeError("DEFAULT_ADMIN_PASSWORD is required when bootstrapping the first admin user")
                password_error = get_password_strength_error(DEFAULT_ADMIN_PASSWORD)
                if password_error:
                    raise RuntimeError(f"DEFAULT_ADMIN_PASSWORD is too weak: {password_error}")
                cur.execute(
                    "INSERT INTO users (username, password_hash, role, is_locked, failed_attempts) VALUES (%s, %s, 'admin', 0, 0)",
                    (bootstrap_admin_username, hash_password(DEFAULT_ADMIN_PASSWORD)),
                )
                conn.commit()
        finally:
            cur.close()
            conn.close()
        return
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS `users` (
              `username` VARCHAR(64) NOT NULL,
              `password_hash` VARCHAR(255) NOT NULL,
              `role` VARCHAR(16) NOT NULL DEFAULT 'user',
              `is_locked` TINYINT(1) NOT NULL DEFAULT 0,
              `failed_attempts` INT NOT NULL DEFAULT 0,
              `temp_lock_until` DATETIME NULL,
              `last_login_at` DATETIME NULL,
              `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `updated_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`username`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS `user_sessions` (
              `id` BIGINT NOT NULL AUTO_INCREMENT,
              `username` VARCHAR(64) NOT NULL,
              `session_hash` CHAR(64) NOT NULL,
              `expires_at` DATETIME NOT NULL,
              `revoked` TINYINT(1) NOT NULL DEFAULT 0,
              `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `last_seen_at` DATETIME NULL,
              PRIMARY KEY (`id`),
              UNIQUE KEY `uk_session_hash` (`session_hash`),
              KEY `idx_session_user` (`username`),
              CONSTRAINT `fk_sessions_user` FOREIGN KEY (`username`) REFERENCES `users` (`username`) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS `auth_audit_logs` (
              `id` BIGINT NOT NULL AUTO_INCREMENT,
              `event_type` VARCHAR(64) NOT NULL,
              `username` VARCHAR(64) NULL,
              `actor` VARCHAR(64) NULL,
              `success` TINYINT(1) NOT NULL DEFAULT 1,
              `detail` VARCHAR(255) NULL,
              `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`),
              KEY `idx_auth_event` (`event_type`, `created_at`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM users")
        has_user = int(cur.fetchone()[0]) > 0
        if not has_user:
            bootstrap_admin_username = normalize_username(DEFAULT_ADMIN_USERNAME) or "admin"
            if not DEFAULT_ADMIN_PASSWORD:
                raise RuntimeError(
                    "DEFAULT_ADMIN_PASSWORD is required when bootstrapping the first admin user"
                )
            password_error = get_password_strength_error(DEFAULT_ADMIN_PASSWORD)
            if password_error:
                raise RuntimeError(f"DEFAULT_ADMIN_PASSWORD is too weak: {password_error}")
            cur.execute(
                "INSERT INTO users (username, password_hash, role, is_locked, failed_attempts) VALUES (%s, %s, 'admin', 0, 0)",
                (bootstrap_admin_username, hash_password(DEFAULT_ADMIN_PASSWORD)),
            )
            conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_user_hidden_object_table() -> None:
    if IS_SQLITE:
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS `user_hidden_object` (
                  `bw_object` TEXT NOT NULL,
                  `sourcesys` TEXT NOT NULL DEFAULT '',
                  `created_at` TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`bw_object`, `sourcesys`)
                )
                """
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
                        CREATE TABLE IF NOT EXISTS `user_hidden_object` (
              `bw_object` VARCHAR(40) NOT NULL COMMENT 'BW Object',
                            `sourcesys` VARCHAR(25) NOT NULL DEFAULT '' COMMENT 'Source System',
                            `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            PRIMARY KEY (`bw_object`, `sourcesys`)
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='用户隐藏对象清单';
            """
        )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'user_del_flag'
            """,
            (DB_CONFIG["database"],),
        )
        has_legacy_table = int(cur.fetchone()[0]) > 0

        if has_legacy_table:
            cur.execute(
                """
                INSERT IGNORE INTO `user_hidden_object` (`bw_object`, `sourcesys`)
                SELECT
                    UPPER(TRIM(`bw_object`)) AS bw_object,
                    COALESCE(NULLIF(UPPER(TRIM(`sourcesys`)), ''), '') AS sourcesys
                FROM `user_del_flag`
                WHERE `bw_object` IS NOT NULL AND TRIM(`bw_object`) <> ''
                """
            )
            cur.execute("DROP TABLE IF EXISTS `user_del_flag`")

        conn.commit()
    finally:
        cur.close()
        conn.close()


def audit_log(event_type: str, username: str | None, success: bool, detail: str = "", actor: str | None = None) -> None:
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO auth_audit_logs (event_type, username, actor, success, detail)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (event_type, normalize_username(username) if username else None, normalize_username(actor) if actor else None, 1 if success else 0, detail[:255] if detail else None),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def fetch_user_by_username(username: str) -> Dict[str, object] | None:
    name = normalize_username(username)
    if not name:
        return None
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT username, password_hash, role, is_locked, failed_attempts, temp_lock_until, last_login_at
            FROM users
            WHERE username = %s
            """,
            (name,),
        )
        row = cur.fetchone()
        return row
    finally:
        cur.close()
        conn.close()


def create_session(username: str) -> tuple[str, datetime]:
    token = secrets.token_urlsafe(32)
    token_hash = hash_session_token(token)
    expires_at = utcnow() + timedelta(hours=AUTH_SESSION_HOURS)

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO user_sessions (username, session_hash, expires_at, revoked, last_seen_at)
            VALUES (%s, %s, %s, 0, %s)
            """,
            (normalize_username(username), token_hash, expires_at, utcnow()),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return token, expires_at


def revoke_session(token: str) -> None:
    token_hash = hash_session_token(token)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE user_sessions SET revoked = 1 WHERE session_hash = %s", (token_hash,))
        conn.commit()
    finally:
        cur.close()
        conn.close()


def get_request_session_token(request: Request) -> str:
    raw_token = str(request.cookies.get(AUTH_COOKIE_NAME) or "").strip()
    if raw_token:
        return raw_token

    auth_header = str(request.headers.get("Authorization") or "").strip()
    if auth_header.lower().startswith("bearer "):
        return auth_header[7:].strip()

    return ""


def resolve_user_from_request(request: Request) -> Dict[str, object] | None:
    raw_token = get_request_session_token(request)
    if not raw_token:
        return None

    token_hash = hash_session_token(raw_token)
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT s.username, s.expires_at, s.revoked, u.role, u.is_locked
            FROM user_sessions s
            JOIN users u ON u.username = s.username
            WHERE s.session_hash = %s
            LIMIT 1
            """,
            (token_hash,),
        )
        row = cur.fetchone()
        if not row:
            return None

        if int(row.get("revoked") or 0) == 1:
            return None
        expires_at = row.get("expires_at")
        if not isinstance(expires_at, datetime) or expires_at <= utcnow():
            cur.execute("UPDATE user_sessions SET revoked = 1 WHERE session_hash = %s", (token_hash,))
            conn.commit()
            return None
        if int(row.get("is_locked") or 0) == 1:
            return None

        cur.execute("UPDATE user_sessions SET last_seen_at = %s WHERE session_hash = %s", (utcnow(), token_hash))
        conn.commit()
        return {"username": row["username"], "role": row["role"]}
    finally:
        cur.close()
        conn.close()


def get_table_columns(table_name: str) -> List[str]:
    if IS_SQLITE:
        conn = get_conn()
        resolved_name = sqlite_resolve_table_name(conn, table_name)
        cur = conn.cursor()
        try:
            cur.execute(f"PRAGMA table_info(`{resolved_name}`)")
            rows = cur.fetchall()
            return [str(row[1] or "").strip() for row in rows if str(row[1] or "").strip()]
        finally:
            cur.close()
            conn.close()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
        ORDER BY ORDINAL_POSITION
        """,
        (DB_CONFIG["database"], table_name),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [r[0] for r in rows]


def table_exists(table_name: str) -> bool:
    normalized_table_name = str(table_name or "").strip().lower()
    if not normalized_table_name:
        return False
    cached = _TABLE_EXISTS_CACHE.get(normalized_table_name)
    if cached is not None:
        return cached

    if IS_SQLITE:
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND LOWER(name) = LOWER(?) LIMIT 1",
                (normalized_table_name,),
            )
            exists = bool(cur.fetchone())
        finally:
            cur.close()
            conn.close()
        _TABLE_EXISTS_CACHE[normalized_table_name] = exists
        return exists
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT 1
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
            LIMIT 1
            """,
            (DB_CONFIG["database"], table_name),
        )
        exists = bool(cur.fetchone())
    finally:
        cur.close()
        conn.close()
    _TABLE_EXISTS_CACHE[normalized_table_name] = exists
    return exists


def get_table_schema(table_name: str) -> List[Dict[str, object]]:
    ensure_import_target_table_schema(table_name)
    if IS_SQLITE:
        conn = get_conn()
        resolved_name = sqlite_resolve_table_name(conn, table_name)
        cur = conn.cursor()
        try:
            cur.execute(f"PRAGMA table_info(`{resolved_name}`)")
            rows = cur.fetchall()
        finally:
            cur.close()
            conn.close()

        field_text_lookup = fetch_dd03t_table_field_text_lookup([table_name])
        fallback_field_text_lookup = fetch_dd03t_field_text_fallback_lookup([str(row[1] or "") for row in rows])
        normalized_table_name = str(table_name or "").strip().upper()
        rs_field_text_lookup = fetch_rs_table_field_text_lookup_by_rollname(normalized_table_name) if normalized_table_name.startswith("RS") else {}
        return [
            {
                "name": str(row[1] or "").strip(),
                "comment": "",
                "field_text": (
                    rs_field_text_lookup.get(str(row[1] or "").strip().upper(), "")
                    or field_text_lookup.get(
                        (normalized_table_name, str(row[1] or "").strip().upper()),
                        fallback_field_text_lookup.get(str(row[1] or "").strip().upper(), ""),
                    )
                ),
                "data_type": str(row[2] or "").strip().lower(),
                "column_type": str(row[2] or "").strip().lower(),
                "default": row[4],
                "nullable": int(row[3] or 0) == 0,
                "is_key": int(row[5] or 0) > 0,
                "ordinal_position": int(row[0] or 0) + 1,
            }
            for row in rows
            if str(row[1] or "").strip()
        ]
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT
            COLUMN_NAME,
            COLUMN_COMMENT,
            DATA_TYPE,
            COLUMN_TYPE,
            COLUMN_DEFAULT,
            IS_NULLABLE,
            COLUMN_KEY,
            ORDINAL_POSITION
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
        ORDER BY ORDINAL_POSITION
        """,
        (DB_CONFIG["database"], table_name),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    field_text_lookup = fetch_dd03t_table_field_text_lookup([table_name])
    fallback_field_text_lookup = fetch_dd03t_field_text_fallback_lookup([str(row.get("COLUMN_NAME") or "") for row in rows])
    normalized_table_name = str(table_name or "").strip().upper()
    rs_field_text_lookup = fetch_rs_table_field_text_lookup_by_rollname(normalized_table_name) if normalized_table_name.startswith("RS") else {}
    return [
        {
            "name": str(row.get("COLUMN_NAME") or "").strip(),
            "comment": str(row.get("COLUMN_COMMENT") or "").strip(),
            "field_text": (
                rs_field_text_lookup.get(str(row.get("COLUMN_NAME") or "").strip().upper(), "")
                or field_text_lookup.get(
                    (normalized_table_name, str(row.get("COLUMN_NAME") or "").strip().upper()),
                    fallback_field_text_lookup.get(str(row.get("COLUMN_NAME") or "").strip().upper(), ""),
                )
            ),
            "data_type": str(row.get("DATA_TYPE") or "").strip(),
            "column_type": str(row.get("COLUMN_TYPE") or "").strip(),
            "default": row.get("COLUMN_DEFAULT"),
            "nullable": str(row.get("IS_NULLABLE") or "").strip().upper() == "YES",
            "is_key": str(row.get("COLUMN_KEY") or "").strip().upper() == "PRI",
            "ordinal_position": int(row.get("ORDINAL_POSITION") or 0),
        }
        for row in rows
        if str(row.get("COLUMN_NAME") or "").strip()
    ]


def get_table_column_names(table_name: str) -> Set[str]:
    normalized_table_name = str(table_name or "").strip().lower()
    if not normalized_table_name:
        return set()
    cached = _TABLE_COLUMN_NAMES_CACHE.get(normalized_table_name)
    if cached is not None:
        return set(cached)

    if IS_SQLITE:
        conn = get_conn()
        resolved_name = sqlite_resolve_table_name(conn, table_name)
        cur = conn.cursor()
        try:
            cur.execute(f"PRAGMA table_info(`{resolved_name}`)")
            rows = cur.fetchall()
            columns = {str(row[1] or "").strip().upper() for row in rows if str(row[1] or "").strip()}
        finally:
            cur.close()
            conn.close()
    else:
        conn = get_conn()
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute(
                """
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """,
                (DB_CONFIG["database"], table_name),
            )
            rows = cur.fetchall()
            columns = {str(row.get("COLUMN_NAME") or "").strip().upper() for row in rows if str(row.get("COLUMN_NAME") or "").strip()}
        finally:
            cur.close()
            conn.close()

    _TABLE_COLUMN_NAMES_CACHE[normalized_table_name] = set(columns)
    return columns


def get_primary_keys(table_name: str) -> List[str]:
    if IS_SQLITE:
        conn = get_conn()
        resolved_name = sqlite_resolve_table_name(conn, table_name)
        cur = conn.cursor()
        try:
            cur.execute(f"PRAGMA table_info(`{resolved_name}`)")
            rows = cur.fetchall()
            pk_rows = sorted((row for row in rows if int(row[5] or 0) > 0), key=lambda item: int(item[5] or 0))
            return [str(row[1] or "").strip() for row in pk_rows if str(row[1] or "").strip()]
        finally:
            cur.close()
            conn.close()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SHOW KEYS FROM `{table_name}` WHERE Key_name = 'PRIMARY'")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [r[4] for r in rows]


def count_table_rows(table_name: str) -> int:
    if not table_exists(table_name):
        return 0
    conn = get_conn()
    cur = conn.cursor()
    if table_name == "bw_object_name":
        cur.execute(
            """
            SELECT COUNT(*)
            FROM `bw_object_name`
            WHERE COALESCE(TRIM(`NAME_EN`), '') <> ''
               OR COALESCE(TRIM(`NAME_DE`), '') <> ''
            """
        )
    else:
        cur.execute(f"SELECT COUNT(*) FROM `{table_name}`")
    value = int(cur.fetchone()[0])
    cur.close()
    conn.close()
    return value


class MaterializedReadModelNotReadyError(RuntimeError):
    def __init__(self, missing_tables: List[str], empty_tables: List[str], counts: Dict[str, int]):
        self.missing_tables = list(missing_tables)
        self.empty_tables = list(empty_tables)
        self.counts = dict(counts)
        problems: List[str] = []
        if self.missing_tables:
            problems.append(f"missing={', '.join(self.missing_tables)}")
        if self.empty_tables:
            problems.append(f"empty={', '.join(self.empty_tables)}")
        detail = "materialized read model not ready"
        if problems:
            detail = f"{detail}: {'; '.join(problems)}"
        super().__init__(detail)


def build_materialization_not_ready_http_exception(error: MaterializedReadModelNotReadyError) -> HTTPException:
    return HTTPException(
        status_code=503,
        detail={
            "code": "materialized_read_model_not_ready",
            "message": "物化读模型未就绪，运行时查询已被禁止。请先重建字段库存和全量映射表。",
            "missing_tables": error.missing_tables,
            "empty_tables": error.empty_tables,
            "table_counts": error.counts,
            "required_actions": [
                "/api/import/rebuild-bw-object-field-inventory",
                "/api/import/rebuild-rstran-mapping-rule-full",
            ],
        },
    )


def assert_materialized_read_model_ready(required_tables: Optional[List[str]] = None) -> Dict[str, int]:
    normalized_required_tables = required_tables or [
        BW_OBJECT_FIELD_INVENTORY_TABLE,
        RSTRAN_MAPPING_RULE_FULL_TABLE,
    ]
    missing_tables: List[str] = []
    empty_tables: List[str] = []
    counts: Dict[str, int] = {}

    for table_name in normalized_required_tables:
        normalized_table_name = str(table_name or "").strip()
        if not normalized_table_name:
            continue
        if not table_exists(normalized_table_name):
            missing_tables.append(normalized_table_name)
            counts[normalized_table_name] = 0
            continue
        row_count = count_table_rows(normalized_table_name)
        counts[normalized_table_name] = row_count
        if row_count <= 0:
            empty_tables.append(normalized_table_name)

    if missing_tables or empty_tables:
        raise MaterializedReadModelNotReadyError(missing_tables, empty_tables, counts)

    return counts


def list_queryable_table_names() -> List[str]:
    if IS_SQLITE:
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
            rows = cur.fetchall()
        finally:
            cur.close()
            conn.close()

        normalized_names = []
        for row in rows:
            table_name = str(row[0] or "").strip()
            if not table_name or table_name.startswith("sqlite_"):
                continue
            normalized_name = table_name.lower()
            if normalized_name in DATA_QUERY_EXCLUDED_TABLES or normalized_name.startswith("auth_"):
                continue
            normalized_names.append(normalized_name)
        return sorted(set(normalized_names))
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT TABLE_NAME
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = %s
              AND TABLE_TYPE = 'BASE TABLE'
            ORDER BY TABLE_NAME
            """,
            (DB_CONFIG["database"],),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return [
        str(row[0] or "").strip()
        for row in rows
        if str(row[0] or "").strip()
        and str(row[0] or "").strip() not in DATA_QUERY_EXCLUDED_TABLES
        and not str(row[0] or "").strip().lower().startswith("auth_")
    ]


def normalize_query_table_name(value: object) -> str:
    return str(value or "").strip().lower()


def ensure_queryable_table_name(value: object) -> str:
    normalized_name = normalize_query_table_name(value)
    if not normalized_name:
        raise HTTPException(status_code=400, detail="Main table is required")
    if normalized_name in DATA_QUERY_EXCLUDED_TABLES or normalized_name.startswith("auth_"):
        raise HTTPException(status_code=400, detail=f"Table is not queryable: {normalized_name}")
    if not table_exists(normalized_name):
        raise HTTPException(status_code=404, detail=f"Table not found: {normalized_name}")
    return normalized_name


def build_schema_lookup(schema_rows: List[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    lookup: Dict[str, Dict[str, object]] = {}
    for row in schema_rows:
        column_name = str(row.get("name") or "").strip()
        if not column_name:
            continue
        lookup[column_name.upper()] = row
    return lookup


def normalize_query_field_name(value: object, schema_lookup: Dict[str, Dict[str, object]], field_label: str) -> str:
    normalized_value = str(value or "").strip()
    if not normalized_value:
        raise HTTPException(status_code=400, detail=f"{field_label} is required")
    matched = schema_lookup.get(normalized_value.upper())
    if not matched:
        raise HTTPException(status_code=400, detail=f"Unknown field: {normalized_value}")
    return str(matched.get("name") or "").strip()


def normalize_query_limit(value: object, default_limit: int) -> int:
    try:
        normalized_limit = int(value)
    except (TypeError, ValueError):
        normalized_limit = default_limit
    if normalized_limit <= 0:
        normalized_limit = default_limit
    return max(1, min(DATA_QUERY_MAX_LIMIT, normalized_limit))


def normalize_query_offset(value: object) -> int:
    try:
        normalized_offset = int(value)
    except (TypeError, ValueError):
        normalized_offset = 0
    return max(0, normalized_offset)


def parse_query_field_ref(field_ref: object) -> Tuple[str, str]:
    raw = str(field_ref or "").strip()
    if not raw or "." not in raw:
        raise HTTPException(status_code=400, detail="Invalid filter field reference")
    source, field_name = raw.split(".", 1)
    normalized_source = str(source or "").strip().lower()
    normalized_field_name = str(field_name or "").strip()
    if normalized_source not in {"main", "join"} or not normalized_field_name:
        raise HTTPException(status_code=400, detail="Invalid filter field reference")
    return normalized_source, normalized_field_name


def normalize_query_filter_values(values: List[str]) -> List[str]:
    normalized_values: List[str] = []
    for item in values or []:
        chunks = re.split(r"[\r\n\t,;]+", str(item or ""))
        for chunk in chunks:
            value = str(chunk or "").strip()
            if value:
                normalized_values.append(value)
    return normalized_values


def build_data_query_select_columns(table_name: str, table_alias: str, source_key: str, schema_rows: List[Dict[str, object]]) -> Tuple[List[str], List[Dict[str, object]]]:
    select_clauses: List[str] = []
    column_meta: List[Dict[str, object]] = []
    for row in schema_rows:
        column_name = str(row.get("name") or "").strip()
        if not column_name:
            continue
        result_key = f"{source_key}__{column_name}"
        select_clauses.append(f"`{table_alias}`.`{column_name}` AS `{result_key}`")
        column_meta.append(
            {
                "key": result_key,
                "label": f"{table_name}.{column_name}",
                "table_name": table_name,
                "field_name": column_name,
                "field_ref": f"{source_key}.{column_name}",
                "source": source_key,
                "comment": str(row.get("comment") or "").strip(),
                "field_text": str(row.get("field_text") or "").strip(),
                "data_type": str(row.get("data_type") or "").strip(),
            }
        )
    return select_clauses, column_meta


def build_data_query_filter_clauses(
    filters: List[DataQueryFilterRequest],
    main_schema_lookup: Dict[str, Dict[str, object]],
    join_schema_lookup: Dict[str, Dict[str, object]],
    has_join: bool,
) -> Tuple[List[str], List[object]]:
    where_clauses: List[str] = []
    params: List[object] = []

    for filter_item in filters or []:
        source_key, field_name = parse_query_field_ref(filter_item.field_ref)
        if source_key == "join" and not has_join:
            raise HTTPException(status_code=400, detail="Join filters require a join table")

        schema_lookup = main_schema_lookup if source_key == "main" else join_schema_lookup
        resolved_field_name = normalize_query_field_name(field_name, schema_lookup, "Filter field")
        operator = str(filter_item.operator or "in").strip().lower()
        if operator not in DATA_QUERY_ALLOWED_FILTER_OPERATORS:
            raise HTTPException(status_code=400, detail=f"Unsupported filter operator: {operator}")

        table_alias = "m" if source_key == "main" else "j"
        qualified_column = f"`{table_alias}`.`{resolved_field_name}`"
        if operator == "in":
            normalized_values = normalize_query_filter_values(filter_item.values)
            if not normalized_values:
                continue
            comparable_column = f"UPPER(TRIM(CAST(COALESCE({qualified_column}, '') AS CHAR)))"
            comparison_clauses: List[str] = []
            for value in normalized_values:
                normalized_value = str(value or "").strip()
                if not normalized_value:
                    continue
                if any(token in normalized_value for token in ("*", "?", "%", "_")):
                    like_pattern = normalized_value.replace("*", "%").replace("?", "_").upper()
                    comparison_clauses.append(f"{comparable_column} LIKE %s")
                    params.append(like_pattern)
                else:
                    comparison_clauses.append(f"{comparable_column} = %s")
                    params.append(normalized_value.upper())
            if comparison_clauses:
                where_clauses.append("(" + " OR ".join(comparison_clauses) + ")")
            continue

        range_start = str(filter_item.range_start or "").strip()
        range_end = str(filter_item.range_end or "").strip()
        if range_start:
            where_clauses.append(f"{qualified_column} >= %s")
            params.append(range_start)
        if range_end:
            where_clauses.append(f"{qualified_column} <= %s")
            params.append(range_end)

    return where_clauses, params


def resolve_data_query_select_fields(
    select_fields: List[str],
    main_schema_lookup: Dict[str, Dict[str, object]],
    join_schema_lookup: Dict[str, Dict[str, object]],
    has_join: bool,
) -> List[Tuple[str, str]]:
    normalized_select_fields: List[Tuple[str, str]] = []
    seen: Set[Tuple[str, str]] = set()
    for field_ref in select_fields or []:
        source_key, field_name = parse_query_field_ref(field_ref)
        if source_key == "join" and not has_join:
            raise HTTPException(status_code=400, detail="Join select fields require a join table")
        schema_lookup = main_schema_lookup if source_key == "main" else join_schema_lookup
        resolved_field_name = normalize_query_field_name(field_name, schema_lookup, "Select field")
        normalized_key = (source_key, resolved_field_name.upper())
        if normalized_key in seen:
            continue
        seen.add(normalized_key)
        normalized_select_fields.append((source_key, resolved_field_name))
    return normalized_select_fields


def resolve_data_query_join_conditions(
    payload: DataQueryRequest,
    main_schema_lookup: Dict[str, Dict[str, object]],
    join_schema_lookup: Dict[str, Dict[str, object]],
) -> List[Tuple[str, str]]:
    raw_join_conditions = [
        (
            str(join_item.main_field or "").strip(),
            str(join_item.join_field or "").strip(),
        )
        for join_item in (payload.join_conditions or [])
    ]
    if not raw_join_conditions and (payload.main_join_field or payload.join_join_field):
        raw_join_conditions = [
            (
                str(payload.main_join_field or "").strip(),
                str(payload.join_join_field or "").strip(),
            )
        ]

    normalized_join_conditions: List[Tuple[str, str]] = []
    seen_conditions: Set[Tuple[str, str]] = set()
    for main_field_raw, join_field_raw in raw_join_conditions:
        if not main_field_raw and not join_field_raw:
            continue
        if not main_field_raw or not join_field_raw:
            raise HTTPException(status_code=400, detail="Each join condition requires both main and join fields")
        main_field = normalize_query_field_name(main_field_raw, main_schema_lookup, "Main join field")
        join_field = normalize_query_field_name(join_field_raw, join_schema_lookup, "Join table field")
        normalized_pair = (main_field, join_field)
        if normalized_pair in seen_conditions:
            continue
        seen_conditions.add(normalized_pair)
        normalized_join_conditions.append(normalized_pair)

    if not normalized_join_conditions:
        raise HTTPException(status_code=400, detail="Join table selected but no join conditions were provided")

    return normalized_join_conditions


def execute_data_query(payload: DataQueryRequest) -> Dict[str, object]:
    main_table = ensure_queryable_table_name(payload.main_table)
    join_table = normalize_query_table_name(payload.join_table)
    join_type_key = str(payload.join_type or "left").strip().lower()
    join_type_sql = DATA_QUERY_ALLOWED_JOIN_TYPES.get(join_type_key)
    if join_table:
        join_table = ensure_queryable_table_name(join_table)
        if not join_type_sql:
            raise HTTPException(status_code=400, detail="Unsupported join type")
    else:
        join_type_key = ""
        join_type_sql = ""

    main_schema_rows = get_table_schema(main_table)
    if not main_schema_rows:
        raise HTTPException(status_code=404, detail=f"Table has no schema: {main_table}")
    main_schema_lookup = build_schema_lookup(main_schema_rows)

    join_schema_rows: List[Dict[str, object]] = []
    join_schema_lookup: Dict[str, Dict[str, object]] = {}
    join_clause_sql = ""
    if join_table:
        join_schema_rows = get_table_schema(join_table)
        if not join_schema_rows:
            raise HTTPException(status_code=404, detail=f"Table has no schema: {join_table}")
        join_schema_lookup = build_schema_lookup(join_schema_rows)
        join_conditions = resolve_data_query_join_conditions(payload, main_schema_lookup, join_schema_lookup)
        join_clause_sql = (
            f" {join_type_sql} `{join_table}` AS `j`"
            + " ON "
            + " AND ".join([f"`m`.`{main_field}` = `j`.`{join_field}`" for main_field, join_field in join_conditions])
        )

    select_main_clauses, select_main_meta = build_data_query_select_columns(main_table, "m", "main", main_schema_rows)
    select_join_clauses: List[str] = []
    select_join_meta: List[Dict[str, object]] = []
    if join_table:
        select_join_clauses, select_join_meta = build_data_query_select_columns(join_table, "j", "join", join_schema_rows)

    selected_fields = resolve_data_query_select_fields(
        payload.select_fields,
        main_schema_lookup,
        join_schema_lookup,
        bool(join_table),
    )

    select_clauses = select_main_clauses + select_join_clauses
    select_meta = select_main_meta + select_join_meta
    if selected_fields:
        main_select_map = {
            str(meta.get("field_name") or "").strip().upper(): (clause, meta)
            for clause, meta in zip(select_main_clauses, select_main_meta)
            if str(meta.get("field_name") or "").strip()
        }
        join_select_map = {
            str(meta.get("field_name") or "").strip().upper(): (clause, meta)
            for clause, meta in zip(select_join_clauses, select_join_meta)
            if str(meta.get("field_name") or "").strip()
        }
        select_clauses = []
        select_meta = []
        for source_key, field_name in selected_fields:
            source_map = main_select_map if source_key == "main" else join_select_map
            selected_tuple = source_map.get(str(field_name or "").strip().upper())
            if not selected_tuple:
                continue
            select_clause, selected_meta = selected_tuple
            select_clauses.append(select_clause)
            select_meta.append(selected_meta)

    if not select_clauses:
        raise HTTPException(status_code=400, detail="No selectable fields available for current tables")

    where_clauses, where_params = build_data_query_filter_clauses(
        payload.filters,
        main_schema_lookup,
        join_schema_lookup,
        bool(join_table),
    )

    limit = normalize_query_limit(payload.limit, DATA_QUERY_DEFAULT_LIMIT)
    offset = normalize_query_offset(payload.offset)
    sql_parts = [
        "SELECT",
        ", ".join(select_clauses),
        f"FROM `{main_table}` AS `m`",
    ]
    if join_clause_sql:
        sql_parts.append(join_clause_sql)
    if where_clauses:
        sql_parts.append("WHERE " + " AND ".join(where_clauses))
    sql_parts.append(f"LIMIT {limit + 1} OFFSET {offset}")
    sql = "\n".join(sql_parts)

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, tuple(where_params))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    has_more = len(rows) > limit
    if has_more:
        rows = rows[:limit]

    return {
        "main_table": main_table,
        "join_table": join_table,
        "join_type": join_type_key,
        "columns": select_meta,
        "rows": rows,
        "returned_count": len(rows),
        "limit": limit,
        "offset": offset,
        "has_more": has_more,
    }


def get_text_column_capacity(column_meta: Dict[str, object]) -> Tuple[str, int] | None:
    column_type = str(column_meta.get("column_type") or "").strip().lower()
    match = re.fullmatch(r"(char|varchar)\((\d+)\)", column_type)
    if match:
        return match.group(1), int(match.group(2))

    text_capacities = {
        "tinytext": 255,
        "text": 65535,
        "mediumtext": 16777215,
        "longtext": 4294967295,
    }
    if column_type in text_capacities:
        return column_type, text_capacities[column_type]

    return None


def get_import_series_max_length(series: pd.Series) -> int:
    max_length = 0
    for value in series.tolist():
        try:
            if pd.isna(value):
                continue
        except Exception:
            pass
        text = str(value)
        if len(text) > max_length:
            max_length = len(text)
    return max_length


def get_import_values_max_length(values: List[object]) -> int:
    max_length = 0
    for value in values:
        if value is None:
            continue
        text = str(value)
        if len(text) > max_length:
            max_length = len(text)
    return max_length


def build_text_capacity_alter_clauses(
    schema_rows: List[Dict[str, object]],
    values_by_column: Dict[str, List[object]],
) -> List[str]:
    schema_map = {str(item.get("name") or "").strip(): item for item in schema_rows}
    alter_clauses: List[str] = []
    for column_name, values in values_by_column.items():
        normalized_column = str(column_name or "").strip()
        column_meta = schema_map.get(normalized_column, {})
        capacity_info = get_text_column_capacity(column_meta)
        if not capacity_info:
            continue

        type_name, current_length = capacity_info
        max_length = get_import_values_max_length(values)
        if max_length <= current_length:
            continue

        if max_length <= 255 and type_name == "char":
            target_sql = f"CHAR({max_length})"
        elif max_length <= 65535:
            target_sql = f"VARCHAR({max_length})" if type_name in {"char", "varchar"} else "TEXT"
        elif max_length <= 16777215:
            target_sql = "MEDIUMTEXT"
        else:
            target_sql = "LONGTEXT"

        null_sql = "NULL" if bool(column_meta.get("nullable", True)) else "NOT NULL"
        comment = str(column_meta.get("comment") or normalized_column).replace("'", "''")
        alter_clauses.append(f"MODIFY COLUMN `{normalized_column}` {target_sql} {null_sql} COMMENT '{comment}'")
    return alter_clauses


def ensure_import_text_capacity(table_name: str, schema_rows: List[Dict[str, object]], mapped_df: pd.DataFrame) -> None:
    if IS_SQLITE:
        return
    alter_clauses = build_text_capacity_alter_clauses(
        schema_rows,
        {
            str(column_name or "").strip(): mapped_df[str(column_name or "").strip()].tolist()
            for column_name in mapped_df.columns
            if str(column_name or "").strip() in mapped_df.columns
        },
    )

    if not alter_clauses:
        return

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"ALTER TABLE `{table_name}`\n  " + ",\n  ".join(alter_clauses))
        conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_import_row_text_capacity(
    table_name: str,
    schema_rows: List[Dict[str, object]],
    db_columns: List[str],
    rows: List[Tuple[object, ...]],
    only_columns: Set[str] | None = None,
) -> None:
    if IS_SQLITE:
        return
    values_by_column: Dict[str, List[object]] = {}
    selected_columns = {str(column or "").strip() for column in (only_columns or set()) if str(column or "").strip()}
    for index, column_name in enumerate(db_columns):
        normalized_column = str(column_name or "").strip()
        if selected_columns and normalized_column not in selected_columns:
            continue
        values_by_column[normalized_column] = [row[index] for row in rows]

    alter_clauses = build_text_capacity_alter_clauses(schema_rows, values_by_column)

    if not alter_clauses:
        return

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"ALTER TABLE `{table_name}`\n  " + ",\n  ".join(alter_clauses))
        conn.commit()
    finally:
        cur.close()
        conn.close()


def choose_preserved_column_type(existing_column_type: str, desired_column_type: str) -> str:
    existing = str(existing_column_type or "").strip().lower()
    desired = str(desired_column_type or "").strip().lower()
    if not existing:
        return desired_column_type

    desired_match = re.fullmatch(r"(char|varchar)\((\d+)\)", desired)
    existing_match = re.fullmatch(r"(char|varchar)\((\d+)\)", existing)
    if desired_match and existing_match and int(existing_match.group(2)) > int(desired_match.group(2)):
        return existing_column_type

    text_order = {"text": 1, "mediumtext": 2, "longtext": 3}
    if desired in text_order and existing in text_order and text_order[existing] > text_order[desired]:
        return existing_column_type

    return desired_column_type


def get_import_fallback_value(column_meta: Dict[str, object]) -> object:
    default_value = column_meta.get("default")
    if default_value is not None:
        return default_value
    if column_meta.get("nullable"):
        return None

    data_type = str(column_meta.get("data_type") or "").strip().lower()
    if data_type in {"char", "varchar", "text", "tinytext", "mediumtext", "longtext", "enum", "set"}:
        return ""
    return None


def is_numeric_timestamp_column(column_meta: Dict[str, object]) -> bool:
    data_type = str(column_meta.get("data_type") or "").strip().lower()
    if data_type not in {"decimal", "numeric", "int", "integer", "bigint", "smallint", "tinyint", "mediumint"}:
        return False

    column_name = str(column_meta.get("name") or "").strip().upper()
    column_comment = str(column_meta.get("comment") or "").strip().upper()
    if (
        "TIMESTMP" in column_name
        or "TIMESTAMP" in column_name
        or "TSTP" in column_name
        or "TIME STAMP" in column_comment
        or "TSTP" in column_comment
    ):
        return True
    return column_name.endswith("_AT") and get_decimal_scale(column_meta) == 0


def try_format_datetime_like_numeric(value: object) -> object:
    if isinstance(value, datetime):
        return int(value.strftime("%Y%m%d%H%M%S"))
    if isinstance(value, pd.Timestamp):
        return int(value.to_pydatetime().strftime("%Y%m%d%H%M%S"))
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return value
        try:
            parsed = pd.to_datetime(candidate, errors="raise")
        except Exception:
            return value
        if pd.isna(parsed):
            return None
        if isinstance(parsed, pd.Timestamp):
            parsed = parsed.to_pydatetime()
        return int(parsed.strftime("%Y%m%d%H%M%S"))
    return value


def get_decimal_scale(column_meta: Dict[str, object]) -> int | None:
    column_type = str(column_meta.get("column_type") or "").strip().lower()
    match = re.match(r"decimal\(\s*\d+\s*,\s*(\d+)\s*\)", column_type)
    if not match:
        return None
    return int(match.group(1))


def normalize_numeric_import_value(value: object, column_meta: Dict[str, object]) -> object:
    data_type = str(column_meta.get("data_type") or "").strip().lower()
    if is_numeric_timestamp_column(column_meta):
        value = try_format_datetime_like_numeric(value)
        if value is None:
            return None

    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
    else:
        candidate = value

    integer_like_types = {"int", "integer", "bigint", "smallint", "tinyint", "mediumint"}
    if data_type in integer_like_types:
        try:
            return int(Decimal(str(candidate)))
        except (InvalidOperation, TypeError, ValueError):
            return candidate

    if data_type in {"decimal", "numeric"}:
        try:
            decimal_value = Decimal(str(candidate))
        except (InvalidOperation, TypeError, ValueError):
            return candidate

        if get_decimal_scale(column_meta) == 0:
            return int(decimal_value)
        return decimal_value

    if data_type in {"float", "double", "real"}:
        try:
            return float(candidate)
        except (TypeError, ValueError):
            return candidate

    return candidate


def normalize_temporal_import_value(value: object, column_meta: Dict[str, object]) -> object:
    data_type = str(column_meta.get("data_type") or "").strip().lower()
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
    else:
        candidate = value

    try:
        parsed = pd.to_datetime(candidate, errors="raise")
    except Exception:
        return candidate

    if pd.isna(parsed):
        return None

    if data_type == "date":
        return parsed.strftime("%Y-%m-%d")
    if data_type in {"datetime", "timestamp"}:
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    if data_type == "year":
        return int(parsed.strftime("%Y"))
    return candidate


IMPORT_KEEP_CASE_COLUMNS = {
    "TXTLG",
    "TXTMD",
    "TXTSH",
    "DDTEXT",
    "REPTEXT",
    "SCRTEXT_S",
    "SCRTEXT_M",
    "SCRTEXT_L",
}

RAW_IMPORT_TABLES = {"rstranstepcnst", "rstransteprout"}

FAST_IMPORT_UPSERT_TABLES = {"dd03l"}
FAST_IMPORT_SKIP_EXISTING_KEY_SCAN_ROW_THRESHOLD = 20000
IMPORT_READ_CHUNK_SIZE = max(250, int(os.getenv("IMPORT_READ_CHUNK_SIZE", "2000")))
DD03L_FAST_WRITE_BATCH_SIZE = max(200, int(os.getenv("DD03L_FAST_WRITE_BATCH_SIZE", "1000")))


def normalize_import_param(value: object, column_meta: Dict[str, object], table_name: str = "") -> object:
    table_name = str(table_name or "").strip().lower()
    if table_name in RAW_IMPORT_TABLES:
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        return value

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    data_type = str(column_meta.get("data_type") or "").strip().lower()
    column_name = str(column_meta.get("name") or "").strip().upper()
    should_preserve_case = table_name in {"rsoadsot", "rsdst", "rsdssegfdt", "rsksnewt", "rsksfieldnewt", "dd02t", "dd03t", "dd04t", "rsdiobjt"} and column_name in IMPORT_KEEP_CASE_COLUMNS

    if isinstance(value, str):
        normalized = value if should_preserve_case else value.upper()
        if normalized.strip() in {"NAN", "NAT", "NONE", "NULL"}:
            return None
        value = normalized.strip() if data_type in {"decimal", "numeric", "int", "integer", "bigint", "smallint", "tinyint", "mediumint", "float", "double", "real", "date", "datetime", "timestamp", "time", "year"} else normalized

    if data_type in {"decimal", "numeric", "int", "integer", "bigint", "smallint", "tinyint", "mediumint", "float", "double", "real"}:
        return normalize_numeric_import_value(value, column_meta)

    if data_type in {"date", "datetime", "timestamp", "year"}:
        return normalize_temporal_import_value(value, column_meta)

    return value


def normalize_type_code(raw_type: str | None) -> str:
    txt = str(raw_type or "").strip().upper()
    if not txt:
        return "UNKNOWN"
    return txt


def normalize_iobj_subtype(raw_subtype: object) -> str:
    normalized = str(raw_subtype or "").strip().upper()
    return normalized if normalized in {"ATTR", "TEXT"} else ""


def row_has_transformation_logic(start_routine: object, end_routine: object, expert: object) -> bool:
    return any(str(value or "").strip() for value in (start_routine, end_routine, expert))


def extract_transformation_logic_details(start_routine: object, end_routine: object, expert: object) -> List[Dict[str, str]]:
    details: List[Dict[str, str]] = []
    for kind, value in (("start", start_routine), ("end", end_routine), ("expert", expert)):
        txt = str(value or "").strip()
        if txt:
            details.append({"kind": kind, "id": txt})
    return details


def make_path_node_key(object_name: str, object_type: str = "", object_system: str = "") -> str:
    normalized_name = normalize_bw_object_lookup(object_name)
    normalized_type = normalize_type_code(object_type)
    normalized_system = normalize_bw_object_lookup(object_system)
    if normalized_type in {"RSDS", "DATASOURCE", "SOURCE"} and normalized_system:
        return f"RSDS::{normalized_name}::{normalized_system}"
    return normalized_name


def get_segment_path_value(segment: PathSegmentRequest, field_name: str, legacy_name: str = "") -> Any:
    value = getattr(segment, field_name, None)
    if value is not None and value != "":
        return value
    if legacy_name:
        return getattr(segment, legacy_name, None)
    return None


def get_segment_source_name(segment: PathSegmentRequest) -> str:
    return str(get_segment_path_value(segment, "SOURCE", "source") or "").strip()


def get_segment_target_name(segment: PathSegmentRequest) -> str:
    return str(get_segment_path_value(segment, "TARGETNAME", "target") or "").strip()


def get_segment_source_type(segment: PathSegmentRequest) -> str:
    return str(get_segment_path_value(segment, "SOURCETYPE", "source_type") or "").strip()


def get_segment_target_type(segment: PathSegmentRequest) -> str:
    return str(get_segment_path_value(segment, "TARGETTYPE", "target_type") or "").strip()


def get_segment_source_subtype(segment: PathSegmentRequest) -> str:
    return normalize_iobj_subtype(get_segment_path_value(segment, "SOURCESUBTYPE", "source_subtype"))


def get_segment_target_subtype(segment: PathSegmentRequest) -> str:
    return normalize_iobj_subtype(get_segment_path_value(segment, "TARGETSUBTYPE", "target_subtype"))


def get_segment_source_system(segment: PathSegmentRequest) -> str:
    return str(get_segment_path_value(segment, "SOURCESYS", "source_system") or "").strip()


def get_segment_target_system(segment: PathSegmentRequest) -> str:
    return str(get_segment_path_value(segment, "TARGETSYSTEM", "target_system") or "").strip()


def get_segment_tran_ids(segment: PathSegmentRequest) -> List[str]:
    value = get_segment_path_value(segment, "TRANIDS", "tran_ids")
    return list(value or [])


def build_path_logic_entry_payload(entry: Dict[str, object]) -> Dict[str, object]:
    return {
        "TRANID": str(entry.get("tran_id") or "").strip(),
        "RULEID": int(entry.get("rule_id") or 0),
        "STEPID": int(entry.get("step_id") or 0),
        "KIND": str(entry.get("kind") or "").strip(),
        "TITLE": str(entry.get("title") or "").strip(),
        "CONTENT_RAW": str(entry.get("content_raw") or ""),
        "CONTENT_DISPLAY": str(entry.get("content_display") or ""),
        "LANGUAGE": str(entry.get("language") or "").strip(),
        "ON_HANA": str(entry.get("on_hana") or "").strip(),
        "CODE_ID": str(entry.get("code_id") or "").strip(),
        "SOURCE_TABLE": str(entry.get("source_table") or "").strip(),
        "IS_CONSTANT": bool(entry.get("is_constant")),
    }


def build_path_diagnostics_payload(diagnostics: Dict[str, object]) -> Dict[str, object]:
    return {
        "OBJECT": str(diagnostics.get("object") or "").strip(),
        "OBJECT_DISPLAY_NAME": str(diagnostics.get("object_display_name") or "").strip(),
        "OBJECT_TYPE": str(diagnostics.get("object_type") or "").strip(),
        "OBJECT_SYSTEM": str(diagnostics.get("object_system") or "").strip(),
        "FIELD_KEY": str(diagnostics.get("field_key") or "").strip(),
        "UNIQUE_FIELD_COUNT": int(diagnostics.get("unique_field_count") or 0),
        "COMPARISON_AVAILABLE": bool(diagnostics.get("comparison_available")),
        "INVENTORY_FIELD_COUNT": diagnostics.get("inventory_field_count"),
        "DIFFERENCE_COUNT": diagnostics.get("difference_count"),
        "MISSING_COUNT": diagnostics.get("missing_count"),
        "EXTRA_COUNT": diagnostics.get("extra_count"),
        "HAS_DIFFERENCE": bool(diagnostics.get("has_difference")),
        "INVENTORY_ORIGIN": str(diagnostics.get("inventory_origin") or "").strip(),
        "INVENTORY_LABEL": str(diagnostics.get("inventory_label") or "").strip(),
        "ADSO_TABLE_SUFFIX": str(diagnostics.get("adso_table_suffix") or "").strip(),
        "ADSO_TABLE_NAME": str(diagnostics.get("adso_table_name") or "").strip(),
    }


def build_path_mapping_row_payload(row: Dict[str, object]) -> Dict[str, object]:
    logic_entries = [
        build_path_logic_entry_payload(entry)
        for entry in (row.get("logic_entries") or [])
        if isinstance(entry, dict)
    ]
    return {
        "TRANID": str(row.get("tran_id") or "").strip(),
        "RULEID": int(row.get("rule_id") or 0),
        "STEPID": int(row.get("step_id") or 0),
        "SEGID": str(row.get("seg_id") or "").strip(),
        "PAIR_INDEX": int(row.get("pair_index") or 0),
        "RULEPOSIT": str(row.get("ruleposit") or "").strip(),
        "SOURCE_FIELD": str(row.get("source_field") or "").strip(),
        "TARGET_FIELD": str(row.get("target_field") or "").strip(),
        "SOURCE_FIELDTYPE": str(row.get("source_fieldtype") or "").strip(),
        "TARGET_FIELDTYPE": str(row.get("target_fieldtype") or "").strip(),
        "SOURCE_TEXT": str(row.get("source_text") or "").strip(),
        "TARGET_TEXT": str(row.get("target_text") or "").strip(),
        "SOURCE_DATATYPE": str(row.get("source_datatype") or "").strip(),
        "SOURCE_LENGTH": row.get("source_length"),
        "SOURCE_DECIMALS": row.get("source_decimals"),
        "TARGET_DATATYPE": str(row.get("target_datatype") or "").strip(),
        "TARGET_LENGTH": row.get("target_length"),
        "TARGET_DECIMALS": row.get("target_decimals"),
        "SOURCE_KEY": str(row.get("source_key") or "").strip(),
        "TARGET_KEY": str(row.get("target_key") or row.get("target_keyflag") or "").strip(),
        "SOURCE_FIELD_ORIGIN": str(row.get("source_field_origin") or "").strip(),
        "TARGET_FIELD_ORIGIN": str(row.get("target_field_origin") or "").strip(),
        "SOURCE_FIELD_MATCHED": int(row.get("source_field_matched") or 0),
        "TARGET_FIELD_MATCHED": int(row.get("target_field_matched") or 0),
        "RULE": str(row.get("rule") or row.get("rule_type") or "").strip(),
        "GROUPTYPE": str(row.get("group_type") or row.get("GROUPTYPE") or "").strip(),
        "AGGR": str(row.get("aggr") or "").strip(),
        "ROW_KIND": str(row.get("row_kind") or "").strip(),
        "LOGIC_ENTRIES": logic_entries,
        "HAS_LOGIC_ENTRY": bool(row.get("has_logic_entry")),
    }


def build_path_step_logic_group_payload(group: Dict[str, object]) -> Dict[str, object]:
    return {
        "TRANID": str(group.get("tran_id") or "").strip(),
        "ENTRY_COUNT": int(group.get("entry_count") or 0),
        "ENTRIES": [
            build_path_logic_entry_payload(entry)
            for entry in (group.get("entries") or [])
            if isinstance(entry, dict)
        ],
    }


def append_or_update_edge(
    edges: List[Dict[str, object]],
    edge_by_key: Dict[Tuple[str, str, str, str, str, str], Dict[str, object]],
    source_name: str,
    target_name: str,
    source_type: str,
    target_type: str,
    source_subtype: str = "",
    target_subtype: str = "",
    source_system: str = "",
    target_system: str = "",
    tran_id: object = None,
    start_routine: object = None,
    end_routine: object = None,
    expert: object = None,
    source_node_key: str = "",
    target_node_key: str = "",
) -> None:
    normalized_source_node_key = str(source_node_key or source_name or "").strip()
    normalized_target_node_key = str(target_node_key or target_name or "").strip()
    normalized_source_subtype = normalize_iobj_subtype(source_subtype)
    normalized_target_subtype = normalize_iobj_subtype(target_subtype)
    edge_key = (normalized_source_node_key, normalized_target_node_key, source_type, target_type, normalized_source_subtype, normalized_target_subtype)
    logic_details = extract_transformation_logic_details(start_routine, end_routine, expert)
    tran_id_text = str(tran_id or "").strip()
    existing = edge_by_key.get(edge_key)
    if existing is None:
        edge = {
            "SOURCE": source_name,
            "TARGETNAME": target_name,
            "SOURCETYPE": source_type,
            "TARGETTYPE": target_type,
            "SOURCESUBTYPE": normalized_source_subtype,
            "TARGETSUBTYPE": normalized_target_subtype,
            "SOURCESYS": str(source_system or "").strip(),
            "TARGETSYSTEM": str(target_system or "").strip(),
            "TRANIDS": [tran_id_text] if tran_id_text else [],
            "LOGIC_IDS": [detail["id"] for detail in logic_details],
            "HAS_LOGIC": bool(logic_details),
            "logic_details": logic_details,
            "_source_node_key": normalized_source_node_key,
            "_target_node_key": normalized_target_node_key,
        }
        edges.append(edge)
        edge_by_key[edge_key] = edge
        return

    if tran_id_text:
        existing_tran_ids_upper = existing.setdefault("TRANIDS", [])
        if tran_id_text not in existing_tran_ids_upper:
            existing_tran_ids_upper.append(tran_id_text)

    if logic_details:
        existing["HAS_LOGIC"] = True
        existing_details = existing.setdefault("logic_details", [])
        existing_keys = {(str(detail.get("kind") or ""), str(detail.get("id") or "")) for detail in existing_details}
        for detail in logic_details:
            detail_key = (str(detail.get("kind") or ""), str(detail.get("id") or ""))
            if detail_key in existing_keys:
                continue
            existing_details.append(detail)
            existing_keys.add(detail_key)
        existing["LOGIC_IDS"] = [detail.get("id", "") for detail in existing_details if str(detail.get("id") or "").strip()]


def fetch_bw_object_name_map(node_names: List[str]) -> Dict[str, str]:
    normalized_names: List[str] = []
    seen_names: Set[str] = set()
    for name in node_names:
        normalized_name = normalize_bw_object_lookup(name)
        if not normalized_name or normalized_name in seen_names:
            continue
        seen_names.add(normalized_name)
        normalized_names.append(normalized_name)
    if not normalized_names:
        return {}

    conn = get_conn()
    cur = conn.cursor()
    object_name_map: Dict[str, str] = {}
    try:
        if IS_SQLITE:
            cur.execute(
                """
                SELECT BW_OBJECT_NORM, MAX(NAME_EN) AS NAME_EN
                FROM bw_object_name
                WHERE BW_OBJECT_NORM IN (SELECT value FROM json_each(?))
                GROUP BY BW_OBJECT_NORM
                """,
                (json.dumps(normalized_names),)
            )
        else:
            placeholders = ", ".join(["%s"] * len(normalized_names))
            cur.execute(
                f"""
                SELECT UPPER(TRIM(BW_OBJECT_NORM)) AS BW_OBJECT_NORM, MAX(NAME_EN) AS NAME_EN
                FROM bw_object_name
                WHERE UPPER(TRIM(BW_OBJECT_NORM)) IN ({placeholders})
                GROUP BY UPPER(TRIM(BW_OBJECT_NORM))
                """,
                tuple(normalized_names),
            )
        for bw_object, object_name in cur.fetchall():
            key = normalize_bw_object_lookup(bw_object)
            value = str(object_name or "").strip()
            if key and value:
                object_name_map[key] = value
    except (sqlite3.Error, mysql.connector.Error):
        return {}
    finally:
        cur.close()
        conn.close()

    return object_name_map


def fetch_path_selection_metadata_by_tran_id(tran_id: str) -> Dict[str, str]:
    normalized_tran_id = str(tran_id or "").strip()
    if not normalized_tran_id:
        return {}

    active_metadata = fetch_active_tran_metadata().get(normalized_tran_id)
    if active_metadata:
        return {
            "tran_id": normalized_tran_id,
            "source_type": normalize_type_code(active_metadata.get("source_type")),
            "target_type": normalize_type_code(active_metadata.get("target_type")),
            "source_subtype": normalize_iobj_subtype(active_metadata.get("source_subtype")),
            "target_subtype": normalize_iobj_subtype(active_metadata.get("target_subtype")),
            "source_object": normalize_bw_object_lookup(active_metadata.get("source_object")),
            "source_system": normalize_bw_object_lookup(active_metadata.get("source_system")),
            "target_object": normalize_bw_object_lookup(active_metadata.get("target_object")),
            "target_system": normalize_bw_object_lookup(active_metadata.get("target_system")),
        }

    if not table_exists(RSTRAN_MAPPING_RULE_FULL_TABLE):
        return {}

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
                 SELECT tran_id, source_object, source_system, target_object, target_system, source_type, target_type,
                     COALESCE(source_subtype, '') AS source_subtype,
                     COALESCE(target_subtype, '') AS target_subtype
            FROM `{RSTRAN_MAPPING_RULE_FULL_TABLE}`
            WHERE UPPER(TRIM(tran_id)) = UPPER(%s)
            LIMIT 1
            """,
            (normalized_tran_id,),
        )
        row = cur.fetchone() or {}
    finally:
        cur.close()
        conn.close()

    if not row:
        return {}

    return {
        "tran_id": normalized_tran_id,
        "source_type": normalize_type_code(row.get("source_type")),
        "target_type": normalize_type_code(row.get("target_type")),
        "source_subtype": normalize_iobj_subtype(row.get("source_subtype")),
        "target_subtype": normalize_iobj_subtype(row.get("target_subtype")),
        "source_object": normalize_bw_object_lookup(row.get("source_object")),
        "source_system": normalize_bw_object_lookup(row.get("source_system")),
        "target_object": normalize_bw_object_lookup(row.get("target_object")),
        "target_system": normalize_bw_object_lookup(row.get("target_system")),
    }


def fetch_path_selection_metadata_batch(tran_ids: List[str]) -> Dict[str, Dict[str, str]]:
    """Fetch metadata for multiple TRANIDs in a single query."""
    normalized_tran_ids = sorted({
        str(tran_id or "").strip()
        for tran_id in tran_ids
        if str(tran_id or "").strip()
    })
    
    if not normalized_tran_ids:
        return {}
    
    result: Dict[str, Dict[str, str]] = {}
    
    # Try to get from active metadata first (batch)
    active_metadata = fetch_active_tran_metadata()
    for tran_id in normalized_tran_ids:
        if tran_id in active_metadata:
            meta = active_metadata[tran_id]
            result[tran_id] = {
                "tran_id": tran_id,
                "source_type": normalize_type_code(meta.get("source_type")),
                "target_type": normalize_type_code(meta.get("target_type")),
                "source_subtype": normalize_iobj_subtype(meta.get("source_subtype")),
                "target_subtype": normalize_iobj_subtype(meta.get("target_subtype")),
                "source_object": normalize_bw_object_lookup(meta.get("source_object")),
                "source_system": normalize_bw_object_lookup(meta.get("source_system")),
                "target_object": normalize_bw_object_lookup(meta.get("target_object")),
                "target_system": normalize_bw_object_lookup(meta.get("target_system")),
            }
    
    # For remaining TRANIDs, query from materialized table
    remaining_tran_ids = [tid for tid in normalized_tran_ids if tid not in result]
    if remaining_tran_ids and table_exists(RSTRAN_MAPPING_RULE_FULL_TABLE):
        placeholders = ", ".join(["%s"] * len(remaining_tran_ids))
        conn = get_conn()
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute(
                f"""
                  SELECT DISTINCT tran_id, source_object, source_system, target_object, target_system, source_type, target_type,
                      COALESCE(source_subtype, '') AS source_subtype,
                      COALESCE(target_subtype, '') AS target_subtype
                FROM `{RSTRAN_MAPPING_RULE_FULL_TABLE}`
                WHERE UPPER(TRIM(tran_id)) IN ({placeholders})
                """,
                tuple(remaining_tran_ids),
            )
            rows = cur.fetchall()
            for row in rows:
                tran_id = str(row.get("tran_id") or "").strip()
                if tran_id and tran_id not in result:
                    result[tran_id] = {
                        "tran_id": tran_id,
                        "source_type": normalize_type_code(row.get("source_type")),
                        "target_type": normalize_type_code(row.get("target_type")),
                        "source_subtype": normalize_iobj_subtype(row.get("source_subtype")),
                        "target_subtype": normalize_iobj_subtype(row.get("target_subtype")),
                        "source_object": normalize_bw_object_lookup(row.get("source_object")),
                        "source_system": normalize_bw_object_lookup(row.get("source_system")),
                        "target_object": normalize_bw_object_lookup(row.get("target_object")),
                        "target_system": normalize_bw_object_lookup(row.get("target_system")),
                    }
        finally:
            cur.close()
            conn.close()
    
    return result


def search_candidate_paths_by_tran_id(tran_id: str) -> Dict[str, object]:
    metadata = fetch_path_selection_metadata_by_tran_id(tran_id)
    normalized_tran_id = str(tran_id or "").strip()
    if not metadata:
        raise HTTPException(status_code=404, detail="No path selection data found for the provided transformation ID")

    source_name = str(metadata.get("source_object") or "").strip()
    source_system = str(metadata.get("source_system") or "").strip()
    target_name = str(metadata.get("target_object") or "").strip()
    source_type = normalize_type_code(metadata.get("source_type"))
    target_type = normalize_type_code(metadata.get("target_type"))
    source_subtype = normalize_iobj_subtype(metadata.get("source_subtype"))
    target_subtype = normalize_iobj_subtype(metadata.get("target_subtype"))
    start_name = resolve_rstran_start_name(source_name, source_system, source_type)
    node_name_map = fetch_bw_object_name_map([start_name, target_name])

    candidate_path = {
        "id": "path-1",
        "index": 1,
        "SEGMENT_COUNT": 1,
        "NODE_SEQUENCE": [
            {
                "id": start_name,
                "OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(start_name), ""),
                "TYPE": source_type,
                "SUBTYPE": source_subtype,
                "SOURCESYS": source_system,
            },
            {
                "id": target_name,
                "OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(target_name), ""),
                "TYPE": target_type,
                "SUBTYPE": target_subtype,
            },
        ],
        "WAYPOINTS_HIT": [],
        "SEGMENTS": [
            {
                "SOURCE": start_name,
                "TARGETNAME": target_name,
                "SOURCETYPE": source_type,
                "TARGETTYPE": target_type,
                "SOURCESUBTYPE": source_subtype,
                "TARGETSUBTYPE": target_subtype,
                "SOURCESYS": source_system,
                "TARGETSYSTEM": str(metadata.get("target_system") or "").strip(),
                "SOURCE_OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(start_name), ""),
                "TARGET_OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(target_name), ""),
                "TRANIDS": [normalized_tran_id],
                "LOGIC_IDS": [],
                "HAS_LOGIC": False,
            }
        ],
    }

    return {
        "SOURCE": source_name,
        "SOURCESYS": source_system,
        "TARGETNAME": target_name,
        "WAYPOINTS": [],
        "CANDIDATE_COUNT": 1,
        "RESOLVED_START_NAME": start_name,
        "CANDIDATE_PATHS": [candidate_path],
        "TRUNCATED": False,
        "SEARCH_STATS": {
            "STATES_VISITED": 1,
            "EDGE_COUNT": 1,
            "NODE_COUNT": 2,
        },
    }


def resolve_rstran_start_name(start_name: str, start_source: str = "", start_type: str = "") -> str:
    normalized_start_name = normalize_bw_object_lookup(start_name)
    normalized_source = normalize_bw_object_lookup(start_source)
    if not normalized_start_name:
        return ""

    if not table_exists("rstran"):
        return normalized_start_name

    rstran_columns = set(get_table_columns("rstran"))
    source_col = "SOURCE" if "SOURCE" in rstran_columns else ("DATASOURCE" if "DATASOURCE" in rstran_columns else "")
    if not source_col:
        return normalized_start_name

    conn = get_conn()
    cur = conn.cursor()
    try:
        if normalized_source and "SOURCESYS" in rstran_columns:
            source_placeholder = "?" if IS_SQLITE else "%s"
            cur.execute(
                f"""
                SELECT 1
                FROM rstran
                WHERE UPPER(TRIM({source_col})) = UPPER({source_placeholder})
                  AND UPPER(TRIM(COALESCE(SOURCESYS, ''))) = UPPER({source_placeholder})
                LIMIT 1
                """,
                (normalized_start_name, normalized_source),
            )
            if cur.fetchone():
                return normalized_start_name

        name_placeholder = "?" if IS_SQLITE else "%s"
        cur.execute(
            f"""
            SELECT 1
            FROM rstran
            WHERE UPPER(TRIM({source_col})) = UPPER({name_placeholder})
               OR UPPER(TRIM(TARGETNAME)) = UPPER({name_placeholder})
            LIMIT 1
            """,
            (normalized_start_name, normalized_start_name),
        )
        if cur.fetchone():
            return normalized_start_name
    finally:
        cur.close()
        conn.close()

    return normalized_start_name


def path_contains_waypoints(node_ids: List[str], waypoints: List[str]) -> bool:
    if not waypoints:
        return True

    waypoint_index = 0
    for node_id in node_ids:
        if waypoint_index >= len(waypoints):
            break
        if normalize_bw_object_lookup(node_id) == waypoints[waypoint_index]:
            waypoint_index += 1
    return waypoint_index >= len(waypoints)


def is_rsds_source_name(source_name: str) -> bool:
    normalized_name = normalize_bw_object_lookup(source_name)
    if not normalized_name:
        return False

    conn = get_conn()
    cur = conn.cursor()
    try:
        if table_exists("rsds"):
            cur.execute(
                """
                SELECT 1
                FROM rsds
                WHERE UPPER(TRIM(DATASOURCE)) = UPPER(%s)
                LIMIT 1
                """,
                (normalized_name,),
            )
            if cur.fetchone():
                return True

        cur.execute(
            """
            SELECT 1
            FROM bw_object_name
            WHERE UPPER(TRIM(BW_OBJECT_NORM)) = UPPER(%s)
              AND UPPER(TRIM(COALESCE(BW_OBJECT_TYPE, ''))) = 'RSDS'
            LIMIT 1
            """,
            (normalized_name,),
        )
        if cur.fetchone():
            return True

        return False
    finally:
        cur.close()
        conn.close()


def search_candidate_paths(
    source_name: str,
    source_system: str,
    target_name: str,
    waypoints: List[str] | None = None,
    max_paths: int = 12,
    max_depth: int = 10,
    max_states: int = 25000,
) -> Dict[str, object]:
    normalized_source_name = normalize_bw_object_lookup(source_name)
    normalized_source_system = normalize_bw_object_lookup(source_system)
    normalized_target_name = normalize_bw_object_lookup(target_name)
    normalized_waypoints = [normalize_bw_object_lookup(item) for item in (waypoints or []) if normalize_bw_object_lookup(item)]

    if not table_exists("rstran"):
        raise HTTPException(status_code=400, detail="路径查询依赖 RSTRAN 数据，请先导入 rstran 表")

    rstran_columns = set(get_table_columns("rstran"))
    source_col = "SOURCE" if "SOURCE" in rstran_columns else ("DATASOURCE" if "DATASOURCE" in rstran_columns else "")
    source_subtype_sql = ", SOURCESUBTYPE" if "SOURCESUBTYPE" in rstran_columns else ", '' AS SOURCESUBTYPE"
    target_subtype_sql = ", TARGETSUBTYPE" if "TARGETSUBTYPE" in rstran_columns else ", '' AS TARGETSUBTYPE"
    if not source_col:
        raise HTTPException(status_code=400, detail="RSTRAN 缺少 SOURCE 字段，无法按新规则生成路径")

    start_name = resolve_rstran_start_name(normalized_source_name, normalized_source_system, "")
    target_key = make_path_node_key(normalized_target_name, "", "")

    total_start = time.time()
    conn = get_conn()
    cur = conn.cursor()
    adjacency: Dict[str, List[Dict[str, object]]] = {}
    edge_by_key: Dict[Tuple[str, str, str, str, str, str], Dict[str, object]] = {}
    edges: List[Dict[str, object]] = []
    node_types: Dict[str, str] = {}
    node_labels: Dict[str, str] = {}
    node_systems: Dict[str, str] = {}
    exact_start_node_keys: List[str] = []
    exact_start_node_key_set: Set[str] = set()

    try:
        step1_start = time.time()
        cur.execute(
            f"""
                        SELECT {source_col} AS SOURCE, SOURCESYS, TARGETNAME, SOURCETYPE, TARGETTYPE{source_subtype_sql}{target_subtype_sql}, TRANID, STARTROUTINE, ENDROUTINE, EXPERT
            FROM rstran
            WHERE OBJVERS = 'A'
              AND {source_col} IS NOT NULL
              AND TARGETNAME IS NOT NULL
            ORDER BY {source_col}, TRANID
            """
        )
        step1_end = time.time()

        step3_start = time.time()
        for row in cur.fetchall():
            row_source_type = normalize_type_code(row[3])
            row_target_type = normalize_type_code(row[4])
            row_source_subtype = normalize_iobj_subtype(row[5])
            row_target_subtype = normalize_iobj_subtype(row[6])

            source_value = str(row[0] or "").strip()
            source_system_value = ""
            if row_source_type == "RSDS":
                source_value, source_system_value = parse_rsds_identity(source_value, row[1])

            target_value = str(row[2] or "").strip()
            target_system_value = ""
            if row_target_type == "RSDS":
                target_value, target_system_value = parse_rsds_identity(target_value, None)

            if not source_value or not target_value:
                continue

            source_node_key = make_path_node_key(source_value, row_source_type, source_system_value)
            target_node_key = make_path_node_key(target_value, row_target_type, target_system_value)
            node_labels[source_node_key] = source_value
            node_labels[target_node_key] = target_value
            node_types[source_node_key] = row_source_type or node_types.get(source_node_key, "UNKNOWN")
            node_types[target_node_key] = row_target_type or node_types.get(target_node_key, "UNKNOWN")
            if source_system_value:
                node_systems[source_node_key] = source_system_value
            if target_system_value:
                node_systems[target_node_key] = target_system_value

            if source_value == normalized_source_name and source_system_value == normalized_source_system and source_node_key not in exact_start_node_key_set:
                exact_start_node_key_set.add(source_node_key)
                exact_start_node_keys.append(source_node_key)

            edge_key = (source_node_key, target_node_key, row_source_type, row_target_type)
            edge_key = (source_node_key, target_node_key, row_source_type, row_target_type, row_source_subtype, row_target_subtype)
            is_new_edge = edge_key not in edge_by_key

            append_or_update_edge(
                edges,
                edge_by_key,
                source_value,
                target_value,
                row_source_type,
                row_target_type,
                row_source_subtype,
                row_target_subtype,
                source_system_value,
                target_system_value,
                row[7],
                row[8],
                row[9],
                row[10],
                source_node_key=source_node_key,
                target_node_key=target_node_key,
            )

            if is_new_edge and source_value != target_value:
                adjacency.setdefault(source_node_key, []).append(edge_by_key[edge_key])
        step3_end = time.time()
    finally:
        cur.close()
        conn.close()

    step4_start = time.time()
    candidate_paths: List[Dict[str, object]] = []
    queue: deque[Tuple[List[str], List[str], List[Dict[str, object]]]] = deque()
    for start_node_key in exact_start_node_keys:
        queue.append(([start_node_key], [node_labels.get(start_node_key, start_name)], []))
    seen_signatures: Set[str] = set()
    states_visited = 0
    truncated = False

    while queue and len(candidate_paths) < max_paths:
        node_key_path, node_label_path, edge_path = queue.popleft()
        states_visited += 1
        if states_visited > max_states:
            truncated = True
            break

        current_node_key = node_key_path[-1]
        if current_node_key == target_key and edge_path:
            if not path_contains_waypoints(node_label_path, normalized_waypoints):
                continue
            signature = "=>".join(node_key_path)
            if signature in seen_signatures:
                continue
            seen_signatures.add(signature)
            candidate_paths.append(
                {
                    "id": f"path-{len(candidate_paths) + 1}",
                    "node_ids": node_label_path,
                    "node_keys": node_key_path,
                    "segment_count": len(edge_path),
                    "segments": [dict(segment) for segment in edge_path],
                }
            )
            continue

        if len(edge_path) >= max_depth:
            continue

        for edge in adjacency.get(current_node_key, []):
            next_node_key = str(edge.get("_target_node_key") or "").strip()
            next_node_label = str(edge.get("TARGETNAME") or edge.get("target") or "").strip()
            if not next_node_key or next_node_key in node_key_path:
                continue
            queue.append((node_key_path + [next_node_key], node_label_path + [next_node_label], edge_path + [edge]))

    step4_end = time.time()

    step5_start = time.time()
    node_name_map = fetch_bw_object_name_map(
        [node_id for path in candidate_paths for node_id in path.get("node_ids", [])]
    )
    step5_end = time.time()

    step6_start = time.time()
    enriched_paths: List[Dict[str, object]] = []
    for index, path in enumerate(candidate_paths, start=1):
        node_ids = [str(node_id or "").strip() for node_id in path.get("node_ids", []) if str(node_id or "").strip()]
        node_keys = [str(node_key or "").strip() for node_key in path.get("node_keys", []) if str(node_key or "").strip()]
        segments = []
        for segment in path.get("segments", []):
            source_id = str(segment.get("SOURCE") or "").strip()
            target_id = str(segment.get("TARGETNAME") or "").strip()
            segments.append(
                {
                    "SOURCE": source_id,
                    "TARGETNAME": target_id,
                    "SOURCETYPE": str(segment.get("SOURCETYPE") or node_types.get(str(segment.get("_source_node_key") or "").strip(), "UNKNOWN")),
                    "TARGETTYPE": str(segment.get("TARGETTYPE") or node_types.get(str(segment.get("_target_node_key") or "").strip(), "UNKNOWN")),
                    "SOURCESYS": str(segment.get("SOURCESYS") or "").strip(),
                    "TARGETSYSTEM": str(segment.get("TARGETSYSTEM") or "").strip(),
                    "SOURCE_OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(source_id), ""),
                    "TARGET_OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(target_id), ""),
                    "TRANIDS": [str(item).strip() for item in (segment.get("TRANIDS") or []) if str(item).strip()],
                    "LOGIC_IDS": [str(item).strip() for item in (segment.get("LOGIC_IDS") or []) if str(item).strip()],
                    "HAS_LOGIC": bool(segment.get("HAS_LOGIC")),
                }
            )

        enriched_paths.append(
            {
                "id": str(path.get("id") or f"path-{index}"),
                "index": index,
                "SEGMENT_COUNT": int(path.get("segment_count") or len(segments)),
                "NODE_SEQUENCE": [
                    {
                        "id": node_id,
                        "OBJECT_NAME": node_name_map.get(normalize_bw_object_lookup(node_id), ""),
                        "TYPE": node_types.get(node_key, "UNKNOWN"),
                        "SOURCESYS": node_systems.get(node_key, ""),
                    }
                    for node_key, node_id in zip(node_keys, node_ids)
                ],
                "WAYPOINTS_HIT": [item for item in normalized_waypoints if item in {normalize_bw_object_lookup(node_id) for node_id in node_ids}],
                "SEGMENTS": segments,
            }
        )

    step6_end = time.time()
    total_end = time.time()

    return {
        "SOURCE": normalized_source_name,
        "SOURCESYS": normalized_source_system,
        "TARGETNAME": normalized_target_name,
        "WAYPOINTS": normalized_waypoints,
        "CANDIDATE_COUNT": len(enriched_paths),
        "RESOLVED_START_NAME": start_name,
        "CANDIDATE_PATHS": enriched_paths,
        "TRUNCATED": truncated,
        "SEARCH_STATS": {
            "STATES_VISITED": states_visited,
            "EDGE_COUNT": len(edges),
            "NODE_COUNT": len(node_types),
        },
    }


def get_tran_mapping_rule_select_sql(filter_by_tran: bool = True) -> str:
    tran_filter = "AND field.TRANID = %s" if filter_by_tran else ""
    return f"""
            WITH source_rows AS (
                SELECT
                    field.TRANID AS tran_id,
                    field.RULEID AS rule_id,
                    field.STEPID AS step_id,
                    field.SEGID AS seg_id,
                    field.FIELDNM AS source_field,
                    COALESCE(NULLIF(TRIM(field.FIELDTYPE), ''), '') AS source_fieldtype,
                    field.KEYFLAG AS source_keyflag,
                    field.RULEPOSIT AS source_ruleposit,
                    ROW_NUMBER() OVER (
                        PARTITION BY field.TRANID, field.RULEID, field.STEPID, field.SEGID
                        ORDER BY CAST(COALESCE(NULLIF(TRIM(field.RULEPOSIT), ''), '0') AS UNSIGNED), field.PARAMNM, field.FIELDNM
                    ) AS pair_index
                FROM rstranfield AS field
                                INNER JOIN rstranrule AS rule
                                    ON field.TRANID = rule.TRANID
                                 AND field.OBJVERS = rule.OBJVERS
                                 AND field.RULEID = rule.RULEID
                WHERE field.OBJVERS = 'A'
                  {tran_filter}
                  AND field.PARAMTYPE = '0'
                                                                        AND rule.GROUPTYPE IN ('N', 'S')
            ),
            target_rows AS (
                SELECT
                    field.TRANID AS tran_id,
                    field.RULEID AS rule_id,
                    field.STEPID AS step_id,
                    field.SEGID AS seg_id,
                    field.FIELDNM AS target_field,
                    COALESCE(NULLIF(TRIM(field.FIELDTYPE), ''), '') AS target_fieldtype,
                    field.KEYFLAG AS target_keyflag,
                    field.RULEPOSIT AS target_ruleposit,
                    ROW_NUMBER() OVER (
                        PARTITION BY field.TRANID, field.RULEID, field.STEPID, field.SEGID
                        ORDER BY CAST(COALESCE(NULLIF(TRIM(field.RULEPOSIT), ''), '0') AS UNSIGNED), field.PARAMNM, field.FIELDNM
                    ) AS pair_index
                FROM rstranfield AS field
                                INNER JOIN rstranrule AS rule
                                    ON field.TRANID = rule.TRANID
                                 AND field.OBJVERS = rule.OBJVERS
                                 AND field.RULEID = rule.RULEID
                WHERE field.OBJVERS = 'A'
                  {tran_filter}
                  AND field.PARAMTYPE = '1'
                                                                        AND rule.GROUPTYPE IN ('N', 'S')
            ),
            target_group_stats AS (
                SELECT
                    tran_id,
                    rule_id,
                    step_id,
                    seg_id,
                    COUNT(*) AS target_count,
                    MIN(pair_index) AS single_target_pair_index
                FROM target_rows
                GROUP BY tran_id, rule_id, step_id, seg_id
            ),
            paired_rows AS (
                SELECT
                    source_rows.tran_id AS tran_id,
                    source_rows.rule_id AS rule_id,
                    source_rows.step_id AS step_id,
                    source_rows.seg_id AS seg_id,
                    source_rows.pair_index AS pair_index,
                    source_rows.source_field AS source_field,
                    source_rows.source_fieldtype AS source_fieldtype,
                    source_rows.source_keyflag AS source_keyflag,
                    source_rows.source_ruleposit AS source_ruleposit,
                    target_rows.target_field AS target_field,
                    target_rows.target_fieldtype AS target_fieldtype,
                    target_rows.target_keyflag AS target_keyflag,
                    target_rows.target_ruleposit AS target_ruleposit
                FROM source_rows
                LEFT JOIN target_group_stats
                  ON target_group_stats.tran_id = source_rows.tran_id
                 AND target_group_stats.rule_id = source_rows.rule_id
                 AND target_group_stats.step_id = source_rows.step_id
                 AND target_group_stats.seg_id = source_rows.seg_id
                LEFT JOIN target_rows
                  ON target_rows.tran_id = source_rows.tran_id
                 AND target_rows.rule_id = source_rows.rule_id
                 AND target_rows.step_id = source_rows.step_id
                 AND target_rows.seg_id = source_rows.seg_id
                 AND target_rows.pair_index = CASE
                     WHEN COALESCE(target_group_stats.target_count, 0) = 1
                         THEN target_group_stats.single_target_pair_index
                     ELSE source_rows.pair_index
                 END

                UNION ALL

                SELECT
                    target_rows.tran_id AS tran_id,
                    target_rows.rule_id AS rule_id,
                    target_rows.step_id AS step_id,
                    target_rows.seg_id AS seg_id,
                    target_rows.pair_index AS pair_index,
                    source_rows.source_field AS source_field,
                    source_rows.source_fieldtype AS source_fieldtype,
                    source_rows.source_keyflag AS source_keyflag,
                    source_rows.source_ruleposit AS source_ruleposit,
                    target_rows.target_field AS target_field,
                    target_rows.target_fieldtype AS target_fieldtype,
                    target_rows.target_keyflag AS target_keyflag,
                    target_rows.target_ruleposit AS target_ruleposit
                FROM target_rows
                LEFT JOIN source_rows
                  ON source_rows.tran_id = target_rows.tran_id
                 AND source_rows.rule_id = target_rows.rule_id
                 AND source_rows.step_id = target_rows.step_id
                 AND source_rows.seg_id = target_rows.seg_id
                 AND source_rows.pair_index = target_rows.pair_index
                WHERE source_rows.rule_id IS NULL
            )
            SELECT
                paired_rows.tran_id AS tran_id,
                paired_rows.rule_id AS rule_id,
                paired_rows.step_id AS step_id,
                paired_rows.seg_id AS seg_id,
                paired_rows.pair_index AS pair_index,
                paired_rows.source_field AS source_field,
                paired_rows.source_fieldtype AS source_fieldtype,
                paired_rows.source_keyflag AS source_keyflag,
                paired_rows.target_field AS target_field,
                paired_rows.target_fieldtype AS target_fieldtype,
                paired_rows.target_keyflag AS target_keyflag,
                COALESCE(paired_rows.target_ruleposit, paired_rows.source_ruleposit, '') AS ruleposit,
                rule.RULETYPE AS rule_type,
                                COALESCE(NULLIF(TRIM(rule.GROUPTYPE), ''), '') AS group_type,
                COALESCE(NULLIF(TRIM(rule.AGGR), ''), '') AS aggr
            FROM paired_rows
            INNER JOIN rstranrule AS rule
              ON paired_rows.tran_id = rule.TRANID
             AND rule.OBJVERS = 'A'
             AND paired_rows.rule_id = rule.RULEID
            WHERE COALESCE(TRIM(paired_rows.source_field), '') <> ''
               OR COALESCE(TRIM(paired_rows.target_field), '') <> ''
            ORDER BY paired_rows.tran_id, paired_rows.rule_id, paired_rows.step_id, paired_rows.seg_id, paired_rows.pair_index
    """


def normalize_tran_mapping_rows(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    normalized_rows: List[Dict[str, object]] = []
    for index, row in enumerate(rows, start=1):
        source_field = str(row.get("source_field") or "").strip()
        target_field = str(row.get("target_field") or "").strip()
        if not source_field and not target_field:
            continue
        raw_display_order = row.get("display_order")
        display_order = int(raw_display_order or 0) if str(raw_display_order or "").strip() else index
        normalized_rows.append(
            {
                "tran_id": str(row.get("tran_id") or "").strip(),
                "rule_id": int(row.get("rule_id") or 0),
                "step_id": int(row.get("step_id") or 0),
                "seg_id": str(row.get("seg_id") or "").strip(),
                "pair_index": int(row.get("pair_index") or 0),
                "display_order": display_order,
                "ruleposit": str(row.get("ruleposit") or "").strip(),
                "source_system": str(row.get("source_system") or "").strip(),
                "source_field": source_field,
                "source_match_field": str(row.get("source_match_field") or "").strip(),
                "source_iobj_name": str(row.get("source_iobj_name") or "").strip(),
                "source_fieldtype": str(row.get("source_fieldtype") or "").strip(),
                "source_key": str(row.get("source_keyflag") or row.get("source_key") or "").strip(),
                "target_field": target_field,
                "target_match_field": str(row.get("target_match_field") or "").strip(),
                "target_iobj_name": str(row.get("target_iobj_name") or "").strip(),
                "target_fieldtype": str(row.get("target_fieldtype") or "").strip(),
                "target_key": str(row.get("target_keyflag") or row.get("target_key") or "").strip(),
                "rule": str(row.get("rule_type") or row.get("rule") or "").strip(),
                "group_type": str(row.get("group_type") or "").strip(),
                "aggr": str(row.get("aggr") or "").strip(),
                "source_field_origin": str(row.get("source_field_origin") or "").strip(),
                "target_field_origin": str(row.get("target_field_origin") or "").strip(),
                "row_kind": str(row.get("row_kind") or "").strip(),
            }
        )
    return normalized_rows


def fetch_materialized_tran_mapping_rows(tran_id: str) -> List[Dict[str, object]]:
    normalized_tran_id = str(tran_id or "").strip()
    if not normalized_tran_id:
        return []

    assert_materialized_read_model_ready()

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
                SELECT tran_id, rule_id, step_id, seg_id, pair_index, display_order, ruleposit,
                             source_system,
                             source_field,
                             source_match_field,
                             source_iobj_name,
                             source_fieldtype,
                             source_field_is_key AS source_keyflag,
                             target_field,
                             target_match_field,
                             target_iobj_name,
                             target_fieldtype,
                             target_field_is_key AS target_keyflag,
                             rule_type, group_type, aggr, source_field_origin, target_field_origin, row_kind,
                             source_object, target_object
                        FROM `{RSTRAN_MAPPING_RULE_FULL_TABLE}` AS base_map
            WHERE tran_id = %s
            ORDER BY display_order, rule_id, step_id, seg_id, pair_index, source_field, target_field
            """,
            (normalized_tran_id,),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    return normalize_tran_mapping_rows(rows)


def fetch_tran_mapping_rows(tran_id: str) -> List[Dict[str, object]]:
    normalized_tran_id = str(tran_id or "").strip()
    if not normalized_tran_id:
        return []
    return fetch_materialized_tran_mapping_rows_batch([normalized_tran_id]).get(normalized_tran_id, [])


def fetch_materialized_tran_mapping_rows_batch(tran_ids: List[str]) -> Dict[str, List[Dict[str, object]]]:
    normalized_tran_ids: List[str] = []
    seen_tran_ids: Set[str] = set()
    for tran_id in tran_ids:
        normalized_tran_id = str(tran_id or "").strip()
        if not normalized_tran_id or normalized_tran_id in seen_tran_ids:
            continue
        seen_tran_ids.add(normalized_tran_id)
        normalized_tran_ids.append(normalized_tran_id)

    if not normalized_tran_ids:
        return {}

    assert_materialized_read_model_ready()

    placeholders = ", ".join(["%s"] * len(normalized_tran_ids))
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"""
            SELECT tran_id, rule_id, step_id, seg_id, pair_index, display_order, ruleposit,
                                     source_type, target_type,
                                                                         source_system, source_field, source_match_field, source_iobj_name, source_fieldtype, source_field_is_key AS source_keyflag,
                                                                         target_field, target_match_field, target_iobj_name, target_fieldtype, target_field_is_key AS target_keyflag,
                     rule_type, group_type, aggr, source_field_origin, target_field_origin, row_kind,
                   source_object, target_object
            FROM `{RSTRAN_MAPPING_RULE_FULL_TABLE}`
            WHERE tran_id IN ({placeholders})
            ORDER BY tran_id, display_order, rule_id, step_id, seg_id, pair_index, source_field, target_field
            """,
            tuple(normalized_tran_ids),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    if not rows:
        return {}

    rows_by_tran: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        tran_id = str(row.get("tran_id") or "").strip()
        normalized_rows = normalize_tran_mapping_rows([
            {
                **row,
                "source_fieldtype": str(row.get("source_fieldtype") or "").strip(),
                "source_keyflag": str(row.get("source_keyflag") or "").strip(),
                "target_fieldtype": str(row.get("target_fieldtype") or "").strip(),
                "target_keyflag": str(row.get("target_keyflag") or "").strip(),
            }
        ])
        if normalized_rows:
            rows_by_tran[tran_id].extend(normalized_rows)

    return rows_by_tran


def normalize_routine_kind(value: object) -> str:
    return str(value or "").strip().upper()


def decode_routine_code_for_display(value: object) -> str:
    return str(value or "").replace("##", "\n")


def detect_logic_language(kind: str, content: str, on_hana: object = "") -> str:
    normalized_kind = normalize_routine_kind(kind)
    normalized_content = str(content or "")
    upper_content = normalized_content.upper()
    on_hana_flag = str(on_hana or "").strip().upper()

    if normalized_kind == "CONSTANT":
        return "plaintext"
    if on_hana_flag in {"X", "1", "TRUE", "Y", "YES"}:
        return "sql"
    if normalized_kind == "FORMULA":
        if re.search(r"\b(SELECT|FROM|WHERE|JOIN|UNION|CASE|COALESCE|GROUP BY|ORDER BY)\b", upper_content):
            return "sql"
        return "plaintext"
    if re.search(r"\b(SELECT|FROM|WHERE|JOIN|UNION|CASE|COALESCE|GROUP BY|ORDER BY)\b", upper_content) and not re.search(r"\b(METHOD|FORM|ENDFORM|DATA:|FIELD-SYMBOLS|LOOP AT|READ TABLE|CLEAR|APPEND)\b", upper_content):
        return "sql"
    return "abap"


def build_logic_entry(
    *,
    tran_id: object,
    rule_id: object = 0,
    step_id: object = 0,
    kind: object = "",
    name: object = "",
    raw_content: object = "",
    on_hana: object = "",
    code_id: object = "",
    source_table: str,
    is_constant: bool = False,
) -> Dict[str, object]:
    normalized_kind = normalize_routine_kind(kind)
    content_raw = str(raw_content or "")
    content_display = content_raw if is_constant else decode_routine_code_for_display(content_raw)
    language = detect_logic_language(normalized_kind or ("CONSTANT" if is_constant else ""), content_display, on_hana)
    title = str(name or "").strip() or normalized_kind or ("CONSTANT" if is_constant else "CODE")
    return {
        "tran_id": str(tran_id or "").strip(),
        "rule_id": int(rule_id or 0),
        "step_id": int(step_id or 0),
        "kind": normalized_kind or ("CONSTANT" if is_constant else ""),
        "title": title,
        "content_raw": content_raw,
        "content_display": content_display,
        "language": language,
        "on_hana": str(on_hana or "").strip(),
        "code_id": str(code_id or "").strip(),
        "source_table": source_table,
        "is_constant": bool(is_constant),
    }


def sort_logic_entries(entries: List[Dict[str, object]]) -> List[Dict[str, object]]:
    return sorted(
        entries,
        key=lambda item: (
            ROUTINE_KIND_SORT_ORDER.get(str(item.get("kind") or "").strip().upper(), 999),
            str(item.get("title") or "").strip().upper(),
            str(item.get("code_id") or "").strip().upper(),
        ),
    )


def get_rule_logic_lookup_key(tran_id: str, rule_id: int) -> Tuple[str, int] | None:
    normalized_tran_id = str(tran_id or "").strip()
    normalized_rule_id = int(rule_id or 0)
    if not normalized_tran_id or normalized_rule_id <= 0:
        return None
    return (normalized_tran_id, normalized_rule_id)


def fetch_tran_logic_details(tran_ids: List[str]) -> Dict[str, object]:
    normalized_tran_ids = sorted({str(item or "").strip() for item in (tran_ids or []) if str(item or "").strip()})
    if not normalized_tran_ids:
        return {"step_programs": {}, "rule_entries": {}}

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    step_programs_map: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    rule_entries_map: Dict[Tuple[str, int], List[Dict[str, object]]] = defaultdict(list)
    try:
        placeholders = ",".join(["%s"] * len(normalized_tran_ids))
        upper_tran_ids = tuple(str(item).upper() for item in normalized_tran_ids)

        cur.execute(
            f"""
            SELECT TRANID, RULEID, STEPID, ON_HANA, CODEID, KIND, NAME, CODE
            FROM rstransteprout
            WHERE UPPER(TRIM(TRANID)) IN ({placeholders})
            ORDER BY TRANID, STEPID, RULEID, KIND, CODEID
            """,
            upper_tran_ids,
        )
        for row in cur.fetchall():
            tran_id = str(row.get("TRANID") or "").strip()
            rule_id = int(row.get("RULEID") or 0)
            step_id = int(row.get("STEPID") or 0)
            kind = normalize_routine_kind(row.get("KIND"))
            entry = build_logic_entry(
                tran_id=tran_id,
                rule_id=rule_id,
                step_id=step_id,
                kind=kind,
                name=row.get("NAME"),
                raw_content=row.get("CODE"),
                on_hana=row.get("ON_HANA"),
                code_id=row.get("CODEID"),
                source_table="rstransteprout",
            )
            if kind in STEP_LEVEL_ROUTINE_KINDS:
                step_programs_map[tran_id].append(entry)
            else:
                lookup_key = get_rule_logic_lookup_key(tran_id, rule_id)
                if lookup_key:
                    rule_entries_map[lookup_key].append(entry)

        cur.execute(
            f"""
            SELECT TRANID, RULEID, STEPID, VALUE
            FROM rstranstepcnst
            WHERE UPPER(TRIM(TRANID)) IN ({placeholders})
            ORDER BY TRANID, STEPID, RULEID
            """,
            upper_tran_ids,
        )
        for row in cur.fetchall():
            tran_id = str(row.get("TRANID") or "").strip()
            rule_id = int(row.get("RULEID") or 0)
            step_id = int(row.get("STEPID") or 0)
            entry = build_logic_entry(
                tran_id=tran_id,
                rule_id=rule_id,
                step_id=step_id,
                kind="CONSTANT",
                name="Constant",
                raw_content=row.get("VALUE"),
                source_table="rstranstepcnst",
                is_constant=True,
            )
            lookup_key = get_rule_logic_lookup_key(tran_id, rule_id)
            if lookup_key:
                rule_entries_map[lookup_key].append(entry)
    finally:
        cur.close()
        conn.close()

    return {
        "step_programs": {tran_id: sort_logic_entries(entries) for tran_id, entries in step_programs_map.items()},
        "rule_entries": {key: sort_logic_entries(entries) for key, entries in rule_entries_map.items()},
    }


def fetch_tran_logic_summary(tran_ids: List[str]) -> Dict[str, object]:
    normalized_tran_ids = sorted({str(item or "").strip() for item in (tran_ids or []) if str(item or "").strip()})
    if not normalized_tran_ids:
        return {"step_entry_count_by_tran": {}, "rule_entry_keys": set()}

    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    step_entry_count_by_tran: Dict[str, int] = defaultdict(int)
    rule_entry_keys: Set[Tuple[str, int]] = set()
    try:
        placeholders = ",".join(["%s"] * len(normalized_tran_ids))
        upper_tran_ids = tuple(str(item).upper() for item in normalized_tran_ids)

        cur.execute(
            f"""
            SELECT TRANID, RULEID, STEPID, KIND
            FROM rstransteprout
            WHERE UPPER(TRIM(TRANID)) IN ({placeholders})
            ORDER BY TRANID, STEPID, RULEID, KIND
            """,
            upper_tran_ids,
        )
        for row in cur.fetchall():
            tran_id = str(row.get("TRANID") or "").strip()
            rule_id = int(row.get("RULEID") or 0)
            step_id = int(row.get("STEPID") or 0)
            kind = normalize_routine_kind(row.get("KIND"))
            if kind in STEP_LEVEL_ROUTINE_KINDS:
                step_entry_count_by_tran[tran_id] += 1
            else:
                lookup_key = get_rule_logic_lookup_key(tran_id, rule_id)
                if lookup_key:
                    rule_entry_keys.add(lookup_key)

        cur.execute(
            f"""
            SELECT TRANID, RULEID, STEPID
            FROM rstranstepcnst
            WHERE UPPER(TRIM(TRANID)) IN ({placeholders})
            ORDER BY TRANID, STEPID, RULEID
            """,
            upper_tran_ids,
        )
        for row in cur.fetchall():
            tran_id = str(row.get("TRANID") or "").strip()
            rule_id = int(row.get("RULEID") or 0)
            step_id = int(row.get("STEPID") or 0)
            lookup_key = get_rule_logic_lookup_key(tran_id, rule_id)
            if lookup_key:
                rule_entry_keys.add(lookup_key)
    finally:
        cur.close()
        conn.close()

    return {
        "step_entry_count_by_tran": dict(step_entry_count_by_tran),
        "rule_entry_keys": rule_entry_keys,
    }


def normalize_segment_tran_ids(tran_ids: List[str] | None) -> List[str]:
    normalized_tran_ids: List[str] = []
    seen_tran_ids: Set[str] = set()
    for item in tran_ids or []:
        normalized_tran_id = str(item or "").strip()
        if not normalized_tran_id or normalized_tran_id in seen_tran_ids:
            continue
        seen_tran_ids.add(normalized_tran_id)
        normalized_tran_ids.append(normalized_tran_id)
    return normalized_tran_ids


def filter_segment_mapping_rows_by_subtype(
    rows: List[Dict[str, object]],
    segment_source_type: str,
    segment_source_subtype: str,
    source_inventory: Dict[str, Dict[str, object]],
    segment_target_type: str,
    segment_target_subtype: str,
    target_inventory: Dict[str, Dict[str, object]],
) -> List[Dict[str, object]]:
    normalized_source_type = normalize_type_code(segment_source_type)
    normalized_target_type = normalize_type_code(segment_target_type)
    normalized_source_subtype = normalize_iobj_subtype(segment_source_subtype)
    normalized_target_subtype = normalize_iobj_subtype(segment_target_subtype)
    if normalized_source_type != "IOBJ" and normalized_target_type != "IOBJ":
        return rows

    filtered_rows: List[Dict[str, object]] = []
    for row in rows:
        row_kind = str(row.get("row_kind") or "").strip()
        seg_id = str(row.get("seg_id") or "").strip()
        if normalized_source_type == "IOBJ" and normalized_source_subtype and row_kind == "source_only":
            source_field = normalize_field_name(row.get("source_field"))
            if source_field and not inventory_contains_field(source_inventory, seg_id, source_field):
                continue
        if normalized_target_type == "IOBJ" and normalized_target_subtype and row_kind == "target_only":
            target_field = normalize_field_name(row.get("target_field"))
            if target_field and not inventory_contains_field(target_inventory, seg_id, target_field):
                continue
        filtered_rows.append(row)
    return filtered_rows


def build_path_mapping_payload(
    segments: List[PathSegmentRequest],
    include_logic: bool = False,
    include_text: bool = False,
) -> Dict[str, object]:
    assert_materialized_read_model_ready()
    total_start = time.time()
    
    prepared_segments: List[Dict[str, object]] = []
    adso_objects: Set[str] = set()
    odso_objects: Set[str] = set()
    trcs_objects: Set[str] = set()
    iobj_objects: Set[Tuple[str, str]] = set()
    rsds_pairs: Set[Tuple[str, str]] = set()
    all_tran_ids: Set[str] = set()
    tran_metadata_cache: Dict[str, Dict[str, str]] = {}

    step1_start = time.time()
    
    # Pre-fetch metadata for all TRANIDs in batch (SINGLE QUERY!)
    all_requested_tran_ids: Set[str] = set()
    for segment in segments:
        tran_ids = normalize_segment_tran_ids(get_segment_tran_ids(segment))
        all_requested_tran_ids.update(tran_ids)
    
    step1a_start = time.time()
    tran_metadata_cache = fetch_path_selection_metadata_batch(list(all_requested_tran_ids))
    step1a_end = time.time()
    
    step1b_start = time.time()
    for segment in segments:
        tran_ids = normalize_segment_tran_ids(get_segment_tran_ids(segment))
        segment_source = get_segment_source_name(segment)
        segment_target = get_segment_target_name(segment)
        segment_source_type = get_segment_source_type(segment)
        segment_target_type = get_segment_target_type(segment)
        segment_source_subtype = get_segment_source_subtype(segment)
        segment_target_subtype = get_segment_target_subtype(segment)
        source_system = get_segment_source_system(segment)
        target_system = get_segment_target_system(segment)

        if tran_ids:
            for tran_id in tran_ids:
                normalized_tran_id = str(tran_id or "").strip()
                if not normalized_tran_id:
                    continue
                if normalized_tran_id not in tran_metadata_cache:
                    tran_metadata_cache[normalized_tran_id] = fetch_path_selection_metadata_by_tran_id(normalized_tran_id)
                tran_meta = tran_metadata_cache.get(normalized_tran_id) or {}
                if not tran_meta:
                    continue
                if not segment_source:
                    segment_source = str(tran_meta.get("source_object") or "").strip()
                if not segment_target:
                    segment_target = str(tran_meta.get("target_object") or "").strip()
                if not segment_source_type:
                    segment_source_type = str(tran_meta.get("source_type") or "").strip()
                if not segment_target_type:
                    segment_target_type = str(tran_meta.get("target_type") or "").strip()
                if not segment_source_subtype:
                    segment_source_subtype = str(tran_meta.get("source_subtype") or "").strip()
                if not segment_target_subtype:
                    segment_target_subtype = str(tran_meta.get("target_subtype") or "").strip()
                if not source_system:
                    source_system = str(tran_meta.get("source_system") or "").strip()
                if not target_system:
                    target_system = str(tran_meta.get("target_system") or "").strip()
                if segment_source and segment_target and segment_source_type and segment_target_type and (source_system or normalize_type_code(segment_source_type) != "RSDS"):
                    break

        normalized_source_type = normalize_type_code(segment_source_type)
        normalized_target_type = normalize_type_code(segment_target_type)
        normalized_source_subtype = normalize_iobj_subtype(segment_source_subtype)
        normalized_target_subtype = normalize_iobj_subtype(segment_target_subtype)

        if normalized_source_type in {"RSDS", "DATASOURCE", "SOURCE"}:
            segment_source, source_system = parse_rsds_identity(segment_source, source_system)
        if normalized_target_type in {"RSDS", "DATASOURCE", "SOURCE"}:
            segment_target, target_system = parse_rsds_identity(segment_target, target_system)

        normalized_source = normalize_bw_object_lookup(segment_source)
        normalized_target = normalize_bw_object_lookup(segment_target)

        if normalized_source_type == "ADSO" and normalized_source:
            adso_objects.add(normalized_source)
        elif normalized_source_type == "ODSO" and normalized_source:
            odso_objects.add(normalized_source)
        elif normalized_source_type == "TRCS" and normalized_source:
            trcs_objects.add(normalized_source)
        elif normalized_source_type == "IOBJ" and normalized_source:
            iobj_objects.add((normalized_source, normalized_source_subtype))
        elif normalized_source_type in {"RSDS", "DATASOURCE", "SOURCE"} and normalized_source and source_system:
            rsds_pairs.add((normalized_source, source_system))

        if normalized_target_type == "ADSO" and normalized_target:
            adso_objects.add(normalized_target)
        elif normalized_target_type == "ODSO" and normalized_target:
            odso_objects.add(normalized_target)
        elif normalized_target_type == "TRCS" and normalized_target:
            trcs_objects.add(normalized_target)
        elif normalized_target_type == "IOBJ" and normalized_target:
            iobj_objects.add((normalized_target, normalized_target_subtype))
        elif normalized_target_type in {"RSDS", "DATASOURCE", "SOURCE"} and normalized_target and target_system:
            rsds_pairs.add((normalized_target, target_system))

        prepared_segments.append(
            {
                "SOURCE": segment_source,
                "TARGETNAME": segment_target,
                "SOURCETYPE": segment_source_type,
                "TARGETTYPE": segment_target_type,
                "SOURCESUBTYPE": normalized_source_subtype,
                "TARGETSUBTYPE": normalized_target_subtype,
                "SOURCESYS": source_system,
                "TARGETSYSTEM": target_system,
                "TRANIDS": tran_ids,
            }
        )
        all_tran_ids.update(tran_ids)
    step1b_end = time.time()
    step1_end = step1b_end
    
    step2_start = time.time()
    rsds_name_lookup = fetch_rsds_object_text_lookup(sorted(rsds_pairs)) if rsds_pairs else {}
    step2_end = time.time()
    
    step3_start = time.time()
    adso_name_lookup = fetch_adso_object_text_lookup(sorted(adso_objects)) if adso_objects else {}
    step3_end = time.time()
    
    step4_start = time.time()
    trcs_name_lookup = fetch_trcs_object_text_lookup(sorted(trcs_objects)) if trcs_objects else {}
    step4_end = time.time()
    
    step5_start = time.time()
    iobj_name_lookup = fetch_iobj_object_text_lookup(sorted(iobj_objects)) if iobj_objects else {}
    step5_end = time.time()
    
    step6_start = time.time()
    object_display_name_map: Dict[str, str] = {}
    activate_data_by_object = fetch_rsoadso_activate_data_by_object(sorted(adso_objects)) if adso_objects else {}
    step6_end = time.time()
    
    step7_start = time.time()
    rsds_inventory = fetch_rsds_field_inventory_for_pairs(sorted(rsds_pairs)) if rsds_pairs else {}
    step7_end = time.time()
    
    step8_start = time.time()
    rsiobj_names = fetch_rsiobj_name_set() if (adso_objects or odso_objects) else set()
    step8_end = time.time()
    
    step9_start = time.time()
    dd03l_inventory: Dict[str, Dict[str, Dict[str, object]]] = {}
    if odso_objects:
        dd03l_inventory.update(fetch_materialized_dd03l_field_inventory("ODSO", sorted(odso_objects)))
    if adso_objects:
        dd03l_inventory.update(fetch_materialized_dd03l_field_inventory("ADSO", sorted(adso_objects)))
    step9_end = time.time()
    
    step10_start = time.time()
    trcs_inventory = fetch_materialized_trcs_field_inventory(sorted(trcs_objects)) if trcs_objects else {}
    step10_end = time.time()
    
    step11_start = time.time()
    iobj_inventory = fetch_materialized_iobj_field_inventory(sorted(iobj_objects)) if iobj_objects else {}
    step11_end = time.time()
    
    step12_start = time.time()
    tran_mapping_rows_by_tran = fetch_materialized_tran_mapping_rows_batch(sorted(all_tran_ids)) if all_tran_ids else {}
    step12_end = time.time()
    
    step13_start = time.time()
    logic_details = (
        fetch_tran_logic_details(sorted(all_tran_ids))
        if include_logic and all_tran_ids
        else {"step_programs": {}, "rule_entries": {}}
    )
    step13_end = time.time()
    
    step14_start = time.time()
    logic_summary = fetch_tran_logic_summary(sorted(all_tran_ids)) if all_tran_ids else {"step_entry_count_by_tran": {}, "rule_entry_keys": set()}
    step14_end = time.time()
    step_programs_by_tran = logic_details.get("step_programs") or {}
    rule_entries_by_key = logic_details.get("rule_entries") or {}
    step_entry_count_by_tran = logic_summary.get("step_entry_count_by_tran") or {}
    rule_entry_keys = logic_summary.get("rule_entry_keys") or set()
    rsiobj_type_lookup = fetch_rsiobj_type_lookup() if include_text else {}
    step15_start = time.time()
    field_text_lookup_cache: Dict[Tuple[str, str, str], Dict[Tuple[str, str], str]] = {}
    tran_name_lookup = fetch_rstrant_text_lookup(sorted(all_tran_ids)) if all_tran_ids else {}

    items: List[Dict[str, object]] = []
    for index, segment in enumerate(prepared_segments, start=1):
        tran_ids = segment.get("TRANIDS") or []
        mapping_rows: List[Dict[str, object]] = []
        for tran_id in tran_ids:
            normalized_tran_id = str(tran_id or "").strip()
            if normalized_tran_id:
                mapping_rows.extend(tran_mapping_rows_by_tran.get(normalized_tran_id, []))

        segment_source = str(segment.get("SOURCE") or "").strip()
        segment_target = str(segment.get("TARGETNAME") or "").strip()
        segment_source_type = str(segment.get("SOURCETYPE") or "").strip()
        segment_target_type = str(segment.get("TARGETTYPE") or "").strip()
        segment_source_subtype = normalize_iobj_subtype(segment.get("SOURCESUBTYPE") or "")
        segment_target_subtype = normalize_iobj_subtype(segment.get("TARGETSUBTYPE") or "")
        source_system = str(segment.get("SOURCESYS") or "").strip()
        target_system = str(segment.get("TARGETSYSTEM") or "").strip()
        source_display_name = resolve_path_object_display_name(
            segment_source,
            source_system,
            segment_source_type,
            rsds_name_lookup=rsds_name_lookup,
            adso_name_lookup=adso_name_lookup,
            trcs_name_lookup=trcs_name_lookup,
            iobj_name_lookup=iobj_name_lookup,
        )
        target_display_name = resolve_path_object_display_name(
            segment_target,
            target_system,
            segment_target_type,
            rsds_name_lookup=rsds_name_lookup,
            adso_name_lookup=adso_name_lookup,
            trcs_name_lookup=trcs_name_lookup,
            iobj_name_lookup=iobj_name_lookup,
        )
        normalized_segment_source = normalize_bw_object_lookup(segment_source)
        normalized_segment_target = normalize_bw_object_lookup(segment_target)
        if normalized_segment_source and source_display_name and normalized_segment_source not in object_display_name_map:
            object_display_name_map[normalized_segment_source] = source_display_name
        if normalized_segment_target and target_display_name and normalized_segment_target not in object_display_name_map:
            object_display_name_map[normalized_segment_target] = target_display_name

        if include_text:
            source_field_names = {
                normalize_field_name(row.get("source_field"))
                for row in mapping_rows
                if normalize_field_name(row.get("source_field"))
            }
            target_field_names = {
                normalize_field_name(row.get("target_field"))
                for row in mapping_rows
                if normalize_field_name(row.get("target_field"))
            }
            seg_ids = {
                str(row.get("seg_id") or "").strip()
                for row in mapping_rows
                if str(row.get("seg_id") or "").strip()
            }

            source_lookup_key = (segment_source, source_system, segment_source_type)
            target_lookup_key = (segment_target, target_system, segment_target_type)
            if source_lookup_key not in field_text_lookup_cache:
                field_text_lookup_cache[source_lookup_key] = build_field_text_lookup_for_object(
                    segment_source,
                    source_system,
                    segment_source_type,
                    field_names=source_field_names,
                    seg_ids=seg_ids,
                ) or {}
            if target_lookup_key not in field_text_lookup_cache:
                field_text_lookup_cache[target_lookup_key] = build_field_text_lookup_for_object(
                    segment_target,
                    target_system,
                    segment_target_type,
                    field_names=target_field_names,
                    seg_ids=seg_ids,
                ) or {}

            mapping_rows = enrich_mapping_rows_with_field_text(
                mapping_rows,
                segment_source,
                source_system,
                segment_source_type,
                segment_target,
                target_system,
                segment_target_type,
                source_lookup=field_text_lookup_cache.get(source_lookup_key, {}),
                target_lookup=field_text_lookup_cache.get(target_lookup_key, {}),
                rsiobj_type_lookup=rsiobj_type_lookup,
            )

        if include_logic:
            for row in mapping_rows:
                tran_id = str(row.get("tran_id") or "").strip()
                rule_id = int(row.get("rule_id") or 0)
                step_id = int(row.get("step_id") or 0)
                logic_entries = rule_entries_by_key.get((tran_id, rule_id), [])
                row["logic_entries"] = logic_entries
                row["has_logic_entry"] = bool(logic_entries)
        else:
            for row in mapping_rows:
                tran_id = str(row.get("tran_id") or "").strip()
                rule_id = int(row.get("rule_id") or 0)
                step_id = int(row.get("step_id") or 0)
                has_logic_entry = (tran_id, rule_id) in rule_entry_keys
                row["logic_entries"] = []
                row["has_logic_entry"] = has_logic_entry

        segment_step_logic = []
        if include_logic:
            for tran_id in tran_ids:
                entries = step_programs_by_tran.get(str(tran_id or "").strip(), [])
                if not entries:
                    continue
                segment_step_logic.append(
                    build_path_step_logic_group_payload(
                        {
                            "tran_id": str(tran_id or "").strip(),
                            "entry_count": len(entries),
                            "entries": entries,
                        }
                    )
                )
        else:
            for tran_id in tran_ids:
                entry_count = int(step_entry_count_by_tran.get(str(tran_id or "").strip()) or 0)
                if entry_count <= 0:
                    continue
                segment_step_logic.append(
                    build_path_step_logic_group_payload(
                        {
                            "tran_id": str(tran_id or "").strip(),
                            "entry_count": entry_count,
                            "entries": [],
                        }
                    )
                )

        source_inventory, _source_supported, _source_inventory_origin = get_supported_inventory_for_object(
            segment_source,
            source_system,
            segment_source_type,
            segment_source_subtype,
            rsds_inventory,
            dd03l_inventory,
            trcs_inventory,
            iobj_inventory,
        )
        target_inventory, _target_supported, _target_inventory_origin = get_supported_inventory_for_object(
            segment_target,
            target_system,
            segment_target_type,
            segment_target_subtype,
            rsds_inventory,
            dd03l_inventory,
            trcs_inventory,
            iobj_inventory,
        )
        mapping_rows = filter_segment_mapping_rows_by_subtype(
            mapping_rows,
            segment_source_type,
            segment_source_subtype,
            source_inventory,
            segment_target_type,
            segment_target_subtype,
            target_inventory,
        )
        source_diagnostics = build_object_field_diagnostics(
            mapping_rows,
            "source_field",
            segment_source,
            source_system,
            segment_source_type,
            segment_source_subtype,
            object_display_name_map,
            rsds_inventory,
            dd03l_inventory,
            trcs_inventory,
            iobj_inventory,
            activate_data_by_object,
            rsiobj_names,
        )
        target_diagnostics = build_object_field_diagnostics(
            mapping_rows,
            "target_field",
            segment_target,
            target_system,
            segment_target_type,
            segment_target_subtype,
            object_display_name_map,
            rsds_inventory,
            dd03l_inventory,
            trcs_inventory,
            iobj_inventory,
            activate_data_by_object,
            rsiobj_names,
        )
        mapping_rows = enrich_mapping_rows_with_inventory_field_metadata(
            mapping_rows,
            source_inventory,
            segment_source_type,
            target_inventory,
            segment_target_type,
            rsiobj_names,
            segment_source,
            segment_target,
        )

        items.append(
            {
                "INDEX": index,
                "SOURCE": segment_source,
                "TARGETNAME": segment_target,
                "SOURCETYPE": segment_source_type,
                "TARGETTYPE": segment_target_type,
                "SOURCESUBTYPE": segment_source_subtype,
                "TARGETSUBTYPE": segment_target_subtype,
                "SOURCESYS": source_system,
                "TARGETSYSTEM": target_system,
                "SOURCE_DISPLAY_NAME": source_display_name,
                "TARGET_DISPLAY_NAME": target_display_name,
                "TRANIDS": tran_ids,
                "TR_NAMES": [str(tran_name_lookup.get(str(tran_id or "").strip()) or "").strip() for tran_id in tran_ids if str(tran_id or "").strip()],
                "ROWS": [build_path_mapping_row_payload(row) for row in mapping_rows],
                "STEP_LOGIC": segment_step_logic,
                "STEP_LOGIC_ENTRY_COUNT": sum(int(item.get("ENTRY_COUNT") or 0) for item in segment_step_logic),
                "DIAGNOSTICS": {
                    "SOURCE": build_path_diagnostics_payload(source_diagnostics),
                    "TARGET": build_path_diagnostics_payload(target_diagnostics),
                },
            }
        )
    step15_end = time.time()

    result = {
        "SEGMENT_COUNT": len(items),
        "SEGMENTS": items,
        "INCLUDE_LOGIC": bool(include_logic),
        "INCLUDE_TEXT": bool(include_text),
    }
    
    total_end = time.time()
    
    return result


def build_path_text_payload(segments: List[PathSegmentRequest]) -> Dict[str, object]:
    if not segments:
        return {
            "SEGMENT_COUNT": 0,
            "SEGMENTS": [],
            "INCLUDE_TEXT": True,
        }

    mapping_payload = build_path_mapping_payload(
        segments,
        include_logic=False,
        include_text=True,
    )

    text_segments: List[Dict[str, object]] = []
    for index, segment in enumerate(mapping_payload.get("SEGMENTS") or [], start=1):
        rows = []
        for row in segment.get("ROWS") or []:
            rows.append(
                {
                    "TRANID": str(row.get("TRANID") or "").strip(),
                    "RULEID": int(row.get("RULEID") or 0),
                    "STEPID": int(row.get("STEPID") or 0),
                    "SEGID": str(row.get("SEGID") or "").strip(),
                    "PAIR_INDEX": int(row.get("PAIR_INDEX") or 0),
                    "SOURCE_FIELD": str(row.get("SOURCE_FIELD") or "").strip(),
                    "TARGET_FIELD": str(row.get("TARGET_FIELD") or "").strip(),
                    "SOURCE_TEXT": str(row.get("SOURCE_TEXT") or "").strip(),
                    "TARGET_TEXT": str(row.get("TARGET_TEXT") or "").strip(),
                }
            )

        text_segments.append(
            {
                "INDEX": int(segment.get("INDEX") or index),
                "SOURCE": str(segment.get("SOURCE") or "").strip(),
                "TARGETNAME": str(segment.get("TARGETNAME") or "").strip(),
                "SOURCETYPE": str(segment.get("SOURCETYPE") or "").strip(),
                "TARGETTYPE": str(segment.get("TARGETTYPE") or "").strip(),
                "SOURCESUBTYPE": str(segment.get("SOURCESUBTYPE") or "").strip(),
                "TARGETSUBTYPE": str(segment.get("TARGETSUBTYPE") or "").strip(),
                "SOURCESYS": str(segment.get("SOURCESYS") or "").strip(),
                "TARGETSYSTEM": str(segment.get("TARGETSYSTEM") or "").strip(),
                "TRANIDS": [str(item or "").strip() for item in (segment.get("TRANIDS") or []) if str(item or "").strip()],
                "TR_NAMES": [str(item or "").strip() for item in (segment.get("TR_NAMES") or []) if str(item or "").strip()],
                "ROWS": rows,
            }
        )

    return {
        "SEGMENT_COUNT": len(text_segments),
        "SEGMENTS": text_segments,
        "INCLUDE_TEXT": True,
    }


def build_path_logic_payload(segments: List[PathSegmentRequest]) -> Dict[str, object]:
    if not segments:
        return {
            "SEGMENT_COUNT": 0,
            "SEGMENTS": [],
            "INCLUDE_LOGIC": True,
        }

    assert_materialized_read_model_ready()

    prepared_segments: List[Dict[str, object]] = []
    all_tran_ids: Set[str] = set()
    for segment in segments:
        tran_ids = normalize_segment_tran_ids(get_segment_tran_ids(segment))
        segment_source = get_segment_source_name(segment)
        segment_target = get_segment_target_name(segment)
        segment_source_type = get_segment_source_type(segment)
        segment_target_type = get_segment_target_type(segment)
        segment_source_subtype = get_segment_source_subtype(segment)
        segment_target_subtype = get_segment_target_subtype(segment)
        source_system = get_segment_source_system(segment)
        target_system = get_segment_target_system(segment)
        normalized_source_type = normalize_type_code(segment_source_type)
        normalized_target_type = normalize_type_code(segment_target_type)

        if normalized_source_type in {"RSDS", "DATASOURCE", "SOURCE"}:
            segment_source, source_system = parse_rsds_identity(segment_source, source_system)
        if normalized_target_type in {"RSDS", "DATASOURCE", "SOURCE"}:
            segment_target, target_system = parse_rsds_identity(segment_target, target_system)

        prepared_segments.append(
            {
                "SOURCE": segment_source,
                "TARGETNAME": segment_target,
                "SOURCETYPE": segment_source_type,
                "TARGETTYPE": segment_target_type,
                "SOURCESUBTYPE": segment_source_subtype,
                "TARGETSUBTYPE": segment_target_subtype,
                "SOURCESYS": source_system,
                "TARGETSYSTEM": target_system,
                "TRANIDS": tran_ids,
            }
        )
        all_tran_ids.update(tran_ids)

    tran_mapping_rows_by_tran = fetch_materialized_tran_mapping_rows_batch(sorted(all_tran_ids)) if all_tran_ids else {}
    logic_details = fetch_tran_logic_details(sorted(all_tran_ids)) if all_tran_ids else {"step_programs": {}, "rule_entries": {}}
    step_programs_by_tran = logic_details.get("step_programs") or {}
    rule_entries_by_key = logic_details.get("rule_entries") or {}
    tran_name_lookup = fetch_rstrant_text_lookup(sorted(all_tran_ids)) if all_tran_ids else {}

    logic_segments: List[Dict[str, object]] = []
    for index, segment in enumerate(prepared_segments, start=1):
        tran_ids = segment.get("TRANIDS") or []
        mapping_rows: List[Dict[str, object]] = []
        for tran_id in tran_ids:
            normalized_tran_id = str(tran_id or "").strip()
            if normalized_tran_id:
                mapping_rows.extend(tran_mapping_rows_by_tran.get(normalized_tran_id, []))

        rows = []
        for row in mapping_rows:
            tran_id = str(row.get("tran_id") or "").strip()
            rule_id = int(row.get("rule_id") or 0)
            step_id = int(row.get("step_id") or 0)
            logic_entries = rule_entries_by_key.get((tran_id, rule_id), [])
            rows.append(
                {
                    "TRANID": tran_id,
                    "RULEID": rule_id,
                    "STEPID": step_id,
                    "SEGID": str(row.get("seg_id") or "").strip(),
                    "PAIR_INDEX": int(row.get("pair_index") or 0),
                    "SOURCE_FIELD": str(row.get("source_field") or "").strip(),
                    "TARGET_FIELD": str(row.get("target_field") or "").strip(),
                    "LOGIC_ENTRIES": [build_path_logic_entry_payload(entry) for entry in logic_entries],
                    "HAS_LOGIC_ENTRY": bool(logic_entries),
                }
            )

        step_logic = []
        for tran_id in tran_ids:
            entries = step_programs_by_tran.get(str(tran_id or "").strip(), [])
            if not entries:
                continue
            step_logic.append(
                build_path_step_logic_group_payload(
                    {
                        "tran_id": str(tran_id or "").strip(),
                        "entry_count": len(entries),
                        "entries": list(entries),
                    }
                )
            )

        logic_segments.append(
            {
                "INDEX": int(segment.get("INDEX") or index),
                "SOURCE": str(segment.get("SOURCE") or "").strip(),
                "TARGETNAME": str(segment.get("TARGETNAME") or "").strip(),
                "SOURCETYPE": str(segment.get("SOURCETYPE") or "").strip(),
                "TARGETTYPE": str(segment.get("TARGETTYPE") or "").strip(),
                "SOURCESUBTYPE": str(segment.get("SOURCESUBTYPE") or "").strip(),
                "TARGETSUBTYPE": str(segment.get("TARGETSUBTYPE") or "").strip(),
                "SOURCESYS": str(segment.get("SOURCESYS") or "").strip(),
                "TARGETSYSTEM": str(segment.get("TARGETSYSTEM") or "").strip(),
                "TRANIDS": [str(item or "").strip() for item in (segment.get("TRANIDS") or []) if str(item or "").strip()],
                "TR_NAMES": [str(tran_name_lookup.get(str(item or "").strip()) or "").strip() for item in (segment.get("TRANIDS") or []) if str(item or "").strip()],
                "ROWS": rows,
                "STEP_LOGIC": step_logic,
                "STEP_LOGIC_ENTRY_COUNT": sum(int(item.get("ENTRY_COUNT") or 0) for item in step_logic),
            }
        )

    return {
        "SEGMENT_COUNT": len(logic_segments),
        "SEGMENTS": logic_segments,
        "INCLUDE_LOGIC": True,
    }


def _build_graph_engine_by_source(start_name: str, max_nodes: int = 2000, max_edges: int = 5000) -> Dict[str, object]:
    """Wave expansion by SOURCENAME -> TARGETNAME with row-level deduplication."""
    seed = start_name.strip()
    if not seed:
        return {"nodes": [], "edges": [], "color_map": {}}

    conn = get_conn()
    cur = conn.cursor()

    # Row-level deduplication for stable traversal.
    visited_row_indices: Set[int] = set()
    current_keys: List[str] = [seed]
    node_types: Dict[str, str] = {seed: "UNKNOWN"}
    node_levels: Dict[str, int] = {seed: 0}
    node_object_names: Dict[str, str] = {}
    edge_seen: Set[Tuple[str, str, str, str]] = set()
    edge_by_key: Dict[Tuple[str, str, str, str], Dict[str, object]] = {}
    edges: List[Dict[str, object]] = []

    try:
        cur.execute(
            """
            SELECT SOURCENAME, TARGETNAME, SOURCETYPE, TARGETTYPE, TRANID, STARTROUTINE, ENDROUTINE, EXPERT
            FROM rstran
            WHERE SOURCENAME IS NOT NULL AND TARGETNAME IS NOT NULL
            ORDER BY TRANID
            """
        )
        all_rows = cur.fetchall()

        source_to_indices: Dict[str, List[int]] = {}
        normalized_rows: List[Tuple[str, str, str, str, object, object, object, object]] = []
        for idx, row in enumerate(all_rows):
            source_name = str(row[0] or "").strip()
            target_name = str(row[1] or "").strip()
            source_type = normalize_type_code(row[2])
            target_type = normalize_type_code(row[3])
            normalized_rows.append((source_name, target_name, source_type, target_type, row[4], row[5], row[6], row[7]))
            if not source_name:
                continue
            source_to_indices.setdefault(source_name, []).append(idx)

        depth = 0
        while current_keys:
            next_keys: List[str] = []
            for key in current_keys:
                for idx in source_to_indices.get(key, []):
                    if idx in visited_row_indices:
                        continue
                    visited_row_indices.add(idx)

                    source_name, target_name, source_type, target_type, tran_id, start_routine, end_routine, expert = normalized_rows[idx]
                    if not source_name or not target_name:
                        continue

                    edge_key = (source_name, target_name, source_type, target_type)
                    if edge_key not in edge_seen:
                        edge_seen.add(edge_key)
                        append_or_update_edge(
                            edges,
                            edge_by_key,
                            source_name,
                            target_name,
                            source_type,
                            target_type,
                            "",
                            "",
                            tran_id,
                            start_routine,
                            end_routine,
                            expert,
                        )
                    else:
                        append_or_update_edge(
                            edges,
                            edge_by_key,
                            source_name,
                            target_name,
                            source_type,
                            target_type,
                            "",
                            "",
                            tran_id,
                            start_routine,
                            end_routine,
                            expert,
                        )

                    if source_name not in node_types or node_types[source_name] == "UNKNOWN":
                        node_types[source_name] = source_type
                    if target_name not in node_types or node_types[target_name] == "UNKNOWN":
                        node_types[target_name] = target_type

                    if source_name not in node_levels:
                        node_levels[source_name] = depth
                    if target_name not in node_levels:
                        node_levels[target_name] = depth + 1

                    next_keys.append(target_name)

                    if len(node_types) >= max_nodes or len(edges) >= max_edges:
                        break
                if len(node_types) >= max_nodes or len(edges) >= max_edges:
                    break

            current_keys = next_keys
            depth += 1

            if len(node_types) >= max_nodes or len(edges) >= max_edges:
                break

        # Enrich all discovered technical names with NAME_EN from bw_object_name.
        all_node_names = list(node_types.keys())
        batch_size = 500
        for start in range(0, len(all_node_names), batch_size):
            batch = [str(v).strip().upper() for v in all_node_names[start : start + batch_size] if str(v).strip()]
            if not batch:
                continue
            placeholders = ",".join(["%s"] * len(batch))
            cur.execute(
                f"""
                SELECT BW_OBJECT_NORM, MAX(NAME_EN) AS NAME_EN
                FROM bw_object_name
                WHERE BW_OBJECT_NORM IN ({placeholders})
                GROUP BY BW_OBJECT_NORM
                """,
                tuple(batch),
            )
            for bw_object, object_name in cur.fetchall():
                key = normalize_bw_object_lookup(bw_object)
                value = str(object_name or "").strip()
                if key and value:
                    node_object_names[key] = value
    finally:
        cur.close()
        conn.close()

    nodes = [
        {
            "id": name,
            "label": name,
            "object_name": node_object_names.get(name) or node_object_names.get(name.upper(), ""),
            "type": node_types.get(name, "UNKNOWN"),
            "level": int(node_levels.get(name, 0)),
            "lane": idx + 1,
        }
        for idx, name in enumerate(sorted(node_types.keys()))
    ]

    # Color is driven by node type (source/target type from rstran).
    color_map = {
        "RSDS": "#9aa0a6",
        "TRCS": "#ffd54a",
        "ADSO": "#5f8dff",
        "IOBJ": "#52c5a8",
        "HCPR": "#ff9f43",
        "ELEM": "#f36f9a",
        "DEST": "#7dd3fc",
        "UNKNOWN": "#b4bfd6",
    }

    return {
        "nodes": nodes,
        "edges": edges,
        "color_map": color_map,
        "stats": {
            "start": seed,
            "node_count": len(nodes),
            "edge_count": len(edges),
        },
    }


def build_graph_upstream(start_name: str, max_nodes: int = 2000, max_edges: int = 5000) -> Dict[str, object]:
    """Canonical app mode: upstream (向上追溯)."""
    return _build_graph_engine_by_source(start_name, max_nodes=max_nodes, max_edges=max_edges)


def build_graph_downstream(start_name: str, max_nodes: int = 2000, max_edges: int = 5000) -> Dict[str, object]:
    """Canonical app mode: downstream (向下追溯)."""
    return _build_graph_engine_by_target(start_name, max_nodes=max_nodes, max_edges=max_edges)


def build_graph_both(start_name: str, max_nodes: int = 2000, max_edges: int = 5000) -> Dict[str, object]:
    """Canonical app mode: both (向上+向下), aligned to reference mode=3 semantics.

    Execution order follows the reference Python script:
    1) TARGETNAME -> SOURCENAME wave (mode1 path)
    2) SOURCENAME -> TARGETNAME wave (mode2 path)
    Both waves share one visited-row set so each row is emitted at most once.
    """
    seed = start_name.strip()
    if not seed:
        return {"nodes": [], "edges": [], "color_map": {}}

    conn = get_conn()
    cur = conn.cursor()

    visited_row_indices: Set[int] = set()
    node_types: Dict[str, str] = {seed: "UNKNOWN"}
    node_levels: Dict[str, int] = {seed: 0}
    node_object_names: Dict[str, str] = {}
    edge_seen: Set[Tuple[str, str, str, str]] = set()
    edge_by_key: Dict[Tuple[str, str, str, str], Dict[str, object]] = {}
    edges: List[Dict[str, object]] = []

    try:
        cur.execute(
            """
            SELECT SOURCENAME, TARGETNAME, SOURCETYPE, TARGETTYPE, TRANID, STARTROUTINE, ENDROUTINE, EXPERT
            FROM rstran
            WHERE SOURCENAME IS NOT NULL AND TARGETNAME IS NOT NULL
            ORDER BY TRANID
            """
        )
        all_rows = cur.fetchall()

        source_to_indices: Dict[str, List[int]] = {}
        target_to_indices: Dict[str, List[int]] = {}
        normalized_rows: List[Tuple[str, str, str, str, object, object, object, object]] = []

        for idx, row in enumerate(all_rows):
            source_name = str(row[0] or "").strip()
            target_name = str(row[1] or "").strip()
            source_type = normalize_type_code(row[2])
            target_type = normalize_type_code(row[3])
            normalized_rows.append((source_name, target_name, source_type, target_type, row[4], row[5], row[6], row[7]))
            if source_name:
                source_to_indices.setdefault(source_name, []).append(idx)
            if target_name:
                target_to_indices.setdefault(target_name, []).append(idx)

        # Wave-1: TARGET -> SOURCE (reference mode1)
        current_keys: List[str] = [seed]
        depth = 0
        while current_keys:
            next_keys: List[str] = []
            for key in current_keys:
                for idx in target_to_indices.get(key, []):
                    if idx in visited_row_indices:
                        continue
                    visited_row_indices.add(idx)

                    source_name, target_name, source_type, target_type, tran_id, start_routine, end_routine, expert = normalized_rows[idx]
                    if not source_name or not target_name:
                        continue

                    edge_key = (source_name, target_name, source_type, target_type)
                    if edge_key not in edge_seen:
                        edge_seen.add(edge_key)
                    append_or_update_edge(
                        edges,
                        edge_by_key,
                        source_name,
                        target_name,
                        source_type,
                        target_type,
                        "",
                        "",
                        tran_id,
                        start_routine,
                        end_routine,
                        expert,
                    )

                    if source_name not in node_types or node_types[source_name] == "UNKNOWN":
                        node_types[source_name] = source_type
                    if target_name not in node_types or node_types[target_name] == "UNKNOWN":
                        node_types[target_name] = target_type

                    if target_name not in node_levels:
                        node_levels[target_name] = depth
                    if source_name not in node_levels:
                        node_levels[source_name] = depth + 1

                    next_keys.append(source_name)

                    if len(node_types) >= max_nodes or len(edges) >= max_edges:
                        break
                if len(node_types) >= max_nodes or len(edges) >= max_edges:
                    break

            current_keys = next_keys
            depth += 1
            if len(node_types) >= max_nodes or len(edges) >= max_edges:
                break

        # Wave-2: SOURCE -> TARGET (reference mode2) with the SAME visited rows.
        current_keys = [seed]
        depth = 0
        while current_keys:
            next_keys = []
            for key in current_keys:
                for idx in source_to_indices.get(key, []):
                    if idx in visited_row_indices:
                        continue
                    visited_row_indices.add(idx)

                    source_name, target_name, source_type, target_type, tran_id, start_routine, end_routine, expert = normalized_rows[idx]
                    if not source_name or not target_name:
                        continue

                    edge_key = (source_name, target_name, source_type, target_type)
                    if edge_key not in edge_seen:
                        edge_seen.add(edge_key)
                    append_or_update_edge(
                        edges,
                        edge_by_key,
                        source_name,
                        target_name,
                        source_type,
                        target_type,
                        "",
                        "",
                        tran_id,
                        start_routine,
                        end_routine,
                        expert,
                    )

                    if source_name not in node_types or node_types[source_name] == "UNKNOWN":
                        node_types[source_name] = source_type
                    if target_name not in node_types or node_types[target_name] == "UNKNOWN":
                        node_types[target_name] = target_type

                    # Use negative depth for second wave so seed stays near center.
                    if source_name not in node_levels:
                        node_levels[source_name] = -depth
                    if target_name not in node_levels:
                        node_levels[target_name] = -(depth + 1)

                    next_keys.append(target_name)

                    if len(node_types) >= max_nodes or len(edges) >= max_edges:
                        break
                if len(node_types) >= max_nodes or len(edges) >= max_edges:
                    break

            current_keys = next_keys
            depth += 1
            if len(node_types) >= max_nodes or len(edges) >= max_edges:
                break

        all_node_names = list(node_types.keys())
        batch_size = 500
        for start in range(0, len(all_node_names), batch_size):
            batch = [str(v).strip().upper() for v in all_node_names[start : start + batch_size] if str(v).strip()]
            if not batch:
                continue
            placeholders = ",".join(["%s"] * len(batch))
            cur.execute(
                f"""
                SELECT BW_OBJECT_NORM, MAX(NAME_EN) AS NAME_EN
                FROM bw_object_name
                WHERE BW_OBJECT_NORM IN ({placeholders})
                GROUP BY BW_OBJECT_NORM
                """,
                tuple(batch),
            )
            for bw_object, object_name in cur.fetchall():
                key = normalize_bw_object_lookup(bw_object)
                value = str(object_name or "").strip()
                if key and value:
                    node_object_names[key] = value
    finally:
        cur.close()
        conn.close()

    nodes = [
        {
            "id": name,
            "label": name,
            "object_name": node_object_names.get(name) or node_object_names.get(name.upper(), ""),
            "type": node_types.get(name, "UNKNOWN"),
            "level": int(node_levels.get(name, 0)),
            "lane": idx + 1,
        }
        for idx, name in enumerate(sorted(node_types.keys()))
    ]

    color_map = {
        "RSDS": "#9aa0a6",
        "TRCS": "#ffd54a",
        "ADSO": "#5f8dff",
        "IOBJ": "#52c5a8",
        "HCPR": "#ff9f43",
        "ELEM": "#f36f9a",
        "DEST": "#7dd3fc",
        "UNKNOWN": "#b4bfd6",
    }

    return {
        "nodes": nodes,
        "edges": edges,
        "color_map": color_map,
        "stats": {
            "start": seed,
            "mode": "both",
            "node_count": len(nodes),
            "edge_count": len(edges),
        },
    }


def build_graph_full(start_name: str, max_nodes: int = 2000, max_edges: int = 5000) -> Dict[str, object]:
    """Canonical app mode: full (全量数据流), aligned to reference mode=4 semantics.

    Strategy:
    1) Treat each row as an undirected adjacency for reachability expansion.
    2) Starting from seed, expand to all connected neighbors until convergence.
    3) Render the induced directed subgraph on this connected component.

    This means mode=full keeps expanding from newly found nodes in BOTH directions,
    so direction changes are allowed at every hop.
    """
    seed = start_name.strip()
    if not seed:
        return {"nodes": [], "edges": [], "color_map": {}}

    conn = get_conn()
    cur = conn.cursor()

    node_types: Dict[str, str] = {seed: "UNKNOWN"}
    node_object_names: Dict[str, str] = {}
    edges: List[Dict[str, object]] = []

    try:
        cur.execute(
            """
            SELECT SOURCENAME, TARGETNAME, SOURCETYPE, TARGETTYPE, TRANID, STARTROUTINE, ENDROUTINE, EXPERT
            FROM rstran
            WHERE SOURCENAME IS NOT NULL AND TARGETNAME IS NOT NULL
            ORDER BY TRANID
            """
        )
        all_rows = cur.fetchall()

        normalized_rows: List[Tuple[str, str, str, str, object, object, object, object]] = []
        undirected_adj_all: Dict[str, Set[str]] = {}

        for row in all_rows:
            source_name = str(row[0] or "").strip()
            target_name = str(row[1] or "").strip()
            source_type = normalize_type_code(row[2])
            target_type = normalize_type_code(row[3])
            if not source_name or not target_name:
                continue

            normalized_rows.append((source_name, target_name, source_type, target_type, row[4], row[5], row[6], row[7]))
            undirected_adj_all.setdefault(source_name, set()).add(target_name)
            undirected_adj_all.setdefault(target_name, set()).add(source_name)

            if source_name not in node_types or node_types[source_name] == "UNKNOWN":
                node_types[source_name] = source_type
            if target_name not in node_types or node_types[target_name] == "UNKNOWN":
                node_types[target_name] = target_type

        # Full closure: expand connected component with direction switch allowed.
        all_nodes: Set[str] = set()
        queue: List[str] = [seed]
        while queue:
            cur_key = queue.pop(0)
            if cur_key in all_nodes:
                continue
            all_nodes.add(cur_key)
            for nxt in undirected_adj_all.get(cur_key, set()):
                if nxt not in all_nodes:
                    queue.append(nxt)
            if len(all_nodes) >= max_nodes:
                break

        if seed:
            all_nodes.add(seed)

        # Build induced full subgraph over the closure union.
        edge_seen: Set[Tuple[str, str, str, str]] = set()
        edge_by_key: Dict[Tuple[str, str, str, str], Dict[str, object]] = {}
        undirected_adj: Dict[str, Set[str]] = {}
        for source_name, target_name, source_type, target_type, tran_id, start_routine, end_routine, expert in normalized_rows:
            if source_name not in all_nodes or target_name not in all_nodes:
                continue
            key = (source_name, target_name, source_type, target_type)
            if key in edge_seen:
                append_or_update_edge(
                    edges,
                    edge_by_key,
                    source_name,
                    target_name,
                    source_type,
                    target_type,
                    "",
                    "",
                    tran_id,
                    start_routine,
                    end_routine,
                    expert,
                )
                continue
            edge_seen.add(key)
            append_or_update_edge(
                edges,
                edge_by_key,
                source_name,
                target_name,
                source_type,
                target_type,
                "",
                "",
                tran_id,
                start_routine,
                end_routine,
                expert,
            )
            undirected_adj.setdefault(source_name, set()).add(target_name)
            undirected_adj.setdefault(target_name, set()).add(source_name)
            if len(edges) >= max_edges:
                break

        # Level for full mode: shortest undirected distance from seed.
        node_levels: Dict[str, int] = {seed: 0}
        queue = [seed]
        while queue:
            cur_key = queue.pop(0)
            base = node_levels.get(cur_key, 0)
            for nxt in undirected_adj.get(cur_key, set()):
                if nxt in node_levels:
                    continue
                node_levels[nxt] = base + 1
                queue.append(nxt)

        # Enrich object names for all discovered nodes.
        all_node_names = list(all_nodes)
        batch_size = 500
        for start in range(0, len(all_node_names), batch_size):
            batch = [str(v).strip().upper() for v in all_node_names[start : start + batch_size] if str(v).strip()]
            if not batch:
                continue
            placeholders = ",".join(["%s"] * len(batch))
            cur.execute(
                f"""
                SELECT BW_OBJECT_NORM, MAX(NAME_EN) AS NAME_EN
                FROM bw_object_name
                WHERE BW_OBJECT_NORM IN ({placeholders})
                GROUP BY BW_OBJECT_NORM
                """,
                tuple(batch),
            )
            for bw_object, object_name in cur.fetchall():
                key = normalize_bw_object_lookup(bw_object)
                value = str(object_name or "").strip()
                if key and value:
                    node_object_names[key] = value
    finally:
        cur.close()
        conn.close()

    nodes = [
        {
            "id": name,
            "label": name,
            "object_name": node_object_names.get(name) or node_object_names.get(name.upper(), ""),
            "type": node_types.get(name, "UNKNOWN"),
            "level": int(node_levels.get(name, 0)),
            "lane": idx + 1,
        }
        for idx, name in enumerate(sorted(all_nodes))
    ]

    color_map = {
        "RSDS": "#9aa0a6",
        "TRCS": "#ffd54a",
        "ADSO": "#5f8dff",
        "IOBJ": "#52c5a8",
        "HCPR": "#ff9f43",
        "ELEM": "#f36f9a",
        "DEST": "#7dd3fc",
        "UNKNOWN": "#b4bfd6",
    }

    return {
        "nodes": nodes,
        "edges": edges,
        "color_map": color_map,
        "stats": {
            "start": seed,
            "mode": "full",
            "node_count": len(nodes),
            "edge_count": len(edges),
        },
    }


def _build_graph_engine_by_target(start_name: str, max_nodes: int = 2000, max_edges: int = 5000) -> Dict[str, object]:
    """Wave expansion by TARGETNAME -> SOURCENAME with row-level deduplication."""
    seed = start_name.strip()
    if not seed:
        return {"nodes": [], "edges": [], "color_map": {}}

    conn = get_conn()
    cur = conn.cursor()

    visited_row_indices: Set[int] = set()
    current_keys: List[str] = [seed]
    node_types: Dict[str, str] = {seed: "UNKNOWN"}
    node_levels: Dict[str, int] = {seed: 0}
    node_object_names: Dict[str, str] = {}
    edge_seen: Set[Tuple[str, str, str, str]] = set()
    edge_by_key: Dict[Tuple[str, str, str, str], Dict[str, object]] = {}
    edges: List[Dict[str, object]] = []

    try:
        cur.execute(
            """
            SELECT SOURCENAME, TARGETNAME, SOURCETYPE, TARGETTYPE, TRANID, STARTROUTINE, ENDROUTINE, EXPERT
            FROM rstran
            WHERE SOURCENAME IS NOT NULL AND TARGETNAME IS NOT NULL
            ORDER BY TRANID
            """
        )
        all_rows = cur.fetchall()

        target_to_indices: Dict[str, List[int]] = {}
        normalized_rows: List[Tuple[str, str, str, str, object, object, object, object]] = []
        for idx, row in enumerate(all_rows):
            source_name = str(row[0] or "").strip()
            target_name = str(row[1] or "").strip()
            source_type = normalize_type_code(row[2])
            target_type = normalize_type_code(row[3])
            normalized_rows.append((source_name, target_name, source_type, target_type, row[4], row[5], row[6], row[7]))
            if not target_name:
                continue
            target_to_indices.setdefault(target_name, []).append(idx)

        depth = 0
        while current_keys:
            next_keys: List[str] = []
            for key in current_keys:
                for idx in target_to_indices.get(key, []):
                    if idx in visited_row_indices:
                        continue
                    visited_row_indices.add(idx)

                    source_name, target_name, source_type, target_type, tran_id, start_routine, end_routine, expert = normalized_rows[idx]
                    if not source_name or not target_name:
                        continue

                    edge_key = (source_name, target_name, source_type, target_type)
                    if edge_key not in edge_seen:
                        edge_seen.add(edge_key)
                    append_or_update_edge(
                        edges,
                        edge_by_key,
                        source_name,
                        target_name,
                        source_type,
                        target_type,
                        "",
                        "",
                        tran_id,
                        start_routine,
                        end_routine,
                        expert,
                    )

                    if source_name not in node_types or node_types[source_name] == "UNKNOWN":
                        node_types[source_name] = source_type
                    if target_name not in node_types or node_types[target_name] == "UNKNOWN":
                        node_types[target_name] = target_type

                    if target_name not in node_levels:
                        node_levels[target_name] = depth
                    if source_name not in node_levels:
                        node_levels[source_name] = depth + 1

                    # Continue matching next key by SOURCENAME.
                    next_keys.append(source_name)

                    if len(node_types) >= max_nodes or len(edges) >= max_edges:
                        break
                if len(node_types) >= max_nodes or len(edges) >= max_edges:
                    break

            current_keys = next_keys
            depth += 1

            if len(node_types) >= max_nodes or len(edges) >= max_edges:
                break

        all_node_names = list(node_types.keys())
        batch_size = 500
        for start in range(0, len(all_node_names), batch_size):
            batch = [str(v).strip().upper() for v in all_node_names[start : start + batch_size] if str(v).strip()]
            if not batch:
                continue
            placeholders = ",".join(["%s"] * len(batch))
            cur.execute(
                f"""
                SELECT BW_OBJECT_NORM, MAX(NAME_EN) AS NAME_EN
                FROM bw_object_name
                WHERE BW_OBJECT_NORM IN ({placeholders})
                GROUP BY BW_OBJECT_NORM
                """,
                tuple(batch),
            )
            for bw_object, object_name in cur.fetchall():
                key = normalize_bw_object_lookup(bw_object)
                value = str(object_name or "").strip()
                if key and value:
                    node_object_names[key] = value
    finally:
        cur.close()
        conn.close()

    nodes = [
        {
            "id": name,
            "label": name,
            "object_name": node_object_names.get(name) or node_object_names.get(name.upper(), ""),
            "type": node_types.get(name, "UNKNOWN"),
            "level": int(node_levels.get(name, 0)),
            "lane": idx + 1,
        }
        for idx, name in enumerate(sorted(node_types.keys()))
    ]

    color_map = {
        "RSDS": "#9aa0a6",
        "TRCS": "#ffd54a",
        "ADSO": "#5f8dff",
        "IOBJ": "#52c5a8",
        "HCPR": "#ff9f43",
        "ELEM": "#f36f9a",
        "DEST": "#7dd3fc",
        "UNKNOWN": "#b4bfd6",
    }

    return {
        "nodes": nodes,
        "edges": edges,
        "color_map": color_map,
        "stats": {
            "start": seed,
            "node_count": len(nodes),
            "edge_count": len(edges),
        },
    }


def upsert_status(table_name: str, item_count: int = 0) -> Dict[str, str | int]:
    if IS_SQLITE:
        conn = get_conn()
        cur = conn.cursor()
        try:
            last_import_at = utcnow().strftime("%Y-%m-%d %H:%M:%S")
            cur.execute(
                """
                INSERT INTO import_status (table_name, last_import_at, last_import_count)
                VALUES (%s, %s, %s)
                ON CONFLICT(table_name) DO UPDATE SET
                  last_import_at = excluded.last_import_at,
                  last_import_count = excluded.last_import_count,
                  updated_at = CURRENT_TIMESTAMP
                """,
                (table_name, last_import_at, item_count),
            )
            conn.commit()
            cur.execute(
                "SELECT last_import_at, last_import_count FROM import_status WHERE table_name = %s",
                (table_name,),
            )
            row = cur.fetchone()
        finally:
            cur.close()
            conn.close()

        last_update = "--"
        if row and row[0]:
            parsed = parse_datetime_like(row[0])
            last_update = parsed.strftime("%Y-%m-%d %H:%M") if isinstance(parsed, datetime) else str(row[0])[:16].replace("T", " ")
        return {
            "last_update": last_update,
            "last_count": int(row[1]) if row and row[1] is not None else 0,
        }
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO import_status (table_name, last_import_at, last_import_count)
        VALUES (%s, NOW(), %s)
        ON DUPLICATE KEY UPDATE
          last_import_at = VALUES(last_import_at),
          last_import_count = VALUES(last_import_count)
        """,
        (table_name, item_count),
    )
    conn.commit()
    cur.execute(
        """
        SELECT DATE_FORMAT(last_import_at, '%Y-%m-%d %H:%i'), last_import_count
        FROM import_status
        WHERE table_name = %s
        """,
        (table_name,),
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return {
        "last_update": row[0] if row and row[0] else "--",
        "last_count": int(row[1]) if row and row[1] is not None else 0,
    }


def parse_upload_to_dataframe(
    upload_file: UploadFile,
    sheet_name: str | None,
    header_row_num: int = 1,
) -> pd.DataFrame:
    filename = (upload_file.filename or "").lower()
    content = upload_file.file.read()
    header_index = max(0, int(header_row_num or 1) - 1)

    if filename.endswith(".csv"):
        return pd.read_csv(io.BytesIO(content), dtype=str, header=header_index).fillna("")

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        target_sheet = sheet_name if sheet_name else 0
        return pd.read_excel(
            io.BytesIO(content),
            sheet_name=target_sheet,
            dtype=str,
            header=header_index,
        ).fillna("")

    raise HTTPException(status_code=400, detail="Unsupported file type, only xlsx/xls/csv")


def normalize_import_chunk_dataframe(source_df: pd.DataFrame) -> pd.DataFrame:
    if source_df.empty:
        return source_df
    return source_df.fillna("")


def iter_csv_upload_chunks(
    upload_file: UploadFile,
    header_row_num: int,
    chunk_size: int = IMPORT_READ_CHUNK_SIZE,
) -> Iterator[pd.DataFrame]:
    upload_file.file.seek(0)
    text_stream = io.TextIOWrapper(upload_file.file, encoding="utf-8-sig", newline="")
    try:
        chunk_iter = pd.read_csv(
            text_stream,
            dtype=str,
            header=max(0, int(header_row_num or 1) - 1),
            chunksize=chunk_size,
            keep_default_na=False,
        )
        for chunk in chunk_iter:
            normalized = normalize_import_chunk_dataframe(chunk)
            if not normalized.empty:
                yield normalized
    finally:
        try:
            text_stream.detach()
        except Exception:
            pass
        upload_file.file.seek(0)


def iter_xlsx_upload_chunks(
    upload_file: UploadFile,
    sheet_name: str | None,
    header_row_num: int,
    chunk_size: int = IMPORT_READ_CHUNK_SIZE,
) -> Iterator[pd.DataFrame]:
    upload_file.file.seek(0)
    workbook = load_workbook(upload_file.file, read_only=True, data_only=True)
    try:
        available_sheets = list(workbook.sheetnames or [])
        default_sheet = available_sheets[0] if available_sheets else ""
        target_sheet = str(sheet_name or default_sheet).strip()
        if not target_sheet or target_sheet not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail="导入失败：未找到指定的 Sheet。")

        worksheet = workbook[target_sheet]
        header_values: List[str] | None = None
        chunk_rows: List[List[str]] = []
        header_index = max(1, int(header_row_num or 1))

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            normalized_row = ["" if value is None else str(value) for value in (row or ())]
            if row_index < header_index:
                continue
            if row_index == header_index:
                header_values = [str(value or "").strip() for value in normalized_row]
                continue
            if not any(str(value or "").strip() for value in normalized_row):
                continue
            chunk_rows.append(normalized_row)
            if len(chunk_rows) >= chunk_size:
                yield pd.DataFrame(chunk_rows, columns=header_values)
                chunk_rows = []

        if header_values is None:
            raise ValueError(f"标题行数第{header_row_num}行超出文件有效范围")

        if chunk_rows:
            yield pd.DataFrame(chunk_rows, columns=header_values)
    finally:
        workbook.close()
        upload_file.file.seek(0)


def iter_upload_dataframe_chunks(
    upload_file: UploadFile,
    sheet_name: str | None,
    header_row_num: int = 1,
    chunk_size: int = IMPORT_READ_CHUNK_SIZE,
) -> Iterator[pd.DataFrame]:
    filename = (upload_file.filename or "").lower()

    if filename.endswith(".csv"):
        yield from iter_csv_upload_chunks(upload_file, header_row_num, chunk_size=chunk_size)
        return

    if filename.endswith(".xlsx"):
        yield from iter_xlsx_upload_chunks(upload_file, sheet_name, header_row_num, chunk_size=chunk_size)
        return

    if filename.endswith(".xls"):
        source_df = parse_upload_to_dataframe(upload_file, sheet_name, header_row_num)
        if not source_df.empty:
            yield source_df
        return

    raise HTTPException(status_code=400, detail="Unsupported file type, only xlsx/xls/csv")


def apply_rstran_logic(mapped_df: pd.DataFrame) -> pd.DataFrame:
    if "SOURCENAME" in mapped_df.columns:
        split_vals = mapped_df["SOURCENAME"].astype(str).str.strip().str.split(r"\s+", n=1, expand=True)
        first_part = split_vals[0].fillna("") if 0 in split_vals.columns else ""
        second_part = split_vals[1].fillna("") if 1 in split_vals.columns else ""

        if "SOURCE" in mapped_df.columns:
            mapped_df["SOURCE"] = first_part
        if "SOURCESYS" in mapped_df.columns:
            mapped_df["SOURCESYS"] = second_part

    return mapped_df


PUBLIC_API_PATHS = {
    "/api/auth/login",
}


@app.middleware("http")
async def auth_guard(request: Request, call_next):
    path = request.url.path
    if not path.startswith("/api/"):
        return await call_next(request)

    if request.method == "OPTIONS":
        return await call_next(request)

    if path in PUBLIC_API_PATHS:
        return await call_next(request)

    user = resolve_user_from_request(request)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "未登录或会话已过期"})

    request.state.current_user = user

    if path.startswith("/api/admin/") and user.get("role") != "admin":
        return JSONResponse(status_code=403, content={"detail": "需要管理员权限"})

    return await call_next(request)


@app.on_event("startup")
def startup() -> None:
    if IS_SQLITE:
        ensure_status_table()
        ensure_prog_code_schema()
        ensure_rstran_mapping_rule_full_table()
        ensure_bw_object_field_inventory_table()
        ensure_auth_tables()
        ensure_user_hidden_object_table()
        return
    ensure_status_table()
    ensure_prog_code_schema()
    migrate_legacy_infoobject_tables()
    ensure_rstran_schema()
    ensure_rstran_mapping_rule_table()
    ensure_rstran_mapping_rule_full_table()
    ensure_bw_object_field_inventory_table()
    ensure_bw_object_name_schema()
    ensure_dd03l_schema()
    ensure_dd02t_schema()
    ensure_dd03t_schema()
    ensure_dd04t_schema()
    ensure_rsdiobj_schema()
    ensure_rsdiobjt_schema()
    ensure_rsoadso_schema()
    ensure_rsdssegfd_schema()
    ensure_rsdssegfdt_schema()
    ensure_rsksnew_schema()
    ensure_rsksnewt_schema()
    ensure_rsksfieldnew_schema()
    ensure_rsksfieldnewt_schema()
    ensure_rstrant_schema()
    ensure_rstranstepcnst_schema()
    ensure_rstransteprout_schema()
    ensure_auth_tables()
    ensure_user_hidden_object_table()


@app.post("/api/auth/login")
def auth_login(payload: LoginRequest, request: Request, response: Response) -> LoginResponse:
    username = normalize_username(payload.username)
    password = str(payload.password or "")
    if not username or not password:
        raise HTTPException(status_code=400, detail="请输入用户名和密码")

    user = fetch_user_by_username(username)
    if not user:
        audit_log("login", username, False, detail="user_not_found")
        raise HTTPException(status_code=401, detail="用户名或密码错误")

    temp_lock_until = user.get("temp_lock_until")
    if temp_lock_until and isinstance(temp_lock_until, datetime) and temp_lock_until > utcnow():
        remain_seconds = max(1, int((temp_lock_until - utcnow()).total_seconds()))
        remain_minutes = max(1, (remain_seconds + 59) // 60)
        audit_log("login", username, False, detail="temp_locked")
        raise HTTPException(status_code=423, detail=f"登录失败次数过多，请 {remain_minutes} 分钟后再试")
    if int(user.get("is_locked") or 0) == 1:
        audit_log("login", username, False, detail="admin_locked")
        raise HTTPException(status_code=423, detail="用户已被锁定，请联系管理员")

    if not verify_password(password, str(user.get("password_hash") or "")):
        failed_attempts = int(user.get("failed_attempts") or 0) + 1
        lock_until = None
        lock_triggered = False
        if failed_attempts >= AUTH_LOGIN_MAX_FAILS:
            lock_until = utcnow() + timedelta(minutes=AUTH_TEMP_LOCK_MINUTES)
            failed_attempts = 0
            lock_triggered = True

        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                "UPDATE users SET failed_attempts = %s, temp_lock_until = %s WHERE username = %s",
                (failed_attempts, lock_until, username),
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()

        if lock_triggered:
            audit_log("login", username, False, detail="lock_triggered")
            raise HTTPException(status_code=423, detail=f"登录失败次数过多，请 {AUTH_TEMP_LOCK_MINUTES} 分钟后再试")

        remain_attempts = max(0, AUTH_LOGIN_MAX_FAILS - failed_attempts)
        audit_log("login", username, False, detail="invalid_password")
        raise HTTPException(status_code=401, detail=f"用户名或密码错误，还可尝试 {remain_attempts} 次")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE users SET failed_attempts = 0, temp_lock_until = NULL, last_login_at = %s WHERE username = %s",
            (utcnow(), username),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    token, expires_at = create_session(username)
    max_age = int(max(1, (expires_at - utcnow()).total_seconds()))
    response.set_cookie(
        key=AUTH_COOKIE_NAME,
        value=token,
        httponly=True,
        secure=AUTH_COOKIE_SECURE or AUTH_COOKIE_SAMESITE == "none",
        samesite=AUTH_COOKIE_SAMESITE,
        domain=AUTH_COOKIE_DOMAIN,
        max_age=max_age,
        path="/",
    )
    audit_log("login", username, True)
    return {
        "username": username,
        "role": user.get("role", "user"),
        "session_token": token,
    }


@app.post("/api/auth/logout")
def auth_logout(request: Request, response: Response) -> Dict[str, str]:
    raw_token = get_request_session_token(request)
    actor = None
    if hasattr(request.state, "current_user") and request.state.current_user:
        actor = request.state.current_user.get("username")
    if raw_token:
        revoke_session(raw_token)
    response.delete_cookie(AUTH_COOKIE_NAME, path="/", domain=AUTH_COOKIE_DOMAIN)
    audit_log("logout", actor, True)
    return {"message": "ok"}


@app.get("/api/auth/me")
def auth_me(request: Request) -> Dict[str, str]:
    user = getattr(request.state, "current_user", None)
    if not user:
        raise HTTPException(status_code=401, detail="未登录")
    return {"username": user.get("username", ""), "role": user.get("role", "user")}


@app.post("/api/auth/change-password")
def auth_change_password(payload: ChangePasswordRequest, request: Request) -> Dict[str, str]:
    user = getattr(request.state, "current_user", None)
    if not user:
        raise HTTPException(status_code=401, detail="未登录")

    username = normalize_username(user.get("username", ""))
    row = fetch_user_by_username(username)
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")
    if not verify_password(payload.current_password, str(row.get("password_hash") or "")):
        audit_log("change_password", username, False, detail="wrong_current_password", actor=username)
        raise HTTPException(status_code=400, detail="当前密码错误")

    validate_password_strength(payload.new_password)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE users SET password_hash = %s, failed_attempts = 0, temp_lock_until = NULL WHERE username = %s",
            (hash_password(payload.new_password), username),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    audit_log("change_password", username, True, actor=username)
    return {"message": "密码修改成功"}


@app.get("/api/admin/users")
def admin_list_users() -> Dict[str, List[Dict[str, object]]]:
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            """
            SELECT username, role, is_locked, temp_lock_until, last_login_at, created_at
            FROM users
            ORDER BY created_at ASC, username ASC
            """
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    users = []
    for row in rows:
        users.append(
            {
                "username": row.get("username", ""),
                "role": row.get("role", "user"),
                "is_locked": bool(row.get("is_locked") or 0),
                "temp_lock_until": row.get("temp_lock_until").isoformat() if row.get("temp_lock_until") else None,
                "last_login_at": row.get("last_login_at").isoformat() if row.get("last_login_at") else None,
                "created_at": row.get("created_at").isoformat() if row.get("created_at") else None,
            }
        )
    return {"users": users}


@app.post("/api/admin/users")
def admin_create_user(payload: AdminCreateUserRequest, request: Request) -> Dict[str, str]:
    actor = request.state.current_user.get("username") if hasattr(request.state, "current_user") else None
    username = normalize_username(payload.username)
    role = str(payload.role or "").strip().lower()
    if not username:
        raise HTTPException(status_code=400, detail="用户名不能为空")
    if role not in {"admin", "user"}:
        raise HTTPException(status_code=400, detail="角色必须是 admin 或 user")
    validate_password_strength(payload.password)

    if fetch_user_by_username(username):
        raise HTTPException(status_code=409, detail="用户名已存在")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO users (username, password_hash, role, is_locked, failed_attempts) VALUES (%s, %s, %s, 0, 0)",
            (username, hash_password(payload.password), role),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    audit_log("admin_create_user", username, True, actor=actor)
    return {"message": "用户创建成功"}


@app.post("/api/admin/users/{username}/lock")
def admin_lock_user(username: str, request: Request) -> Dict[str, str]:
    actor = request.state.current_user.get("username") if hasattr(request.state, "current_user") else None
    target = normalize_username(username)
    actor_name = normalize_username(actor or "")
    if target == actor_name:
        raise HTTPException(status_code=400, detail="不能锁定当前登录用户")
    if not fetch_user_by_username(target):
        raise HTTPException(status_code=404, detail="用户不存在")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE users SET is_locked = 1 WHERE username = %s", (target,))
        cur.execute("UPDATE user_sessions SET revoked = 1 WHERE username = %s", (target,))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    audit_log("admin_lock_user", target, True, actor=actor)
    return {"message": "用户已锁定"}


@app.post("/api/admin/users/{username}/unlock")
def admin_unlock_user(username: str, request: Request) -> Dict[str, str]:
    actor = request.state.current_user.get("username") if hasattr(request.state, "current_user") else None
    target = normalize_username(username)
    if not fetch_user_by_username(target):
        raise HTTPException(status_code=404, detail="用户不存在")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE users SET is_locked = 0, failed_attempts = 0, temp_lock_until = NULL WHERE username = %s",
            (target,),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    audit_log("admin_unlock_user", target, True, actor=actor)
    return {"message": "用户已解锁"}


@app.post("/api/admin/users/{username}/reset-password")
def admin_reset_password(username: str, payload: AdminResetPasswordRequest, request: Request) -> Dict[str, str]:
    actor = request.state.current_user.get("username") if hasattr(request.state, "current_user") else None
    target = normalize_username(username)
    if not fetch_user_by_username(target):
        raise HTTPException(status_code=404, detail="用户不存在")

    validate_password_strength(payload.new_password)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            UPDATE users
            SET password_hash = %s, failed_attempts = 0, temp_lock_until = NULL
            WHERE username = %s
            """,
            (hash_password(payload.new_password), target),
        )
        cur.execute("UPDATE user_sessions SET revoked = 1 WHERE username = %s", (target,))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    audit_log("admin_reset_password", target, True, actor=actor)
    return {"message": "密码重置成功"}


@app.delete("/api/admin/users/{username}")
def admin_delete_user(username: str, request: Request) -> Dict[str, str]:
    actor = request.state.current_user.get("username") if hasattr(request.state, "current_user") else None
    target = normalize_username(username)
    actor_name = normalize_username(actor or "")
    if not target:
        raise HTTPException(status_code=400, detail="用户名不能为空")
    if target == actor_name:
        raise HTTPException(status_code=400, detail="不能删除当前登录用户")

    row = fetch_user_by_username(target)
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")

    conn = get_conn()
    cur = conn.cursor()
    try:
        if str(row.get("role") or "user").lower() == "admin":
            cur.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
            admin_count = int(cur.fetchone()[0] or 0)
            if admin_count <= 1:
                raise HTTPException(status_code=400, detail="至少保留一个管理员用户")

        cur.execute("DELETE FROM users WHERE username = %s", (target,))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    audit_log("admin_delete_user", target, True, actor=actor)
    return {"message": "用户删除成功"}


@app.get("/api/import-status")
def get_import_status() -> Dict[str, Dict[str, str | int]]:
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT table_name, last_import_at, last_import_count FROM import_status")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    response: Dict[str, Dict[str, str | int]] = {}
    for row in rows:
        dt = row["last_import_at"]
        response[str(row["table_name"] or "").strip()] = {
            "last_update": dt.strftime("%Y-%m-%d %H:%M") if isinstance(dt, datetime) else "--",
            "last_count": int(row["last_import_count"] or 0),
        }

    # Keep card status aligned with actual DB content, including data imported outside this app.
    for table in IMPORT_STATUS_TRACKED_TABLES:
        table_name = str(table or "").strip()
        current = response.get(table_name, {"last_update": "--", "last_count": 0})
        actual_count = count_table_rows(table_name)
        current_count = int(current.get("last_count") or 0)

        if current_count != actual_count:
            synced = upsert_status(table_name, actual_count)
            response[table_name] = {
                "last_update": str(synced.get("last_update") or "--"),
                "last_count": int(synced.get("last_count") or 0),
            }
        else:
            response[table_name] = {
                "last_update": str(current.get("last_update") or "--"),
                "last_count": current_count,
            }

    return response


@app.get("/api/import-schema")
def get_import_schema(table_name: str = Query(...)) -> Dict[str, object]:
    normalized_name = str(table_name or "").strip().lower()
    if normalized_name not in IMPORT_SCHEMA_TABLES:
        raise HTTPException(status_code=400, detail="Unsupported table_name")

    ensure_import_target_table_schema(normalized_name)
    columns = get_table_schema(normalized_name)
    if not columns:
        raise HTTPException(status_code=404, detail=f"Target table not found: {normalized_name}")

    return {
        "table_name": normalized_name,
        "columns": columns,
    }


@app.get("/api/data-query/tables")
def get_data_query_tables() -> Dict[str, object]:
    table_names = list_queryable_table_names()
    return {
        "tables": [
            {
                "name": table_name,
                "label": DDIC_METADATA_TABLES.get(table_name, {}).get("metadata_tabname") or table_name.upper(),
            }
            for table_name in table_names
        ]
    }


@app.get("/api/data-query/schema")
def get_data_query_schema(table_name: str = Query(...)) -> Dict[str, object]:
    normalized_name = ensure_queryable_table_name(table_name)
    columns = get_table_schema(normalized_name)
    return {
        "table_name": normalized_name,
        "columns": columns,
    }


@app.post("/api/data-query/query")
def run_data_query(payload: DataQueryRequest) -> Dict[str, object]:
    return execute_data_query(payload)


@app.post("/api/import-status/upsert")
def upsert_import_status(payload: ImportStatusUpdate) -> Dict[str, str | int]:
    table_name = payload.table_name.strip()
    if table_name not in IMPORT_STATUS_TRACKED_TABLES:
        raise HTTPException(status_code=400, detail="Unsupported table_name")

    status = upsert_status(table_name, count_table_rows(table_name))
    return {"table_name": table_name, "last_update": status["last_update"], "last_count": status["last_count"]}


@app.post("/api/import/prog-code/upsert")
def upsert_prog_code(payload: ProgCodeUpsertRequest) -> Dict[str, str | int]:
    ensure_prog_code_schema()

    prog_id = str(payload.prog_id or "").strip()
    prog_type = str(payload.prog_type or "").strip()
    prog_code = normalize_prog_code_text(payload.prog_code)

    if not prog_id:
        raise HTTPException(status_code=400, detail="Program ID is required")
    if len(prog_id) > 40:
        raise HTTPException(status_code=400, detail="Program ID exceeds 40 characters")
    if len(prog_type) > 20:
        raise HTTPException(status_code=400, detail="Program Type exceeds 20 characters")
    if not prog_code.strip():
        raise HTTPException(status_code=400, detail="Program code is required")

    conn = get_conn()
    cur = conn.cursor()
    try:
        if IS_SQLITE:
            cur.execute(
                f"""
                INSERT INTO `{PROG_CODE_TABLE}` (`prog_id`, `prog_type`, `prog_code`)
                VALUES (%s, %s, %s)
                ON CONFLICT (`prog_id`) DO UPDATE SET
                  `prog_type` = excluded.`prog_type`,
                  `prog_code` = excluded.`prog_code`,
                  `updated_at` = CURRENT_TIMESTAMP
                """,
                (prog_id, prog_type, prog_code),
            )
        else:
            cur.execute(
                f"""
                INSERT INTO `{PROG_CODE_TABLE}` (`prog_id`, `prog_type`, `prog_code`)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE
                  `prog_type` = VALUES(`prog_type`),
                  `prog_code` = VALUES(`prog_code`)
                """,
                (prog_id, prog_type, prog_code),
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    status = upsert_status(PROG_CODE_TABLE, count_table_rows(PROG_CODE_TABLE))
    return {
        "table_name": PROG_CODE_TABLE,
        "prog_id": prog_id,
        "db_count": int(status["last_count"]),
        "last_update": status["last_update"],
        "last_count": int(status["last_count"]),
        "line_count": len(prog_code.split("\n")),
        "message": "Program code saved",
    }


@app.get("/api/import/prog-code/search")
def search_prog_code(q: str = Query(...), limit: int = Query(default=50, ge=1, le=200)) -> Dict[str, object]:
    ensure_prog_code_schema()

    keyword = str(q or "").strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="Query keyword is required")

    items = search_prog_code_entries(keyword, limit)
    return {
        "query": keyword,
        "count": len(items),
        "has_wildcard": "*" in keyword,
        "items": items,
    }


@app.get("/api/import/prog-code/item")
def get_prog_code_item(prog_id: str = Query(...)) -> Dict[str, object]:
    ensure_prog_code_schema()

    normalized_prog_id = str(prog_id or "").strip()
    if not normalized_prog_id:
        raise HTTPException(status_code=400, detail="Program ID is required")

    entry = fetch_prog_code_exact(normalized_prog_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Program not found")

    prog_code = normalize_prog_code_text(entry.get("prog_code"))
    return {
        "prog_id": str(entry.get("prog_id") or "").strip(),
        "prog_type": str(entry.get("prog_type") or "").strip(),
        "prog_code": prog_code,
        "updated_at": str(entry.get("updated_at") or "").strip(),
        "line_count": len(prog_code.split("\n")) if prog_code else 0,
    }


@app.post("/api/abap-to-sql/generate")
def generate_abap_to_sql(payload: AbapToSqlRequest) -> Dict[str, object]:
    ensure_prog_code_schema()

    output_table = str(payload.output_table or "").strip()
    abap_code = normalize_prog_code_text(payload.abap_code)

    if len(output_table) > 20:
        raise HTTPException(status_code=400, detail="Output table exceeds 20 characters")
    if not abap_code.strip():
        raise HTTPException(status_code=400, detail="ABAP code is required")

    resolution = resolve_abap_call_chain(abap_code)
    combined_parts = [abap_code] + [str(entry.get("prog_code") or "") for entry in (resolution.get("matched_entries") or [])]
    combined_code = "\n\n".join(part for part in combined_parts if str(part or "").strip())
    method_rules, ignored_aliases = derive_abap_method_call_rules(abap_code, resolution)
    function_rules, function_ignored_aliases = derive_abap_function_call_rules(abap_code, resolution)
    ignored_aliases.update(function_ignored_aliases)
    rules = [
        rule for rule in collect_abap_field_rules(combined_code)
        if str(rule.get("alias") or "") not in ignored_aliases
    ]
    program_call_rules = method_rules + function_rules
    program_call_rule_aliases = {str(rule.get("alias") or "") for rule in program_call_rules}
    rules = [rule for rule in rules if str(rule.get("alias") or "") not in program_call_rule_aliases] + program_call_rules
    comment_hints = collect_abap_comment_hints(abap_code) or collect_abap_comment_hints(combined_code)
    processing_notes = collect_abap_processing_notes(combined_code)
    rendered_output_table = output_table or "TARGET_TABLE"
    location_text = build_abap_location_text(rendered_output_table, resolution)
    sql_text = build_abap_pseudo_sql(rendered_output_table, rules, resolution, comment_hints, processing_notes, combined_code)

    return {
        "output_table": rendered_output_table,
        "location_text": location_text,
        "sql_text": sql_text,
        "detected_count": len(resolution.get("detected_refs") or []),
        "matched_count": len(resolution.get("matched_entries") or []),
        "missing_count": len(resolution.get("missing_refs") or []),
        "matched_programs": [entry.get("prog_id") for entry in (resolution.get("matched_entries") or [])],
        "missing_programs": [ref.get("name") for ref in (resolution.get("missing_refs") or [])],
    }


@app.post("/api/path-selection/search")
def path_selection_search(payload: PathSelectionRequest) -> Dict[str, object]:
    source_name = normalize_bw_object_lookup(payload.SOURCE)
    source_system = normalize_bw_object_lookup(payload.SOURCESYS)
    target_name = normalize_bw_object_lookup(payload.TARGETNAME)
    tran_id = str(payload.TRANID or "").strip()
    waypoints = [normalize_bw_object_lookup(item) for item in (payload.WAYPOINTS or []) if normalize_bw_object_lookup(item)]

    if tran_id and (source_name or source_system or target_name):
        raise HTTPException(status_code=400, detail="Enter either Source and Target, or Transformation ID")

    if tran_id:
        return search_candidate_paths_by_tran_id(tran_id)

    if not source_name or not target_name:
        raise HTTPException(status_code=400, detail="Enter both Source and Target, or Transformation ID")

    return search_candidate_paths(
        source_name=source_name,
        source_system=source_system,
        target_name=target_name,
        waypoints=waypoints,
    )


@app.post("/api/path-selection/mapping")
def path_selection_mapping(payload: PathMappingRequest) -> Dict[str, object]:
    if not payload.segments:
        return {
            "segment_count": 0,
            "segments": [],
            "include_logic": bool(payload.include_logic),
            "include_text": bool(payload.include_text),
        }
    try:
        return build_path_mapping_payload(
            payload.segments,
            include_logic=bool(payload.include_logic),
            include_text=bool(payload.include_text),
        )
    except MaterializedReadModelNotReadyError as exc:
        raise build_materialization_not_ready_http_exception(exc) from exc


@app.post("/api/path-selection/text")
def path_selection_text(payload: PathMappingRequest) -> Dict[str, object]:
    try:
        return build_path_text_payload(payload.segments)
    except MaterializedReadModelNotReadyError as exc:
        raise build_materialization_not_ready_http_exception(exc) from exc


@app.post("/api/path-selection/logic")
def path_selection_logic(payload: PathMappingRequest) -> Dict[str, object]:
    try:
        return build_path_logic_payload(payload.segments)
    except MaterializedReadModelNotReadyError as exc:
        raise build_materialization_not_ready_http_exception(exc) from exc


@app.post("/api/import/execute")
def execute_import(
    table_name: str = Form(...),
    mapping_json: str = Form(...),
    sheet_name: str = Form(""),
    header_row_num: int = Form(1),
    duplicate_mode: str = Form("fail"),
    file: UploadFile = File(...),
) -> Dict[str, str | int]:
    _ = duplicate_mode

    table_name = table_name.strip()
    if table_name not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Unsupported table_name")

    ensure_import_target_table_schema(table_name)

    try:
        mapping = json.loads(mapping_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid mapping_json") from exc

    if not isinstance(mapping, dict):
        raise HTTPException(status_code=400, detail="mapping_json must be an object")

    header_row_num = max(1, min(int(header_row_num or 1), 10))
    schema_rows = get_table_schema(table_name)
    db_columns = [str(item.get("name") or "").strip() for item in schema_rows if str(item.get("name") or "").strip()]
    if not db_columns:
        raise HTTPException(status_code=400, detail=f"Target table not found: {table_name}")
    schema_map = {str(item.get("name") or "").strip(): item for item in schema_rows}
    explicitly_mapped_fields = {
        col_name
        for col_name, source_field in mapping.items()
        if str(col_name or "").strip() in schema_map and str(source_field or "").strip()
    }
    key_fields = get_primary_keys(table_name)
    if not key_fields:
        raise HTTPException(status_code=400, detail=f"导入失败：表 {table_name} 未配置可用于更新的键字段")

    missing_key_cols = [k for k in key_fields if k not in db_columns]
    if missing_key_cols:
        raise HTTPException(status_code=400, detail=f"导入失败：键字段不存在于目标表：{', '.join(missing_key_cols)}")
    key_indices = [db_columns.index(key_field) for key_field in key_fields]

    col_sql = ", ".join(f"`{c}`" for c in db_columns)
    ph_sql = ", ".join(["%s"] * len(db_columns))
    insert_sql = f"INSERT INTO `{table_name}` ({col_sql}) VALUES ({ph_sql})"

    mutable_fields = [c for c in db_columns if c not in key_fields]
    mutable_update_fields = [c for c in mutable_fields if c in explicitly_mapped_fields]
    if mutable_update_fields:
        upsert_set_sql = ", ".join(f"`{c}` = VALUES(`{c}`)" for c in mutable_update_fields)
    else:
        # Fallback for key-only tables: perform a no-op update expression.
        first_key = key_fields[0]
        upsert_set_sql = f"`{first_key}` = `{first_key}`"
    if IS_SQLITE:
        conflict_target = ", ".join(f"`{c}`" for c in key_fields)
        if mutable_update_fields:
            sqlite_update_sql = ", ".join(f"`{c}` = excluded.`{c}`" for c in mutable_update_fields)
            upsert_sql = f"{insert_sql} ON CONFLICT ({conflict_target}) DO UPDATE SET {sqlite_update_sql}"
        else:
            upsert_sql = f"{insert_sql} ON CONFLICT ({conflict_target}) DO NOTHING"
    else:
        upsert_sql = f"{insert_sql} ON DUPLICATE KEY UPDATE {upsert_set_sql}"
    inserted_count = 0
    updated_count = 0
    db_count = 0
    affected = 0

    def build_rows_from_source_df(source_df: pd.DataFrame, current_schema_rows: List[Dict[str, object]]) -> Tuple[List[Tuple[object, ...]], List[Dict[str, object]]]:
        current_schema_map = {str(item.get("name") or "").strip(): item for item in current_schema_rows}
        mapped_df = pd.DataFrame(index=source_df.index)

        for col in db_columns:
            column_meta = current_schema_map.get(col, {})
            source_field = mapping.get(col, "")
            if isinstance(source_field, str) and source_field.startswith("__LOGIC_"):
                mapped_df[col] = [get_import_fallback_value(column_meta)] * len(source_df.index)
            elif isinstance(source_field, str) and source_field.startswith("__FIXED__:"):
                mapped_df[col] = source_field.replace("__FIXED__:", "", 1)
            elif source_field in source_df.columns:
                mapped_df[col] = source_df[source_field] if table_name in RAW_IMPORT_TABLES else source_df[source_field].astype(str)
            else:
                mapped_df[col] = [get_import_fallback_value(column_meta)] * len(source_df.index)

        if table_name == "rstran":
            mapped_df = apply_rstran_logic(mapped_df)

        mapped_df = mapped_df.astype(object).where(pd.notna(mapped_df), None)

        if table_name not in RAW_IMPORT_TABLES:
            ensure_import_text_capacity(table_name, current_schema_rows, mapped_df)
            current_schema_rows = get_table_schema(table_name)
            current_schema_map = {str(item.get("name") or "").strip(): item for item in current_schema_rows}

        column_meta_sequence = [current_schema_map.get(col, {}) for col in db_columns]
        rows = [
            tuple(
                normalize_import_param(value, column_meta_sequence[index], table_name=table_name)
                for index, value in enumerate(row_values)
            )
            for row_values in mapped_df.loc[:, db_columns].itertuples(index=False, name=None)
        ]

        if rows and table_name not in RAW_IMPORT_TABLES:
            ensure_import_row_text_capacity(table_name, current_schema_rows, db_columns, rows)
            current_schema_rows = get_table_schema(table_name)

        return rows, current_schema_rows

    def execute_import_write(rows_batch: List[Tuple[object, ...]]) -> Tuple[int, int, int, int]:
        conn = get_conn()
        cur = conn.cursor()
        try:
            fast_upsert_mode = table_name in FAST_IMPORT_UPSERT_TABLES or len(rows_batch) >= FAST_IMPORT_SKIP_EXISTING_KEY_SCAN_ROW_THRESHOLD
            if table_name == "dd03l":
                write_batch_size = DD03L_FAST_WRITE_BATCH_SIZE
            else:
                write_batch_size = 2000 if fast_upsert_mode else 500
            key_tuples = [tuple(row[index] for index in key_indices) for row in rows_batch]
            distinct_key_count = len(set(key_tuples))

            if fast_upsert_mode:
                cur.execute(f"SELECT COUNT(*) FROM `{table_name}`")
                before_db_count = int(cur.fetchone()[0])
            else:
                before_db_count = None

            existing_keys: set[tuple] = set()
            if key_tuples and not fast_upsert_mode:
                chunk_size = 300
                key_match_sql = " AND ".join(f"`{k}` <=> %s" for k in key_fields)
                select_cols_sql = ", ".join(f"`{k}`" for k in key_fields)

                for start in range(0, len(key_tuples), chunk_size):
                    chunk = key_tuples[start:start + chunk_size]
                    where_sql = " OR ".join([f"({key_match_sql})" for _ in chunk])
                    params: list = []
                    for key_values in chunk:
                        params.extend(key_values)
                    cur.execute(
                        f"SELECT {select_cols_sql} FROM `{table_name}` WHERE {where_sql}",
                        tuple(params),
                    )
                    existing_keys.update(tuple(row) for row in cur.fetchall())

            local_updated_count = 0
            local_inserted_count = 0
            if fast_upsert_mode:
                # In fast mode we skip expensive existing-key scans for large imports.
                # Inserted rows are inferred from table row-count delta.
                pass
            else:
                seen_keys = set(existing_keys)
                for key_values in key_tuples:
                    if key_values in seen_keys:
                        local_updated_count += 1
                    else:
                        local_inserted_count += 1
                        seen_keys.add(key_values)

            if rows_batch:
                if table_name in RAW_IMPORT_TABLES:
                    for row in rows_batch:
                        cur.execute(upsert_sql, row)
                else:
                    for start in range(0, len(rows_batch), write_batch_size):
                        cur.executemany(upsert_sql, rows_batch[start:start + write_batch_size])

            cur.execute(f"SELECT COUNT(*) FROM `{table_name}`")
            local_db_count = int(cur.fetchone()[0])

            if fast_upsert_mode:
                local_inserted_count = max(0, local_db_count - int(before_db_count or 0))
                local_updated_count = max(0, distinct_key_count - local_inserted_count)

            conn.commit()
            local_affected = local_inserted_count + local_updated_count
            return local_inserted_count, local_updated_count, local_db_count, local_affected
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            raise
        finally:
            try:
                cur.close()
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass

    def raise_import_mysql_error(exc: mysql.connector.Error) -> None:
        if exc.errno == 1062:
            detail = "导入失败：检测到重复键值数据。已回滚到导入前数据，请去重后重新导入。"
        elif exc.errno == 1048:
            exc_text = str(exc)
            col_match = re.search(r"Column\s+'([^']+)'\s+cannot\s+be\s+null", exc_text, flags=re.IGNORECASE)
            col_name = col_match.group(1) if col_match else "未知字段"
            detail = f"导入失败：字段 {col_name} 不能为 NULL（当前数据库表结构限制）。请检查表结构或改为空字符串后重试。"
        elif exc.errno == 1366:
            detail = f"导入失败：字段值与数据库类型不兼容。数据库返回：{exc.msg}"
        elif exc.errno == 1406:
            detail = f"导入失败：字段长度超出数据库定义。数据库返回：{exc.msg}"
        else:
            detail = f"导入失败，已回滚到导入前数据。数据库返回：{exc.msg}"
        raise HTTPException(status_code=400, detail=detail) from exc

    filename = str(file.filename or "").strip() or "<unknown>"
    target_sheet = str(sheet_name or "").strip() or "首个Sheet/CSV"
    excel_count = 0
    chunk_seen = False
    current_schema_rows = schema_rows

    try:
        for source_df in iter_upload_dataframe_chunks(file, sheet_name or None, header_row_num, chunk_size=IMPORT_READ_CHUNK_SIZE):
            if source_df.empty:
                continue
            chunk_seen = True
            rows, current_schema_rows = build_rows_from_source_df(source_df, current_schema_rows)
            if not rows:
                continue

            excel_count += len(rows)

            try:
                chunk_inserted, chunk_updated, db_count, chunk_affected = execute_import_write(rows)
            except sqlite3.Error as exc:
                raise HTTPException(status_code=400, detail=f"导入失败，已回滚到导入前数据。SQLite 返回：{exc}") from exc
            except mysql.connector.Error as exc:
                if exc.errno == 1406:
                    exc_text = str(exc.msg or exc)
                    col_match = re.search(r"column '([^']+)'", exc_text, flags=re.IGNORECASE)
                    offending_column = str(col_match.group(1) if col_match else "").strip()
                    if offending_column:
                        current_schema_rows = get_table_schema(table_name)
                        ensure_import_row_text_capacity(table_name, current_schema_rows, db_columns, rows, only_columns={offending_column})
                        current_schema_rows = get_table_schema(table_name)
                        try:
                            chunk_inserted, chunk_updated, db_count, chunk_affected = execute_import_write(rows)
                        except mysql.connector.Error as retry_exc:
                            raise_import_mysql_error(retry_exc)
                    else:
                        raise_import_mysql_error(exc)
                else:
                    raise_import_mysql_error(exc)

            inserted_count += chunk_inserted
            updated_count += chunk_updated
            affected += chunk_affected
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"导入失败：标题行数第{header_row_num}行超出文件有效范围，请调整后重试。",
        ) from exc

    if not chunk_seen or excel_count == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                f"导入失败：文件 {filename} 在 {target_sheet}（标题行=第{header_row_num}行）未读取到可导入数据行。"
                "请确认文件包含表头且至少有 1 行数据，并检查是否选对了 Sheet。"
            ),
        )

    status = upsert_status(table_name, count_table_rows(table_name))
    return {
        "table_name": table_name,
        "affected_rows": int(affected),
        "inserted_rows": int(inserted_count),
        "updated_rows": int(updated_count),
        "excel_count": int(excel_count),
        "db_count": int(db_count),
        "last_update": status["last_update"],
        "last_count": status["last_count"],
        "message": "Import completed",
    }


@app.post("/api/import/clear-table")
def clear_import_table(table_name: str = Form(...)) -> Dict[str, str | int]:
    table_name = str(table_name or "").strip()
    if table_name not in {"rstran", "rstrant", "rstranrule", "rstranfield", "rstranstepcnst", "rstransteprout", "rsoadso", "rsoadsot", "rsds", "rsdst", "rsdssegfd", "rsdssegfdt", "rsksnew", "rsksnewt", "rsksfieldnew", "rsksfieldnewt", "dd03l", "dd02t", "dd03t", "dd04t", "rsdiobj", "rsdiobjt", PROG_CODE_TABLE}:
        raise HTTPException(status_code=400, detail="Only rstran, rstrant, rstranrule, rstranfield, rstranstepcnst, rstransteprout, rsoadso, rsoadsot, rsds, rsdst, rsdssegfd, rsdssegfdt, rsksnew, rsksnewt, rsksfieldnew, rsksfieldnewt, dd03l, dd02t, dd03t, dd04t, rsdiobj, rsdiobjt and prog_code support clear-table in current workflow")

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"DELETE FROM `{table_name}`")
        cur.execute(f"SELECT COUNT(*) FROM `{table_name}`")
        db_count = int(cur.fetchone()[0])
        conn.commit()
    except mysql.connector.Error as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail="删除失败：数据库执行异常") from exc
    finally:
        cur.close()
        conn.close()

    status = upsert_status(table_name, db_count)
    return {
        "table_name": table_name,
        "db_count": int(db_count),
        "last_update": status["last_update"],
        "last_count": status["last_count"],
        "message": "Clear completed",
    }


@app.post("/api/import/rebuild-rstran-mapping-rule")
def rebuild_rstran_mapping_rule() -> Dict[str, str | int]:
    try:
        result = rebuild_rstran_mapping_rule_table()
        total_rows = count_table_rows(RSTRAN_MAPPING_RULE_TABLE)
    except mysql.connector.Error as exc:
        raise HTTPException(status_code=500, detail=f"重建失败：数据库执行异常。{exc.msg}") from exc

    return {
        "table_name": RSTRAN_MAPPING_RULE_TABLE,
        "inserted_rows": int(result.get("inserted_rows") or 0),
        "tran_count": int(result.get("tran_count") or 0),
        "db_count": int(total_rows),
        "message": "RSTRAN mapping rule rebuild completed after truncate",
    }


@app.post("/api/import/rebuild-rstran-mapping-rule-full")
def rebuild_rstran_mapping_rule_full() -> Dict[str, str | int]:
    try:
        result = rebuild_rstran_mapping_rule_full_table()
        total_rows = count_table_rows(RSTRAN_MAPPING_RULE_FULL_TABLE)
    except mysql.connector.Error as exc:
        raise HTTPException(status_code=500, detail=f"重建失败：数据库执行异常。{exc.msg}") from exc

    return {
        "table_name": RSTRAN_MAPPING_RULE_FULL_TABLE,
        "inserted_rows": int(result.get("inserted_rows") or 0),
        "tran_count": int(result.get("tran_count") or 0),
        "mapped_rows": int(result.get("mapped_rows") or 0),
        "source_completed_rows": int(result.get("source_completed_rows") or 0),
        "target_completed_rows": int(result.get("target_completed_rows") or 0),
        "base_inserted_rows": int(result.get("base_inserted_rows") or 0),
        "base_tran_count": int(result.get("base_tran_count") or 0),
        "inventory_inserted_rows": int(result.get("inventory_inserted_rows") or 0),
        "inventory_object_count": int(result.get("inventory_object_count") or 0),
        "db_count": int(total_rows),
        "message": "RSTRAN full mapping rule rebuild completed after truncate",
    }


@app.post("/api/import/rebuild-bw-object-field-inventory")
def rebuild_bw_object_field_inventory() -> Dict[str, str | int]:
    try:
        result = rebuild_bw_object_field_inventory_table()
        total_rows = count_table_rows(BW_OBJECT_FIELD_INVENTORY_TABLE)
    except mysql.connector.Error as exc:
        raise HTTPException(status_code=500, detail=f"重建失败：数据库执行异常。{exc.msg}") from exc

    return {
        "table_name": BW_OBJECT_FIELD_INVENTORY_TABLE,
        "inserted_rows": int(result.get("inserted_rows") or 0),
        "object_count": int(result.get("object_count") or 0),
        "odso_count": int(result.get("odso_count") or 0),
        "adso_count": int(result.get("adso_count") or 0),
        "trcs_count": int(result.get("trcs_count") or 0),
        "iobj_count": int(result.get("iobj_count") or 0),
        "db_count": int(total_rows),
        "message": "BW object field inventory rebuild completed after truncate",
    }
